import json

import pytest
from openpyxl import Workbook, load_workbook

from bedding_order_parser.cli import build_parser as build_main_parser
from bedding_order_parser.dictionaries.loader import (
    DictionaryLoadError,
    bounded_nonempty_bounds,
    compute_sha256,
    load_dictionary_bundle,
)
from bedding_order_parser.dictionaries.models import (
    DictionaryBundle,
    DictionarySource,
    FabricRow,
    RuleRow,
    StyleRow,
)
from bedding_order_parser.dictionaries.writer import write_dictionary_preview
from bedding_order_parser.exceptions import OutputFileError


def _rules_workbook(path) -> None:
    workbook = Workbook()
    rules = workbook.active
    rules.title = "被套 提取规则"
    rules.append(
        [
            "字段名",
            "可能值（尽可能列举所有，\n有缺少的可以插入行补充）",
            "关键描述（尽可能列举可能的关键描述，如果有组合或者关联关系的也请详细列出）",
            "默认值规则",
            "补充说明",
        ]
    )
    rules.append(["产品名称", "被套", "duvet   cover", "没有默认", None])
    for row_index in range(3, 37):
        field_name = "面料" if row_index in {4, 5} else f"字段{row_index}"
        rules.append(
            [
                field_name,
                f"标准值 {row_index}",
                f"关键词\r\n{row_index}",
                " 默认  规则 ",
                "",
            ]
        )
    rules.merge_cells("A4:A5")

    fabrics = workbook.create_sheet("面料类价格表")
    fabrics.append(["面料品类", "面料", "颜色", "涤棉成份", "密度"])
    for row_index in range(2, 77):
        fabrics.append(
            [
                "贡缎 / 缎纹",
                f"贡缎/JC{row_index}S*JC40S/173*120/缎纹",
                "漂白",
                "100%棉",
                "T300",
            ]
        )
    workbook.save(path)


def _style_workbook(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(
        [
            "被套款式",
            "飞边 (Flange)",
            "系带 (Tie)",
            "拉链 (Zipper)",
            "是否口袋式 (Has Pocket)",
            "是否迎宾式 (Is Welcome Style)",
            "其他款式结构 (Other Structure)",
            "备注尺寸 (Dimensions)",
        ]
    )
    sheet.append(
        [
            "无飞边口袋无系带式",
            "无飞边",
            "无系带",
            "",
            "是",
            "否",
            "口袋",
            None,
        ]
    )
    for row_index in range(3, 107):
        sheet.append(
            [
                f"款式{row_index}",
                "单飞边",
                "有系带",
                "无拉链",
                "否",
                "否",
                "",
                f"{row_index}cm",
            ]
        )
    workbook.save(path)


@pytest.fixture()
def dictionary_paths(tmp_path):
    rules = tmp_path / "PI单提取规则.xlsx"
    styles = tmp_path / "款式表_structured.xlsx"
    _rules_workbook(rules)
    _style_workbook(styles)
    return rules, styles


def _load_test_bundle(rules, styles):
    return load_dictionary_bundle(
        rules,
        styles,
        expected_rules_sha256=compute_sha256(rules),
        expected_styles_sha256=compute_sha256(styles),
    )


def _minimal_bundle() -> DictionaryBundle:
    return DictionaryBundle(
        sources=[
            DictionarySource(
                file_name="PI单提取规则.xlsx",
                sha256="abc",
                sheet_name="被套 提取规则",
            )
        ],
        rules=[
            RuleRow(
                source_row=2,
                source_cells={"字段名": "A2"},
                field_name="产品名称",
                rule_description="duvet cover",
                standard_value="被套",
                notes="中文备注",
                raw_values={"字段名": "产品名称"},
            )
        ],
        fabrics=[
            FabricRow(
                source_row=2,
                fabric_family="贡缎 / 缎纹",
                fabric_standard="贡缎/JC60S*JC40S/173*120/缎纹",
                color_standard="漂白",
                composition_raw="100%棉",
                density="T300",
                raw_values={"面料品类": "贡缎 / 缎纹"},
            )
        ],
        styles=[
            StyleRow(
                source_row=2,
                standard_name="无飞边口袋无系带式",
                flange="无飞边",
                tie="无系带",
                zipper="",
                has_pocket="是",
                is_welcome_style="否",
                other_structure="口袋",
                dimensions="",
                raw_values={"被套款式": "无飞边口袋无系带式"},
            )
        ],
        summary={"rule_rows": 1, "fabric_rows": 1, "style_rows": 1},
    )


def test_correct_sha_loads_dictionary_bundle(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    bundle = _load_test_bundle(rules, styles)

    assert bundle.summary == {"rule_rows": 35, "fabric_rows": 75, "style_rows": 105}
    assert [source.sheet_name for source in bundle.sources] == [
        "被套 提取规则",
        "面料类价格表",
        "Sheet1",
    ]


def test_wrong_sha_fails_before_using_dictionary(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    with pytest.raises(DictionaryLoadError, match="SHA-256 mismatch"):
        load_dictionary_bundle(
            rules,
            styles,
            expected_rules_sha256="0" * 64,
            expected_styles_sha256=compute_sha256(styles),
        )


def test_missing_file_fails(dictionary_paths, tmp_path) -> None:
    _, styles = dictionary_paths

    with pytest.raises(DictionaryLoadError, match="does not exist"):
        load_dictionary_bundle(
            tmp_path / "missing.xlsx",
            styles,
            expected_rules_sha256="unused",
            expected_styles_sha256=compute_sha256(styles),
        )


def test_non_xlsx_file_fails(dictionary_paths, tmp_path) -> None:
    _, styles = dictionary_paths
    bad_file = tmp_path / "rules.txt"
    bad_file.write_text("not an xlsx", encoding="utf-8")

    with pytest.raises(DictionaryLoadError, match="must be .xlsx"):
        load_dictionary_bundle(
            bad_file,
            styles,
            expected_rules_sha256="unused",
            expected_styles_sha256=compute_sha256(styles),
        )


def test_merged_cells_inherit_top_left_value(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    bundle = _load_test_bundle(rules, styles)

    inherited = bundle.rules[3]
    assert inherited.source_row == 5
    assert inherited.field_name == "面料"
    assert inherited.source_cells["字段名"] == "A4"


def test_bounded_nonempty_range_ignores_far_cells(dictionary_paths) -> None:
    rules, _ = dictionary_paths
    workbook = load_workbook(rules, read_only=False, data_only=True)
    try:
        sheet = workbook["被套 提取规则"]
        sheet["Z100"] = "outside approved range"

        bounds = bounded_nonempty_bounds(
            sheet,
            min_row=1,
            max_row=36,
            min_column=1,
            max_column=5,
        )
    finally:
        workbook.close()

    assert bounds == {"min_row": 1, "max_row": 36, "min_column": 1, "max_column": 5}


def test_rule_row_fields_are_mapped_by_position(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    row = _load_test_bundle(rules, styles).rules[0]

    assert row.source_row == 2
    assert row.field_name == "产品名称"
    assert row.standard_value == "被套"
    assert row.rule_description == "duvet cover"
    assert row.default_rule == "没有默认"
    assert row.raw_values["可能值"] == "被套"


def test_fabric_row_fields_are_mapped_by_position(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    row = _load_test_bundle(rules, styles).fabrics[0]

    assert row.source_row == 2
    assert row.fabric_family == "贡缎 / 缎纹"
    assert row.fabric_standard.startswith("贡缎/JC2S")
    assert row.color_standard == "漂白"
    assert row.composition_raw == "100%棉"
    assert row.density == "T300"


def test_style_row_fields_are_mapped_by_position(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    row = _load_test_bundle(rules, styles).styles[0]

    assert row.source_row == 2
    assert row.standard_name == "无飞边口袋无系带式"
    assert row.flange == "无飞边"
    assert row.tie == "无系带"
    assert row.zipper == ""
    assert row.has_pocket == "是"
    assert row.is_welcome_style == "否"
    assert row.other_structure == "口袋"


def test_none_values_become_empty_strings(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    bundle = _load_test_bundle(rules, styles)

    assert bundle.rules[0].notes == ""
    assert bundle.styles[0].dimensions == ""


def test_dictionary_preview_json_keeps_chinese_unescaped(tmp_path) -> None:
    output = tmp_path / "preview.json"

    write_dictionary_preview(_minimal_bundle(), output)

    text = output.read_text(encoding="utf-8")
    assert "中文备注" in text
    assert "\\u4e2d" not in text
    payload = json.loads(text)
    assert payload["summary"]["rule_rows"] == 1


def test_dictionary_preview_writer_refuses_overwrite(tmp_path) -> None:
    output = tmp_path / "preview.json"
    write_dictionary_preview(_minimal_bundle(), output)

    with pytest.raises(OutputFileError, match="already exists"):
        write_dictionary_preview(_minimal_bundle(), output)


def test_dictionary_preview_writer_allows_explicit_overwrite(tmp_path) -> None:
    output = tmp_path / "preview.json"
    output.write_text("old", encoding="utf-8")

    write_dictionary_preview(_minimal_bundle(), output, overwrite=True)

    payload = json.loads(output.read_text(encoding="utf-8"))
    assert payload["summary"]["style_rows"] == 1


def test_atomic_write_failure_removes_temp_file(tmp_path, monkeypatch) -> None:
    output = tmp_path / "preview.json"

    def fail_replace(_source, _target):
        raise OSError("replace failed")

    monkeypatch.setattr(
        "bedding_order_parser.dictionaries.writer.os.replace",
        fail_replace,
    )

    with pytest.raises(OutputFileError):
        write_dictionary_preview(_minimal_bundle(), output)

    assert not output.exists()
    assert list(tmp_path.glob("*.tmp")) == []


def test_existing_parse_cli_command_is_still_available() -> None:
    parser = build_main_parser()

    assert "parse" in parser.format_help()


def test_bundle_json_order_matches_preview_contract(dictionary_paths) -> None:
    rules, styles = dictionary_paths

    payload = _load_test_bundle(rules, styles).to_json_dict()

    assert list(payload) == ["sources", "summary", "rules", "fabrics", "styles"]
