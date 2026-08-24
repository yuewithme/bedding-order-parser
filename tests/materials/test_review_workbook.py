from __future__ import annotations

import json
import sqlite3
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from bedding_order_parser.materials.loader import compute_sha256
from bedding_order_parser.materials.review_validator import validate_review_workbook
from bedding_order_parser.materials.review_workbook import (
    CANDIDATE_HEADERS,
    CANDIDATE_SHEET,
    INSTRUCTION_SHEET,
    MATERIAL_INDEX_SHEET,
    REVIEW_CONCLUSIONS,
    REVIEW_HEADERS,
    REVIEW_SHEET,
    build_review_workbook,
)


def write_store(path: Path) -> None:
    with sqlite3.connect(path) as connection:
        connection.execute(
            """
            CREATE TABLE materials (
                material_code TEXT PRIMARY KEY,
                material_name_raw TEXT NOT NULL,
                spec_raw TEXT NOT NULL,
                color_raw TEXT NOT NULL,
                fabric_raw TEXT NOT NULL,
                composition_raw TEXT NOT NULL,
                style_raw TEXT NOT NULL
            )
            """
        )
        connection.executemany(
            "INSERT INTO materials VALUES (?, ?, ?, ?, ?, ?, ?)",
            [
                ("MAT-1", "被套 A", "240*260cm", "漂白色", "贡缎", "C100", "无飞边"),
                ("MAT-2", "被套 B", "240*260cm", "漂白色", "贡缎", "C100", "无飞边"),
                ("MAT-3", "被套 C", "250*270cm", "浅灰色", "斜纹", "C95/T5", "口袋"),
            ],
        )


def candidate(code: str, rank: int = 1, status: str = "exact_match") -> dict[str, object]:
    fields = {
        "spec": {"candidate_value": "240*260cm", "status": status},
        "color": {"candidate_value": "漂白色", "status": "exact_match"},
        "fabric": {"candidate_value": "贡缎", "status": "equivalent_match"},
        "composition": {"candidate_value": "C100", "status": "exact_match"},
        "style": {"candidate_value": "无飞边", "status": "partial_match"},
        "label_method": {"candidate_value": "客标", "status": "no_match"},
        "size_type": {"candidate_value": "交货尺寸", "status": "missing_candidate"},
    }
    return {
        "rank": rank,
        "material_code": code,
        "prototype_match_score": 0.9 - rank / 100,
        "structured_score": 0.8,
        "vector_score": 0.7,
        "duplicate_group_size": 1,
        "ambiguous_duplicate_group": False,
        "duplicate_group": None,
        "fields": fields,
    }


def record(index: int, status: str, candidates: list[dict[str, object]] | None = None) -> dict[str, object]:
    return {
        "source_file": "sample.xlsx",
        "sheet": "PI",
        "行号": str(index),
        "result_json": "sample_gate2d.json",
        "parse_report_json": "sample_gate2d_parse_report.json",
        "query": {
            "product_category": "被套",
            "spec": "240*260cm",
            "color": "漂白色",
            "fabric": "贡缎/T300/C100",
            "composition": "C100",
            "style": "无飞边",
            "label_method": "客标",
            "size_type": "交货尺寸",
        },
        "retrieval": {
            "structured_candidates": 1,
            "vector_candidates": 3,
            "union_candidates": 3,
            "hard_conflict_removed": 0,
            "hard_conflicts_by_field": {},
            "post_filter_candidates": len(candidates or []),
        },
        "decision": {
            "status": status,
            "action": "manual_review",
            "top1_margin": 0.1 if status == "unique_best_candidate" else None,
            "comparable_field_count": 7 if candidates else 0,
            "reason": f"fixture {status}",
        },
        "candidates": candidates or [],
    }


def write_payloads(tmp_path: Path, records: list[dict[str, object]]) -> tuple[Path, Path, Path, Path]:
    output_root = tmp_path / "data" / "output"
    prototype = output_root / "material_match_prototype"
    prototype.mkdir(parents=True)
    candidates_path = prototype / "material_match_candidates.json"
    summary_path = prototype / "material_match_summary.json"
    decisions = Counter(row["decision"]["status"] for row in records)  # type: ignore[index]
    candidates_path.write_text(
        json.dumps({"records": records}, ensure_ascii=False),
        encoding="utf-8",
    )
    summary_path.write_text(
        json.dumps(
            {
                "summary": {
                    "order_records": len(records),
                    "records_with_candidates": sum(bool(row["candidates"]) for row in records),
                    "decision_statuses": dict(decisions),
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    formal_dir = output_root / "gate2d_validation" / "all_results"
    formal_dir.mkdir(parents=True)
    formal_dir.joinpath("sample_gate2d.json").write_text(
        json.dumps(
            [
                {
                    "行号": str(i),
                    "客户": "Fixture Customer",
                    "物料名称": "Fixture Customer 被套",
                }
                for i in range(1, len(records) + 1)
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    store = tmp_path / "material_master.sqlite3"
    write_store(store)
    workbook = tmp_path / "material_match_review.xlsx"
    return candidates_path, summary_path, store, workbook


def test_build_review_workbook_uses_one_row_per_order_and_top10(tmp_path) -> None:
    records = [record(i, "unique_best_candidate", [candidate("MAT-1")]) for i in range(1, 50)]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)

    result = build_review_workbook(candidates_path, summary_path, store, workbook)

    assert result.review_records == 49
    assert result.recommended_codes == 49
    wb = load_workbook(workbook)
    try:
        assert wb.sheetnames == [
            REVIEW_SHEET,
            CANDIDATE_SHEET,
            "重点问题",
            INSTRUCTION_SHEET,
            MATERIAL_INDEX_SHEET,
        ]
        assert wb[MATERIAL_INDEX_SHEET].sheet_state == "hidden"
        review = wb[REVIEW_SHEET]
        detail = wb[CANDIDATE_SHEET]
        assert [cell.value for cell in review[1]] == REVIEW_HEADERS
        assert [cell.value for cell in detail[1]] == CANDIDATE_HEADERS
        assert review.max_row == 50
        assert detail.max_row == 50
        assert review["E2"].value == "Fixture Customer"
        assert review["F2"].value == "Fixture Customer 被套"
        assert review["Q1"].comment is not None
        assert "原型分数" in review["Q1"].comment.text
        assert any(validation.type == "list" for validation in review.data_validations.dataValidation)
        assert any(validation.type == "custom" for validation in review.data_validations.dataValidation)
    finally:
        wb.close()


def test_candidate_status_labels_and_hidden_material_index(tmp_path) -> None:
    records = [record(1, "unique_best_candidate", [candidate("MAT-1", status="no_match")])]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)

    build_review_workbook(candidates_path, summary_path, store, workbook)

    wb = load_workbook(workbook)
    try:
        detail = wb[CANDIDATE_SHEET]
        assert detail["R2"].value == "不一致"
        assert detail["S2"].value == "完全一致"
        assert detail["T2"].value == "等价一致"
        assert detail["U2"].value == "完全一致"
        assert detail["V2"].value == "部分匹配"
        assert detail["W2"].value == "不一致"
        assert detail["X2"].value == "物料信息缺失"
        index = wb[MATERIAL_INDEX_SHEET]
        assert index.max_row == 4
        assert index["A2"].value == "MAT-1"
    finally:
        wb.close()


def test_validate_review_workbook_accepts_core_business_conclusions(tmp_path) -> None:
    records = [
        record(1, "unique_best_candidate", [candidate("MAT-1"), candidate("MAT-2", rank=2)]),
        record(2, "unique_best_candidate", [candidate("MAT-1"), candidate("MAT-2", rank=2)]),
        record(3, "no_candidate", []),
        record(4, "unique_best_candidate", [candidate("MAT-3")]),
    ]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)
    build_review_workbook(candidates_path, summary_path, store, workbook)
    wb = load_workbook(workbook)
    try:
        review = wb[REVIEW_SHEET]
        review["T2"] = "MAT-1"
        review["U2"] = "推荐编码正确"
        review["T3"] = "MAT-2"
        review["U3"] = "Top候选中其他编码正确"
        review["U4"] = "物料库不存在对应物料"
        review["U5"] = "订单字段解析错误"
        review["V5"] = "颜色"
        wb.save(workbook)
    finally:
        wb.close()

    before = compute_sha256(workbook)
    result = validate_review_workbook(workbook, store)

    assert result.ok
    assert result.checked_rows == 4
    assert result.sha256_before == before
    assert result.sha256_after == before


def test_validate_review_workbook_reports_invalid_codes_and_required_fields(tmp_path) -> None:
    records = [record(1, "unique_best_candidate", [candidate("MAT-1"), candidate("MAT-2", rank=2)])]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)
    build_review_workbook(candidates_path, summary_path, store, workbook)
    wb = load_workbook(workbook)
    try:
        review = wb[REVIEW_SHEET]
        review["T2"] = "MAT-X"
        review["U2"] = "推荐编码正确"
        wb.save(workbook)
    finally:
        wb.close()

    result = validate_review_workbook(workbook, store)

    messages = [issue.message for issue in result.errors]
    assert not result.ok
    assert any("不存在于SQLite" in message for message in messages)
    assert any("必须等于推荐物料编码" in message for message in messages)


def test_validate_review_workbook_rejects_other_code_outside_top10(tmp_path) -> None:
    records = [record(1, "unique_best_candidate", [candidate("MAT-1")])]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)
    build_review_workbook(candidates_path, summary_path, store, workbook)
    wb = load_workbook(workbook)
    try:
        review = wb[REVIEW_SHEET]
        review["T2"] = "MAT-3"
        review["U2"] = "Top候选中其他编码正确"
        wb.save(workbook)
    finally:
        wb.close()

    result = validate_review_workbook(workbook, store)

    assert any("不在该订单Top 10候选" in issue.message for issue in result.errors)


def test_validate_review_workbook_requires_fix_field_for_parse_error(tmp_path) -> None:
    records = [record(1, "unique_best_candidate", [candidate("MAT-1")])]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)
    build_review_workbook(candidates_path, summary_path, store, workbook)
    wb = load_workbook(workbook)
    try:
        review = wb[REVIEW_SHEET]
        review["U2"] = "订单字段解析错误"
        wb.save(workbook)
    finally:
        wb.close()

    result = validate_review_workbook(workbook, store)

    assert any("必须填写需要修正" in issue.message for issue in result.errors)


def test_validate_review_workbook_rejects_duplicate_and_missing_audit_ids(tmp_path) -> None:
    records = [
        record(1, "unique_best_candidate", [candidate("MAT-1")]),
        record(2, "unique_best_candidate", [candidate("MAT-1")]),
        record(3, "unique_best_candidate", [candidate("MAT-1")]),
    ]
    candidates_path, summary_path, store, workbook = write_payloads(tmp_path, records)
    build_review_workbook(candidates_path, summary_path, store, workbook)
    wb = load_workbook(workbook)
    try:
        review = wb[REVIEW_SHEET]
        review["A3"] = 1
        review.delete_rows(4)
        wb.save(workbook)
    finally:
        wb.close()

    result = validate_review_workbook(workbook, store)

    messages = [issue.message for issue in result.errors]
    assert any("重复" in message for message in messages)
    assert any("缺失" in message for message in messages)
