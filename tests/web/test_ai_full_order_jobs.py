from __future__ import annotations

from io import BytesIO
import json
import socket
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order import downstream
from bedding_order_parser.ai_full_order.downstream import MaterialMatchOutput, MaterialSelection
from bedding_order_parser.ai_full_order.fake_provider import (
    FakeFullOrderProvider,
    FakeV2CandidateProvider,
)
from bedding_order_parser.ai_full_order.orchestration import build_v2_extraction_units
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook
from bedding_order_parser.ai_full_order.reliability import TransientProviderError
from bedding_order_parser.ai_full_order.structure_manifest import (
    StructureManifestAdapterError,
)
from bedding_order_parser.ai_full_order.reliability_v2 import (
    V2ReliabilityStore,
    build_v2_cache_identity,
)
from bedding_order_parser.web.ai_full_order_service import AIEnhancedDependencies
from bedding_order_parser.web.services import JobService, WebJobError


class _DeferredExecutor:
    def submit(self, _function, *_args) -> None:
        return None

    def shutdown(self, **_kwargs) -> None:
        return None


class _FakeDictionaryValidator:
    def __init__(self) -> None:
        self.calls = 0

    def validate(self, records, evidence):
        self.calls += 1
        assert evidence
        return {
            "validation_version": "1.0",
            "mode": "validation_only",
            "status": "completed",
            "records": [{"行号": record.line_number} for record in records],
        }


class _FakeMaterialMatcher:
    def __init__(self) -> None:
        self.calls = 0

    def match(self, records, resolved):
        self.calls += 1
        selections = {
            item.source_record_id: MaterialSelection(item.source_record_id, "MAT-001", 0.75)
            for item in resolved
        }
        return MaterialMatchOutput(
            selections,
            {
                "mode": "manual_review_only",
                "record_count": len(records),
                "records": [{"行号": item.values["行号"], "candidates": []} for item in records],
            },
            {
                "mode": "manual_review_only",
                "record_count": len(records),
                "accuracy_statement": "相似分数不是准确率，候选只用于人工复核。",
            },
        )


@pytest.fixture(autouse=True)
def _block_network(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        socket,
        "create_connection",
        lambda *_args, **_kwargs: pytest.fail("network forbidden"),
    )


class _TransientSecondChunkProvider(FakeV2CandidateProvider):
    def __init__(self) -> None:
        super().__init__({"candidates": []})
        self.attempts = 0

    def extract_v2(self, request):
        self.attempts += 1
        if self.attempts == 2:
            self.extraction_call_count += 1
            raise TransientProviderError("injected transient failure")
        return super().extract_v2(request)


class _V2OnlyProvider(FakeV2CandidateProvider):
    def __init__(self, payload=None) -> None:
        super().__init__(payload or {"candidates": []})
        self.v1_extraction_call_count = 0

    def extract(self, request):
        self.v1_extraction_call_count += 1
        raise AssertionError("new ai_enhanced jobs must not call V1 extract")

    def extract_v2(self, request):
        return FakeV2CandidateProvider.extract(self, request)


class _SlowV2Provider(_V2OnlyProvider):
    def extract_v2(self, request):
        time.sleep(0.05)
        return super().extract_v2(request)


class _FailingLayoutProvider(_V2OnlyProvider):
    def resolve_structure(self, _manifest):
        self.structure_call_count += 1
        raise RuntimeError("synthetic layout provider failure")


class _SelectingAuxiliaryProvider(_V2OnlyProvider):
    def __init__(self) -> None:
        super().__init__({"candidates": []})
        self.http_attempt_count = 100
        self.usage_summary = {
            "input_tokens": 1000,
            "output_tokens": 500,
            "total_tokens": 1500,
        }

    def resolve_structure(self, manifest):
        self.structure_call_count += 1
        self.http_attempt_count += 1
        self._usage(10, 2)
        sheet = manifest["unresolved_sheets"][0]
        candidate = sheet["candidate_options"][0]
        return {
            "layout_contract_version": "2.0",
            "status": "resolved",
            "decisions": [{
                "sheet_id": sheet["sheet_id"],
                "role": "auxiliary",
                "candidate_id": candidate["candidate_id"],
                "reason": "auxiliary_non_order_content",
            }],
        }

    def extract_v2(self, request):
        result = super().extract_v2(request)
        self.http_attempt_count += 1
        self._usage(5, 1)
        return result

    def _usage(self, input_tokens, output_tokens):
        self.usage_summary["input_tokens"] += input_tokens
        self.usage_summary["output_tokens"] += output_tokens
        self.usage_summary["total_tokens"] += input_tokens + output_tokens


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    for index, (title, line) in enumerate((("PI-A", "1"), ("PI-B", "2"))):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = title
        sheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)"])
        sheet.append(["BUYER:", "", "", "", ""])
        sheet.append([f"Test Hotel {line}", "", "", "", "Contact Person: Aaron Lee"])
        sheet.append(["Delivery date:", "2026-09-30", "", "", ""])
        sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
        sheet.append([line, "Duvet Cover", "200*240", "100% cotton white", "12"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _single_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI-A"
    sheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)"])
    sheet.append(["BUYER:", "", "", "", ""])
    sheet.append(["Test Hotel 1", "", "", "", "Contact Person: Aaron Lee"])
    sheet.append(["Delivery date:", "2026-09-30", "", "", ""])
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet Cover", "200*240", "100% cotton white", "12"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _ambiguous_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet.append(["Synthetic quotation note"])
    sheet.append(["Please prepare a draft for internal review."])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _known_plus_auxiliary_workbook_bytes() -> bytes:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "Orders"
    orders.append(["No.", "Item", "Size", "Specification", "Qty"])
    for number in range(1, 4):
        orders.append([str(number), "Duvet", "200*240", "White cotton", str(number)])
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "Packing schedule"
    notes["A2"] = "Color references"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _job_units(service: JobService, job_id: str):
    job = service._read_job(job_id)
    input_path = service.jobs_root / job_id / "input" / job["file_name"]
    preprocessed = preprocess_workbook(input_path)
    return preprocessed, build_v2_extraction_units(preprocessed)


def _candidate(field_name: str, value: str, evidence_id: str, *, interpretation: str = "direct"):
    return {
        "field_name": field_name,
        "candidate_value": value,
        "evidence_references": [evidence_id],
        "interpretation": interpretation,
        "supporting_quote": "",
    }


def _service(tmp_path: Path, provider=None):
    provider = provider or _V2OnlyProvider()
    dictionary = _FakeDictionaryValidator()
    matcher = _FakeMaterialMatcher()
    service = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=_DeferredExecutor(),
        ai_enhanced_dependencies=AIEnhancedDependencies(
            provider=provider,
            dictionary_validator=dictionary,
            material_matcher=matcher,
        ),
    )
    return service, provider, dictionary, matcher


def test_new_jobs_only_accept_the_two_parse_modes_and_legacy_jobs_remain_read_only(tmp_path: Path) -> None:
    service, *_ = _service(tmp_path)
    standard = service.create_job("standard.xlsx", _workbook_bytes(), parse_mode="standard")
    enhanced = service.create_job("enhanced.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    with pytest.raises(WebJobError, match="仅支持"):
        service.create_job("bad.xlsx", _workbook_bytes(), parse_mode="allow_ai")

    assert standard["parse_mode"] == standard["effective_parse_mode"] == "standard"
    assert enhanced["parse_mode"] == enhanced["effective_parse_mode"] == "ai_enhanced"
    assert enhanced["requested_parse_mode"] == "ai_enhanced"
    assert enhanced["ai_contract_version"] == "2.0"
    assert enhanced["ai_contract_source"] == "new_job_v2"
    assert enhanced["source_identity"]["sha256"]
    assert enhanced["client_idempotency_identity"].startswith("ai-full-order-v2:")
    assert enhanced["business_identity"].startswith("desktop-job-v2:")
    with pytest.raises(WebJobError, match="创建后不可修改"):
        service._update_job(enhanced["id"], {"parse_mode": "standard"})
    legacy = service._read_job(standard["id"])
    for key in ("parse_mode", "effective_parse_mode", "parse_contract_version", "parse_mode_source", "fallback"):
        legacy.pop(key, None)
    service._write_job(legacy)
    before = (service.jobs_root / standard["id"] / "job.json").read_bytes()

    view = service.get_job(standard["id"])

    assert view["parse_mode"] == "standard"
    assert view["parse_mode_source"] == "legacy_default"
    assert view["parse_mode_label"] == "标准解析（历史任务）"
    assert (service.jobs_root / standard["id"] / "job.json").read_bytes() == before


def test_standard_dispatch_remains_on_existing_path(tmp_path: Path, monkeypatch) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("standard.xlsx", _workbook_bytes(), parse_mode="standard")
    calls: list[str] = []
    monkeypatch.setattr(service, "_run_standard_job", lambda _job_id: calls.append("standard"))
    monkeypatch.setattr(service, "_run_ai_enhanced_job", lambda _job_id: pytest.fail("standard must not enter AI chain"))

    service._run_job(job["id"])

    assert calls == ["standard"]


def test_ai_enhanced_runs_fake_chain_publishes_roles_and_skips_layout_ai(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("enhanced.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")

    service._run_job(job["id"])
    completed = service.get_job(job["id"])

    assert completed["status"] == "completed"
    assert completed["parse_mode"] == completed["effective_parse_mode"] == "ai_enhanced"
    assert completed["has_complete_five_results"] is True
    assert all(completed["artifact_roles"].values())
    assert completed["ai_execution"]["logical_call_count"] == 2
    assert completed["ai_execution"]["http_attempt_count"] == 0
    assert completed["ai_execution"]["has_publishable_result"] is True
    assert completed["ai_execution"]["stage_history"] == [
        "preprocessing",
        "structure_resolution",
        "python_shadow",
        "ai_extraction",
        "evidence_binding",
        "field_resolution",
        "cache_revalidation",
        "dictionary_validation",
        "material_matching",
        "publication",
        "completed",
    ]
    assert provider.structure_call_count == 0
    assert provider.v1_extraction_call_count == 0
    assert provider.extraction_call_count == 2
    assert provider.network_call_count == 0
    assert dictionary.calls == matcher.calls == 1
    assert not (service.jobs_root / job["id"] / "ai-advisory").exists()
    assert service.get_preview(job["id"], "official_result")[0]["物料编码"] == "MAT-001"
    assert service.artifact_path(job["id"], "material_summary").name == "material_match_summary.json"
    calls = provider.extraction_call_count
    service.get_job(job["id"])
    service.list_jobs()
    service.get_preview(job["id"], "official_result")
    service.artifact_path(job["id"], "parse_diagnostics")
    review = service.get_ai_review(job["id"])
    assert review["available"] is True
    assert review["summary"]["technical_ready"] is True
    assert completed["ai_review_summary"]["available"] is True
    assert provider.extraction_call_count == calls
    assert all(stage["status"] == "completed" for stage in completed["stages"])


@pytest.mark.parametrize(
    ("stage", "expected_label", "active_index"),
    [
        ("preprocessing", "正在读取订单", 0),
        ("structure_resolution", "正在确认表格结构", 0),
        ("python_shadow", "正在执行本地解析对照", 0),
        ("ai_extraction", "正在提取订单候选字段", 1),
        ("evidence_binding", "正在绑定来源证据", 1),
        ("field_resolution", "正在处理字段差异", 1),
        ("cache_revalidation", "正在复核本地缓存", 1),
        ("dictionary_validation", "正在验证业务字段", 2),
        ("material_matching", "正在匹配参考物料", 3),
        ("publication", "正在生成结果", 4),
    ],
)
def test_ai_v2_stage_updates_keep_the_five_user_steps_in_sync(
    tmp_path: Path,
    stage: str,
    expected_label: str,
    active_index: int,
) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("progress.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")

    service._set_ai_progress(job["id"], stage, completed=1, total=3, calls=1)
    view = service.get_job(job["id"])

    assert view["current_stage"] == expected_label
    assert view["ai_execution"]["stage"] == stage
    assert view["stages"] == [
        {
            "name": name,
            "status": (
                "completed"
                if index < active_index
                else "processing"
                if index == active_index
                else "waiting"
            ),
        }
        for index, name in enumerate(
            ("文件读取", "订单字段提取", "字典校验", "物料匹配", "结果生成")
        )
    ]


def test_standard_progress_stage_rows_remain_unchanged(tmp_path: Path) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("standard-progress.xlsx", _workbook_bytes(), parse_mode="standard")

    service._set_progress(
        job["id"],
        70,
        "正在物料匹配",
        completed_through=2,
        active_stage=3,
    )
    view = service.get_job(job["id"])

    assert view["current_stage"] == "正在物料匹配"
    assert [stage["status"] for stage in view["stages"]] == [
        "completed",
        "completed",
        "completed",
        "processing",
        "waiting",
    ]


def test_ai_current_entry_corruption_or_identity_mismatch_fails_safely(tmp_path: Path) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("enhanced.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    service._run_job(job["id"])
    root = service.jobs_root / job["id"] / "ai-bundle"
    current = root / "CURRENT"
    current.write_text("broken\n", encoding="utf-8")

    with pytest.raises(WebJobError, match="入口身份无效"):
        service.get_preview(job["id"], "official_result")
    assert service.get_job(job["id"])["has_complete_five_results"] is False

    cache_key = next((root / "bundles").iterdir()).name
    current.write_text(cache_key + "\n", encoding="utf-8")
    diagnostic = root / "bundles" / cache_key / "ai_full_order_parse_report.json"
    payload = service._read_json(diagnostic)
    payload["ai_enhanced"]["cache_key"] = "0" * 64
    diagnostic.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(WebJobError, match="身份不一致"):
        service.artifact_path(job["id"], "parse_diagnostics")


def test_failed_ai_waits_for_decision_and_retry_only_executes_remaining_chunk(tmp_path: Path) -> None:
    provider = _TransientSecondChunkProvider()
    service, _provider, dictionary, matcher = _service(tmp_path, provider)
    job = service.create_job("retry.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_TRANSIENT_FAILURE"
    assert paused["has_complete_five_results"] is False
    assert provider.extraction_call_count == 2
    assert dictionary.calls == matcher.calls == 0

    service.retry_missing_chunks(job["id"])
    completed = service.get_job(job["id"])

    assert completed["status"] == "completed"
    assert provider.extraction_call_count == 3
    assert dictionary.calls == matcher.calls == 1
    assert completed["ai_execution"]["resume_count"] == 1
    assert completed["ai_user_decision"]["action"] == "retry_missing_units"


def test_completed_ai_job_cannot_use_standard_reprocess(tmp_path: Path) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("fallback.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    service._run_job(job["id"])
    assert service.get_job(job["id"])["status"] == "completed"

    with pytest.raises(WebJobError, match="不处于可使用标准解析重新处理"):
        service.reprocess_ai_job_as_standard(
            job["id"], operation_id="completed-ai-reprocess-001"
        )
    assert not hasattr(service, "fallback_to_standard")


def test_ai_not_ready_blocks_new_job_and_token_budget_stays_awaiting(tmp_path: Path) -> None:
    no_ready = JobService(
        tmp_path / "not-ready",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=_DeferredExecutor(),
    )
    preflight = no_ready.ai_enhanced_preflight()
    assert preflight["unavailable_reason_code"] == "AI_PROVIDER_DISABLED"
    with pytest.raises(WebJobError, match="尚未在本机启用"):
        no_ready.create_job("not-ready.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    assert no_ready.list_jobs() == []

    provider = _V2OnlyProvider()
    dictionary = _FakeDictionaryValidator()
    matcher = _FakeMaterialMatcher()
    budgeted = JobService(
        tmp_path / "budgeted",
        store_path=tmp_path / "budget.sqlite3",
        index_dir=tmp_path / "budget-index",
        executor=_DeferredExecutor(),
        ai_enhanced_dependencies=AIEnhancedDependencies(
            provider=provider,
            dictionary_validator=dictionary,
            material_matcher=matcher,
            max_logical_calls=1,
        ),
    )
    budget_job = budgeted.create_job("budget.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    budgeted._run_job(budget_job["id"])
    paused = budgeted.get_job(budget_job["id"])
    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "TOKEN_BUDGET_EXCEEDED"
    assert provider.extraction_call_count == provider.structure_call_count == 0


def test_standard_and_ai_jobs_share_the_five_public_artifact_roles(tmp_path: Path) -> None:
    service, *_ = _service(tmp_path)
    job = service.create_job("standard-roles.xlsx", _workbook_bytes(), parse_mode="standard")
    root = service.jobs_root / job["id"]
    names = {
        "business": "formal.json",
        "diagnostic": "diagnostic.json",
        "validation": "validation.json",
        "matches": "candidates.json",
        "match_summary": "summary.json",
    }
    for name in names.values():
        (root / name).write_text("{}", encoding="utf-8")
    stored = service._read_job(job["id"])
    stored["artifacts"] = dict(names)
    service._write_job(stored)

    public = service.get_job(job["id"])

    assert public["artifact_roles"] == {
        "official_result": True,
        "parse_diagnostics": True,
        "dictionary_validation": True,
        "material_candidates": True,
        "material_summary": True,
    }
    assert service.artifact_path(job["id"], "official_result").name == "formal.json"
    assert service.get_ai_review(job["id"])["applicable"] is False


def test_legacy_ai_job_without_contract_version_stays_on_v1(tmp_path: Path) -> None:
    provider = FakeFullOrderProvider()
    dictionary = _FakeDictionaryValidator()
    matcher = _FakeMaterialMatcher()
    service = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=_DeferredExecutor(),
        ai_enhanced_dependencies=AIEnhancedDependencies(
            provider=provider,
            dictionary_validator=dictionary,
            material_matcher=matcher,
            contract_version="1.0",
        ),
    )
    job = service.create_job("legacy.xlsx", _single_workbook_bytes())
    stored = service._read_job(job["id"])
    stored["parse_mode"] = "ai_enhanced"
    stored["requested_parse_mode"] = "ai_enhanced"
    stored["effective_parse_mode"] = "ai_enhanced"
    stored.pop("ai_contract_version")
    stored.pop("ai_contract_source")
    service._write_job(stored)

    service._run_job(job["id"])
    completed = service.get_job(job["id"])

    assert completed["status"] == "completed"
    assert completed["ai_contract_version"] == "1.0"
    assert completed["ai_contract_source"] == "legacy_missing_version"
    assert completed["ai_contract_label"] == "Legacy V1"
    assert completed["ai_execution"]["contract_version"] == "1.0"
    assert provider.extraction_call_count == 1
    assert completed["has_complete_five_results"] is True
    legacy_review = service.get_ai_review(job["id"])
    assert legacy_review["available"] is False
    assert "Legacy V1" in legacy_review["compatibility_message"]


def test_ordinary_v2_candidate_issue_completes_with_isolation_diagnostics(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("isolated.xlsx", _single_workbook_bytes(), parse_mode="ai_enhanced")
    _preprocessed, units = _job_units(service, job["id"])
    evidence = next(item for item in units[0].evidence_catalog if item.original_text)
    provider.payload = {
        "candidates": [_candidate("包装方式", "invented packaging", evidence.evidence_id)]
    }

    service._run_job(job["id"])
    completed = service.get_job(job["id"])
    diagnostic = service.get_preview(job["id"], "parse_diagnostics")

    assert completed["status"] == "completed"
    assert completed["ai_execution"]["isolated_field_count"] == 1
    assert completed["ai_review_summary"]["content_issue_count"] == 1
    assert completed["ai_review_summary"]["review_required_count"] >= 1
    assert diagnostic["ai_enhanced"]["protocol"] == "v2"
    assert diagnostic["ai_enhanced"]["review_summary"]["content_issue_field_count"] == 1
    assert diagnostic["ai_enhanced"]["unit_states"] == ["validated_with_content_issue"]
    assert dictionary.calls == matcher.calls == 1


def test_high_review_conflict_completes_with_ai_value_and_five_results(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("conflict.xlsx", _single_workbook_bytes(), parse_mode="ai_enhanced")
    _preprocessed, units = _job_units(service, job["id"])
    evidence = next(item for item in units[0].evidence_catalog if "Aaron Lee" in item.original_text)
    provider.payload = {
        "candidates": [_candidate("客户", "Aaron Lee", evidence.evidence_id)]
    }

    service._run_job(job["id"])
    completed = service.get_job(job["id"])
    official = service.get_preview(job["id"], "official_result")

    assert completed["status"] == "completed"
    assert completed["has_complete_five_results"] is True
    assert completed["ai_review_summary"]["review_required_count"] >= 1
    assert completed["ai_review_summary"]["high_review_count"] >= 1
    assert official[0]["客户"] == "Aaron Lee"
    assert dictionary.calls == matcher.calls == 1


@pytest.mark.parametrize("kind", ["unknown", "cross_scope"])
def test_hard_evidence_failures_wait_without_partial_publication(tmp_path: Path, kind: str) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    content = _single_workbook_bytes() if kind == "unknown" else _workbook_bytes()
    job = service.create_job(f"{kind}.xlsx", content, parse_mode="ai_enhanced")
    _preprocessed, units = _job_units(service, job["id"])
    evidence_id = "unknown-evidence"
    if kind == "cross_scope":
        evidence_id = units[1].evidence_catalog[0].evidence_id
    provider.payload = {
        "candidates": [_candidate("颜色", "white", evidence_id)]
    }

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_CONTRACT_FAILED"
    assert paused["has_complete_five_results"] is False
    assert paused["ai_review_summary"]["available"] is False
    assert dictionary.calls == matcher.calls == 0
    assert not (service.jobs_root / job["id"] / "ai-bundle" / "CURRENT").exists()


def test_corrupt_cache_requires_explicit_retry_then_reexecutes_safely(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("corrupt.xlsx", _single_workbook_bytes(), parse_mode="ai_enhanced")
    _preprocessed, units = _job_units(service, job["id"])
    identity = build_v2_cache_identity(units)
    store = V2ReliabilityStore(service.jobs_root / job["id"] / "ai-runtime" / "reliability")
    path = store.state_path(identity.cache_key, units[0].extraction_unit_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("{broken", encoding="utf-8")

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_CACHE_CORRUPT"
    assert provider.extraction_call_count == 0
    completed = service.retry_missing_chunks(job["id"])
    assert completed["status"] == "completed"
    assert completed["ai_execution"]["resume_count"] == 1
    assert provider.extraction_call_count == 1
    assert dictionary.calls == matcher.calls == 1


def test_service_restart_resumes_only_unvalidated_v2_units(tmp_path: Path) -> None:
    first_provider = _TransientSecondChunkProvider()
    first, _provider, first_dictionary, first_matcher = _service(tmp_path, first_provider)
    job = first.create_job("restart.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")
    first._run_job(job["id"])
    assert first.get_job(job["id"])["status"] == "awaiting_user_decision"
    assert first_provider.extraction_call_count == 2
    assert first_dictionary.calls == first_matcher.calls == 0

    second_provider = _V2OnlyProvider()
    second_dictionary = _FakeDictionaryValidator()
    second_matcher = _FakeMaterialMatcher()
    second = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=_DeferredExecutor(),
        ai_enhanced_dependencies=AIEnhancedDependencies(
            provider=second_provider,
            dictionary_validator=second_dictionary,
            material_matcher=second_matcher,
        ),
    )

    resumed = second.retry_missing_chunks(job["id"])

    assert resumed["status"] == "completed"
    assert resumed["ai_execution"]["resume_count"] == 1
    assert second_provider.extraction_call_count == 1
    assert second_dictionary.calls == second_matcher.calls == 1


def test_service_restart_moves_stale_active_v2_job_to_recoverable_decision(tmp_path: Path) -> None:
    first, *_ = _service(tmp_path)
    job = first.create_job("stale.xlsx", _single_workbook_bytes(), parse_mode="ai_enhanced")
    first._set_ai_progress(job["id"], "ai_extraction", 0, 1, 0)
    stored = first._read_job(job["id"])
    stored["owner_pid"] = 999_999_999
    stored["owner_session_id"] = "previous-session"
    first._write_job(stored)

    second, provider, dictionary, matcher = _service(tmp_path)
    assert second.recover_stale_jobs() == 1
    recovered = second.get_job(job["id"])

    assert recovered["status"] == "awaiting_user_decision"
    assert recovered["ai_execution"]["safe_error_code"] == "AI_V2_INTERRUPTED"
    assert recovered["ai_user_decision"]["status"] == "pending"
    assert provider.extraction_call_count == 0
    assert dictionary.calls == matcher.calls == 0


def test_concurrent_same_job_does_not_duplicate_v2_unit_calls(tmp_path: Path) -> None:
    provider = _SlowV2Provider()
    service, _provider, dictionary, matcher = _service(tmp_path, provider)
    job = service.create_job("concurrent.xlsx", _workbook_bytes(), parse_mode="ai_enhanced")

    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = [pool.submit(service._run_ai_enhanced_job, job["id"]) for _ in range(2)]
        for future in futures:
            future.result(timeout=5)

    final = service.get_job(job["id"])
    assert final["status"] == "completed"
    assert provider.extraction_call_count == 2
    assert dictionary.calls == matcher.calls == 1


def test_ambiguous_structure_uses_only_fake_layout_and_never_extracts(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("ambiguous.xlsx", _ambiguous_workbook_bytes(), parse_mode="ai_enhanced")

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_STRUCTURE_UNRESOLVED"
    assert paused["ai_execution"]["layout_call_count"] == 1
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == 0
    assert dictionary.calls == matcher.calls == 0


def test_layout_provider_failure_uses_distinct_job_error_code(tmp_path: Path) -> None:
    service, provider, dictionary, matcher = _service(
        tmp_path, _FailingLayoutProvider()
    )
    job = service.create_job(
        "ambiguous-provider.xlsx", _ambiguous_workbook_bytes(), parse_mode="ai_enhanced"
    )

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_STRUCTURE_PROVIDER_FAILED"
    assert paused["ai_execution"]["layout_call_count"] == 1
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == 0
    assert dictionary.calls == matcher.calls == 0


def test_invalid_layout_manifest_stops_before_provider(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job(
        "ambiguous-manifest.xlsx", _ambiguous_workbook_bytes(), parse_mode="ai_enhanced"
    )
    monkeypatch.setattr(
        "bedding_order_parser.web.ai_full_order_service.build_structure_manifest",
        lambda *_args: (_ for _ in ()).throw(
            StructureManifestAdapterError("synthetic invalid manifest")
        ),
    )

    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_STRUCTURE_MANIFEST_INVALID"
    assert paused["ai_execution"]["layout_call_count"] == 0
    assert provider.structure_call_count == 0
    assert provider.extraction_call_count == 0
    assert dictionary.calls == matcher.calls == 0


def test_publication_failure_never_exposes_partial_bundle(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    service, provider, dictionary, matcher = _service(tmp_path)
    job = service.create_job("publish-failure.xlsx", _single_workbook_bytes(), parse_mode="ai_enhanced")
    original = downstream._write_json
    writes = 0

    def fail_third(path, payload, retries):
        nonlocal writes
        writes += 1
        if writes == 3:
            raise PermissionError("synthetic occupied file")
        return original(path, payload, retries)

    monkeypatch.setattr(downstream, "_write_json", fail_third)
    service._run_job(job["id"])
    paused = service.get_job(job["id"])

    assert paused["status"] == "awaiting_user_decision"
    assert paused["ai_execution"]["safe_error_code"] == "AI_V2_PUBLICATION_FAILED"
    assert paused["has_complete_five_results"] is False
    assert dictionary.calls == matcher.calls == 1
    assert provider.extraction_call_count == 1
    assert not (service.jobs_root / job["id"] / "ai-bundle" / "CURRENT").exists()


def test_safe_structure_summary_and_job_local_usage_are_persisted(tmp_path: Path) -> None:
    provider = _SelectingAuxiliaryProvider()
    service, _provider, dictionary, matcher = _service(tmp_path, provider)
    job = service.create_job(
        "known-plus-auxiliary.xlsx",
        _known_plus_auxiliary_workbook_bytes(),
        parse_mode="ai_enhanced",
    )

    service._run_job(job["id"])
    stored = service.get_job(job["id"])
    structure = stored["ai_execution"]["structure_resolution"]

    assert stored["status"] == "completed"
    assert stored["record_count"] == 3
    assert stored["has_complete_five_results"] is True
    assert structure["validation_status"] == "applied"
    assert structure["decisions"][0]["sheet_id"] == "s2"
    assert structure["decisions"][0]["role"] == "auxiliary"
    assert structure["decisions"][0]["candidate_id"].startswith(
        "layout-candidate:s2:"
    )
    assert structure["decisions"][0]["reason"] == "auxiliary_non_order_content"
    assert stored["ai_execution"]["token_summary"] == {
        "input_tokens": 25,
        "output_tokens": 5,
        "total_tokens": 30,
    }
    assert stored["ai_execution"]["http_attempt_count"] == 4
    assert provider.usage_summary["total_tokens"] == 1530
    assert dictionary.calls == matcher.calls == 1
