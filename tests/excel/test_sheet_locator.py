from openpyxl import Workbook

from bedding_order_parser.excel.sheet_locator import locate_pi_sheet


def test_sheet_locator_prefers_pi_update() -> None:
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("PI")
    workbook.create_sheet("PI-update")

    assert locate_pi_sheet(workbook).title == "PI-update"


def test_sheet_locator_falls_back_to_fuzzy_pi() -> None:
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("Customer PI Sheet")

    assert locate_pi_sheet(workbook).title == "Customer PI Sheet"
