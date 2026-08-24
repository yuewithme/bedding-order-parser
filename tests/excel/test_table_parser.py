from openpyxl import Workbook

from bedding_order_parser.excel.table_parser import parse_table


def test_table_parser_locates_header_and_numbered_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["PROFORMA INVOICE"])
    worksheet.append(["No.", "Item", "Size W*L", "Specification", "QTY"])
    worksheet.append(["1", "Duvet Cover", "200*240", "100% cotton after wash", "12"])
    worksheet.append(["", "Total:", "", "", "12"])

    table = parse_table(worksheet)

    assert table.header_index == 1
    assert table.headers[:5] == ["No.", "Item", "Size W*L", "Specification", "QTY"]
    assert len(table.data_rows) == 1
    assert table.data_rows[0].values["Item"] == "Duvet Cover"


def test_table_parser_combines_two_layer_merged_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["PROFORMA INVOICE"])
    worksheet.append(
        [
            "No.\n序号",
            "Linens Description描述",
            "",
            "",
            "Qty\n数量",
        ]
    )
    worksheet.append(
        [
            "",
            "Item\n品名",
            "After Wash Size（WxL)\n洗涤尺寸",
            "Specification\n工艺",
            "",
        ]
    )
    worksheet.append(["1", "Bedsheet Single", "180*270cm", "plain", "20"])
    worksheet.append(["2", "Duvet Cover Single", "180*270cm", "Bag with Hand Holes, no flange", "300"])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("B2:D2")
    worksheet.merge_cells("E2:E3")

    table = parse_table(worksheet)

    assert table.header_index == 1
    assert "Linens Description" in table.headers[1]
    assert "Item" in table.headers[1]
    assert "After Wash Size" in table.headers[2]
    assert "Specification" in table.headers[3]
    assert table.data_rows[1].values[table.headers[1]] == "Duvet Cover Single"


def test_table_parser_continues_after_nonnumeric_category_titles() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Items", "Description", "Picture", "Size", "Quantity"])
    worksheet.append(["Bath Linen", "", "", "", "", ""])
    worksheet.append(["1", "Bath Towel", "100% cotton", "", "70*140cm", "12"])
    worksheet.append(["2", "Hand Towel", "100% cotton", "", "35*75cm", "24"])
    worksheet.append(["Bed Linen King: 180*200cm Twin: 100*200cm", "", "", "", "", ""])
    worksheet.append(["25", "Duvet cover\nKing", "Bag style with 5cm hem.", "", "260*250cm", "36"])
    worksheet.append(["26", "Duvet cover\nTwin", "Bag style with 5cm hem.", "", "180*250cm", "12"])
    worksheet.append(["TOTAL VALUE - FOB", "", "", "", "", "36"])

    table = parse_table(worksheet)

    assert [row.raw_values[0] for row in table.data_rows] == ["1", "2", "25", "26"]
