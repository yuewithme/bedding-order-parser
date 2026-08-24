import json

from openpyxl import Workbook

from bedding_order_parser.excel.workbook_reader import compute_sha256
from bedding_order_parser.models.final_result import FINAL_FIELD_NAMES
from bedding_order_parser.pipeline.order_parser import parse_order


def test_parse_order_outputs_final_json_and_preserves_input_hash(tmp_path) -> None:
    input_path = tmp_path / "20251231 sample.xlsx"
    output_path = tmp_path / "result.json"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PI"
    worksheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)"])
    worksheet.append(["BUYER:", "", "", "", ""])
    worksheet.append(["Test Hotel", "", "", "", "Contact Person: Aaron Lee"])
    worksheet.append(["No.", "Item", "Size W*L", "Specification", "QTY"])
    worksheet.append(["1", "Duvet Cover", "200*240", "100% cotton white after wash", "12"])
    worksheet.append(["", "Total:", "", "", "12"])
    workbook.save(input_path)

    before_hash = compute_sha256(input_path)
    summary = parse_order(input_path, output_path)
    after_hash = compute_sha256(input_path)

    assert summary.record_count == 1
    assert summary.input_sha256_before == before_hash
    report_path = tmp_path / "result_parse_report.json"
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert summary.input_sha256_after == after_hash
    payload = json.loads(output_path.read_text(encoding="utf-8"))
    assert payload[0]["客户"] == "Test Hotel"
    assert payload[0]["币种"] == "美元"
    assert payload[0]["业务员"] == "Aaron Lee"
    assert payload[0]["规格"] == "240*200cm"
    assert payload[0]["表头备注"] == ""
    assert payload[0]["计划发货日期"] == ""
    assert list(payload[0]) == list(FINAL_FIELD_NAMES)
    assert payload[0]["物料编码"] == ""
    assert payload[0]["相似分数"] == 0.0
    assert isinstance(payload[0]["相似分数"], float)
    assert summary.report_path == report_path
    assert report["summary"]["record_count"] == 1
    assert list(report["records"][0]["fields"]) == list(FINAL_FIELD_NAMES)
    assert all(
        report["records"][0]["fields"][field]["value"] == payload[0][field]
        for field in FINAL_FIELD_NAMES
    )
    assert report["records"][0]["fields"]["物料编码"]["status"] == "not_implemented"
    assert report["records"][0]["fields"]["相似分数"]["status"] == "not_implemented"
    statuses = {
        field: diagnostic["status"]
        for field, diagnostic in report["records"][0]["fields"].items()
    }
    assert statuses["客户"] == "extracted"
    assert statuses["币种"] == "normalized"
    assert statuses["物料名称"] == "derived"
    assert statuses["规格"] == "normalized"
    assert statuses["颜色"] == "normalized"
    assert statuses["加标方式"] == "defaulted"
    assert statuses["数量"] == "extracted"
    assert statuses["计划发货日期"] == "source_not_provided"
    assert statuses["包装方式"] == "defaulted"
    assert "Address:" not in json.dumps(report, ensure_ascii=False)
    diagnostic_words = {
        "source_not_provided",
        "unrecognized",
        "ambiguous",
        "not_implemented",
    }
    assert not any(value in diagnostic_words for value in payload[0].values())


def test_parse_order_preserves_contract_with_two_layer_headers_and_remarks(tmp_path) -> None:
    input_path = tmp_path / "two-layer-pi.xlsx"
    output_path = tmp_path / "result.json"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PI"
    worksheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)", ""])
    worksheet.append(["BUYER:", "", "", "", "", ""])
    worksheet.append(["Test Hotel", "", "", "", "Contact Person: Aaron Lee", ""])
    worksheet.append(["No.\n序号", "Linens Description描述", "", "", "Qty\n数量", "Remarks"])
    worksheet.append(["", "Item\n品名", "After Wash Size（WxL)\n洗涤尺寸", "Specification\n工艺", "", "Remarks"])
    worksheet.append(
        [
            "1",
            "Duvet Cover King",
            "180*270cm",
            "300TC cotton after washed size, Bag with Hand Holes, no flange",
            "300",
            "Delivery Size",
        ]
    )
    worksheet.merge_cells("A4:A5")
    worksheet.merge_cells("B4:D4")
    worksheet.merge_cells("E4:E5")
    worksheet.merge_cells("F4:F5")
    workbook.save(input_path)

    summary = parse_order(input_path, output_path)
    payload = json.loads(output_path.read_text(encoding="utf-8"))

    assert summary.record_count == 1
    assert list(payload[0]) == list(FINAL_FIELD_NAMES)
    assert payload[0]["物料编码"] == ""
    assert payload[0]["相似分数"] == 0.0
    assert isinstance(payload[0]["相似分数"], float)
    assert payload[0]["尺寸类型"] == "交货尺寸"
    assert payload[0]["款式"] == "无飞边口袋无系带迎宾式"
