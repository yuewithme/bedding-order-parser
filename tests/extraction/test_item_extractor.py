from openpyxl import Workbook

from bedding_order_parser.excel.table_parser import parse_table
from bedding_order_parser.extraction.item_extractor import extract_raw_items, is_duvet_cover_row


def test_duvet_filter_keeps_covers_and_excludes_non_covers() -> None:
    assert is_duvet_cover_row("Duvet Cover 100% cotton")
    assert is_duvet_cover_row("Dubet cover TWIN")
    assert is_duvet_cover_row("duvet-cover king")
    assert is_duvet_cover_row("quilt cover white")
    assert not is_duvet_cover_row("pillow case 100% cotton")
    assert not is_duvet_cover_row("duvet insert inner")
    assert not is_duvet_cover_row("duvet insert with duvet cover shell")


def test_extract_raw_items_keeps_only_duvet_cover_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Size", "Specification", "QTY"])
    worksheet.append(["1", "Duvet Cover", "200*240", "white", "12"])
    worksheet.append(["2", "Pillow", "50*80", "white", "24"])

    table = parse_table(worksheet)
    items = extract_raw_items(table)

    assert [item.line_number for item in items] == ["1"]


def test_item_name_include_overrides_mattress_words_in_description() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Dimension(cm)", "Description", "UOM", "Total Qty"])
    worksheet.append(
        [
            "4",
            "Duvet Cover King",
            "245*280cm",
            "100% cotton sateen for TOP MATTRESS, bottom opening with hand holes",
            "pcs",
            "160",
        ]
    )
    worksheet.append(["5", "Mattress Protector King", "180*200cm", "water proof", "pcs", "135"])

    table = parse_table(worksheet)
    items = extract_raw_items(table)

    assert [item.line_number for item in items] == ["4"]
    assert items[0].quantity == "160"


def test_strong_non_cover_items_remain_excluded() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Size", "Specification", "QTY"])
    worksheet.append(["1", "Duvet Cover", "200*240", "white", "12"])
    worksheet.append(["2", "Mattress Protector", "180*200", "white", "24"])
    worksheet.append(["3", "Duvet Insert", "200*240", "down filling", "8"])
    worksheet.append(["4", "Duvet Inner", "200*240", "down filling", "9"])

    table = parse_table(worksheet)
    items = extract_raw_items(table)

    assert [item.line_number for item in items] == ["1"]


def test_inherits_adjacent_shared_description_with_duvet_cover_evidence() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Size", "Description", "QTY"])
    worksheet.append(
        [
            "1",
            "Bed Sheet",
            "240*300",
            "100% coton, 400 thread count, sateen, Simple bag style for duvet covers",
            "10",
        ]
    )
    worksheet.append(["2", "Duvet Cover", "205*260", "", "12"])

    items = extract_raw_items(parse_table(worksheet), worksheet)

    assert items[0].inherited_description is True
    assert "100% coton" in items[0].specification
    assert items[0].source_cells["specification"] == ("D2",)


def test_does_not_inherit_description_without_duvet_cover_evidence() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Size", "Description", "QTY"])
    worksheet.append(["1", "Bed Sheet", "240*300", "100% cotton sateen", "10"])
    worksheet.append(["2", "Duvet Cover", "205*260", "", "12"])

    items = extract_raw_items(parse_table(worksheet), worksheet)

    assert items[0].inherited_description is False
    assert items[0].specification == ""


def test_does_not_inherit_description_across_category_title() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["No.", "Item", "Size", "Description", "QTY"])
    worksheet.append(
        ["1", "Bed Sheet", "240*300", "Simple bag style for duvet covers", "10"]
    )
    worksheet.append(["Another Product Group"])
    worksheet.append(["2", "Duvet Cover", "205*260", "", "12"])

    items = extract_raw_items(parse_table(worksheet), worksheet)

    assert items[0].inherited_description is False
    assert items[0].specification == ""
