from __future__ import annotations

import socket
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order.contracts import FullOrderContractError, safe_contract_diagnostic
from bedding_order_parser.ai_full_order.fake_provider import (
    FakeFullOrderProvider,
    FakeV2CandidateProvider,
)
from bedding_order_parser.ai_full_order.preprocessing import PreprocessedWorkbook, preprocess_workbook
from bedding_order_parser.ai_full_order.provenance import (
    CandidateIssueCode,
    CandidateValidationStatus,
    adapt_verified_v1_record_to_v2_candidates,
    bind_v2_candidates,
)


def _write_book(path: Path, *, second_scope: bool = False) -> None:
    workbook = Workbook()
    _add_sheet(workbook.active, "PI-A", "Duvet Cover")
    if second_scope:
        _add_sheet(workbook.create_sheet(), "PI-B", "Pillow Case")
    workbook.save(path)


def _add_sheet(sheet, title: str, item: str) -> None:
    sheet.title = title
    sheet["A1"] = "SYNTHETIC ORDER"
    sheet.merge_cells("A1:D1")
    sheet.append(["No.", "Item", "Specification", "Qty"])
    sheet.append(["1", item, "White cotton", "12"])
    sheet.append(["2", f"{item} Alt", "Blue cotton", "8"])


def _write_merged_header_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MERGED"
    sheet["A2"] = "No."
    sheet["B2"] = "Items"
    sheet["E2"] = "Qty"
    sheet["B3"] = "Description"
    sheet["C3"] = "Size"
    sheet["D3"] = "Specification"
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:D2")
    sheet.merge_cells("E2:E3")
    sheet.append(["1", "Duvet Cover", "200*240", "White cotton", "12"])
    workbook.save(path)


def _preprocessed(tmp_path: Path, *, second_scope: bool = False) -> PreprocessedWorkbook:
    path = tmp_path / "synthetic-v2.xlsx"
    _write_book(path, second_scope=second_scope)
    return preprocess_workbook(path)


def _evidence_id(result: PreprocessedWorkbook, text: str, *, scope_id: str | None = None) -> str:
    for item in result.evidence_catalog:
        if item.normalized_text == text and (scope_id is None or item.scope_id == scope_id):
            return item.evidence_id
    raise AssertionError(f"missing synthetic evidence: {text}")


def _candidate(
    *,
    field_name: str = "物料名称",
    candidate_value: str = "Duvet Cover",
    references: list[str],
    interpretation: str = "direct",
    supporting_quote: str = "",
) -> dict[str, object]:
    return {
        "field_name": field_name,
        "candidate_value": candidate_value,
        "evidence_references": references,
        "interpretation": interpretation,
        "supporting_quote": supporting_quote,
    }


def test_v2_binder_recovers_local_evidence_and_deterministic_spans(tmp_path: Path) -> None:
    result = _preprocessed(tmp_path)
    target = result.records[0]
    item_id = _evidence_id(result, "Duvet Cover")
    specification_id = _evidence_id(result, "White cotton")
    title_id = _evidence_id(result, "Item")

    bound = bind_v2_candidates(
        {
            "candidates": [
                _candidate(
                    references=[item_id, specification_id],
                    supporting_quote="ignored direct-model quote",
                ),
                _candidate(
                    field_name="面料",
                    candidate_value="cotton",
                    references=[specification_id],
                    interpretation="semantic",
                    supporting_quote="White cotton",
                ),
                _candidate(
                    field_name="表头备注",
                    candidate_value="synthetic order note",
                    references=[title_id],
                    interpretation="source_summary",
                    supporting_quote="Item",
                ),
            ]
        },
        target=target,
        evidence_catalog=result.evidence_catalog,
    )

    assert [item.validation_status for item in bound] == [
        CandidateValidationStatus.BOUND,
        CandidateValidationStatus.BOUND,
        CandidateValidationStatus.BOUND,
    ]
    assert bound[0].quote_span is not None
    assert bound[0].quote_span.matched_text == "Duvet Cover"
    assert len(bound[0].evidence_snapshots) == 2
    assert bound[1].quote_span is not None
    assert bound[1].quote_span.start == 0
    assert title_id in target.evidence_ids
    assert bound[2].evidence_snapshots[0].cell_range == "B2"
    assert bound[2].to_dict()["target_record_local_id"] == target.record_local_id


def test_v2_binder_keeps_ordinary_quote_failures_as_candidate_issues(tmp_path: Path) -> None:
    result = _preprocessed(tmp_path)
    target = result.records[0]
    item_id = _evidence_id(result, "Duvet Cover")

    bound = bind_v2_candidates(
        {
            "candidates": [
                _candidate(candidate_value="Imagined Item", references=[item_id]),
                _candidate(
                    field_name="面料",
                    candidate_value="cotton",
                    references=[item_id],
                    interpretation="semantic",
                    supporting_quote="",
                ),
                _candidate(
                    field_name="规格",
                    candidate_value="derived",
                    references=[item_id],
                    interpretation="source_summary",
                    supporting_quote="not in source",
                ),
            ]
        },
        target=target,
        evidence_catalog=result.evidence_catalog,
    )

    assert [item.issue_code for item in bound] == [
        CandidateIssueCode.DIRECT_CANDIDATE_UNTRACEABLE,
        CandidateIssueCode.SUPPORTING_QUOTE_REQUIRED,
        CandidateIssueCode.SUPPORTING_QUOTE_UNTRACEABLE,
    ]
    assert all(item.validation_status is CandidateValidationStatus.ISSUE for item in bound)


def test_v2_binder_accepts_a_merged_header_anchor_in_target_scope(tmp_path: Path) -> None:
    path = tmp_path / "merged-header.xlsx"
    _write_merged_header_book(path)
    result = preprocess_workbook(path)
    target = result.records[0]
    header_id = _evidence_id(result, "Items")

    bound = bind_v2_candidates(
        {"candidates": [_candidate(field_name="表头备注", candidate_value="Items", references=[header_id])]},
        target=target,
        evidence_catalog=result.evidence_catalog,
    )

    assert header_id in target.evidence_ids
    assert bound[0].evidence_snapshots[0].cell_range == "B2"
    assert bound[0].validation_status is CandidateValidationStatus.BOUND


@pytest.mark.parametrize("kind", ["unknown", "cross_scope", "outside_target"])
def test_v2_binder_rejects_hard_evidence_boundary_errors(tmp_path: Path, kind: str) -> None:
    result = _preprocessed(tmp_path, second_scope=True)
    target = result.records[0]
    if kind == "unknown":
        reference = "forged:evidence"
        category = "evidence_id_missing"
    elif kind == "cross_scope":
        other_scope = next(record.scope_id for record in result.records if record.scope_id != target.scope_id)
        reference = _evidence_id(result, "Pillow Case", scope_id=other_scope)
        category = "evidence_cross_scope"
    else:
        own_scope_unrelated = next(
            item.evidence_id
            for item in result.evidence_catalog
            if item.scope_id == target.scope_id and item.evidence_id not in target.evidence_ids
        )
        reference = own_scope_unrelated
        category = "evidence_not_in_target"

    with pytest.raises(FullOrderContractError) as raised:
        bind_v2_candidates(
            {"candidates": [_candidate(references=[reference])]},
            target=target,
            evidence_catalog=result.evidence_catalog,
        )

    assert safe_contract_diagnostic(raised.value)["category"] == category


def test_v2_binder_rejects_an_invalid_local_target(tmp_path: Path) -> None:
    result = _preprocessed(tmp_path)
    invalid_target = replace(result.records[0], evidence_ids=())

    with pytest.raises(FullOrderContractError) as raised:
        bind_v2_candidates(
            {"candidates": []}, target=invalid_target, evidence_catalog=result.evidence_catalog
        )

    assert safe_contract_diagnostic(raised.value) == {
        "stage": "provenance_binding",
        "category": "target_identity_invalid",
        "path": "$.records[].record_local_id",
    }


def test_v1_adapter_only_maps_a_strictly_verified_v1_record(tmp_path: Path) -> None:
    result = _preprocessed(tmp_path)
    request = result.to_request_dict()
    output = FakeFullOrderProvider().extract(request)

    bound = adapt_verified_v1_record_to_v2_candidates(
        output,
        request=request,
        target=result.records[0],
        evidence_catalog=result.evidence_catalog,
    )

    assert len(bound) == 1
    assert bound[0].field_name == "物料名称"
    assert bound[0].interpretation == "direct"
    output["records"][0]["fields"]["物料名称"]["value"] = "invented"
    output["records"][0]["fields"]["物料名称"]["original_value"] = "invented"
    with pytest.raises(FullOrderContractError, match="not traceable"):
        adapt_verified_v1_record_to_v2_candidates(
            output,
            request=request,
            target=result.records[0],
            evidence_catalog=result.evidence_catalog,
        )


def test_fake_v2_provider_is_offline_and_returns_a_copy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    result = _preprocessed(tmp_path)
    item_id = _evidence_id(result, "Duvet Cover")
    provider = FakeV2CandidateProvider({"candidates": [_candidate(references=[item_id])]})

    monkeypatch.setattr(
        socket,
        "create_connection",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("network forbidden")),
    )
    output = provider.extract({"synthetic": "request"})
    output["candidates"].clear()

    assert provider.extraction_call_count == 1
    assert provider.network_call_count == 0
    assert provider.payload["candidates"]
