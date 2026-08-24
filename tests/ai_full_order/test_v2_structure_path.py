from __future__ import annotations

from dataclasses import replace
from io import BytesIO
import json
from pathlib import Path
import socket

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from bedding_order_parser.ai_full_order.downstream import (
    MaterialMatchOutput,
    MaterialSelection,
)
from bedding_order_parser.ai_full_order.fake_provider import FakeV2CandidateProvider
from bedding_order_parser.ai_full_order.orchestration import (
    build_v2_extraction_units,
    run_v2_offline_orchestration,
)
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook
from bedding_order_parser.ai_full_order.standard_geometry import StandardSheetGeometry
from bedding_order_parser.ai_full_order.python_shadow import (
    build_deterministic_python_shadow,
)
from bedding_order_parser.ai_full_order.structure_manifest import (
    STRUCTURE_MANIFEST_VERSION,
    build_structure_manifest,
)
from bedding_order_parser.ai_full_order.volcengine_ark import (
    FULL_ORDER_LAYOUT_FUNCTION,
    FULL_ORDER_V2_EXTRACTION_FUNCTION,
    VolcengineArkFullOrderProvider,
)
from bedding_order_parser.excel.table_parser import parse_table
from bedding_order_parser.extraction.item_extractor import extract_raw_items
from bedding_order_parser.llm.settings import LLMSettings
from bedding_order_parser.llm.transport import JSONHTTPRequest, JSONHTTPResponse
from bedding_order_parser.web import ai_full_order_service as service_module
from bedding_order_parser.web.ai_full_order_service import (
    AIEnhancedDependencies,
    AIEnhancedJobPause,
    run_ai_enhanced_v2_job,
)


@pytest.fixture(autouse=True)
def _block_network(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        socket,
        "create_connection",
        lambda *_args, **_kwargs: pytest.fail("network forbidden"),
    )


class _Dictionary:
    def __init__(self) -> None:
        self.calls = 0

    def validate(self, records, evidence):
        self.calls += 1
        assert evidence
        return {
            "validation_version": "1.0",
            "mode": "validation_only",
            "status": "completed",
            "records": [{"行号": record.line_number} for record in records],
        }


class _Matcher:
    def __init__(self) -> None:
        self.calls = 0

    def match(self, records, resolved):
        self.calls += 1
        return MaterialMatchOutput(
            selections={
                item.source_record_id: MaterialSelection(
                    item.source_record_id, "SYNTHETIC-MAT", 0.75
                )
                for item in resolved
            },
            candidates_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "records": [
                    {"行号": item.values["行号"], "candidates": []}
                    for item in records
                ],
            },
            summary_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "accuracy_statement": "相似分数不是准确率，候选只用于人工复核。",
            },
        )


class _RecordingTransport:
    def __init__(self) -> None:
        self.requests: list[JSONHTTPRequest] = []

    def send(self, request: JSONHTTPRequest) -> JSONHTTPResponse:
        self.requests.append(request)
        sent = json.loads(request.body)
        context = json.loads(sent["input"][1]["content"][0]["text"])
        function_name = sent["tool_choice"]["name"]
        if function_name == FULL_ORDER_V2_EXTRACTION_FUNCTION:
            body = {
                "id": "synthetic-extraction-response",
                "status": "completed",
                "output": [{
                    "type": "function_call",
                    "name": FULL_ORDER_V2_EXTRACTION_FUNCTION,
                    "arguments": json.dumps({"candidates": []}),
                }],
            }
            return JSONHTTPResponse(200, {}, json.dumps(body).encode("utf-8"), 1)
        decisions = []
        for sheet in context["unresolved_sheets"]:
            candidate = sheet["candidate_options"][0]
            decisions.append(
                {
                    "sheet_id": sheet["sheet_id"],
                    "role": candidate["role"],
                    "candidate_id": candidate["candidate_id"],
                    "reason": (
                        "selected_local_order_candidate"
                        if candidate["role"] == "order"
                        else "auxiliary_non_order_content"
                    ),
                }
            )
        body = {
            "id": "synthetic-layout-response",
            "status": "completed",
            "output": [
                {
                    "type": "function_call",
                    "name": FULL_ORDER_LAYOUT_FUNCTION,
                    "arguments": json.dumps(
                        {
                            "layout_contract_version": "2.0",
                            "status": "resolved",
                            "decisions": decisions,
                        }
                    ),
                }
            ],
        }
        return JSONHTTPResponse(200, {}, json.dumps(body).encode("utf-8"), 1)


class _FailingStructureProvider(FakeV2CandidateProvider):
    def resolve_structure(self, _manifest):
        self.structure_call_count += 1
        raise RuntimeError("synthetic provider failure")


def _dependencies(provider):
    return AIEnhancedDependencies(
        provider=provider,
        dictionary_validator=_Dictionary(),
        material_matcher=_Matcher(),
        provider_name=str(getattr(provider, "provider_name", "fake_provider")),
        model_name=str(getattr(provider, "model_name", "offline-test")),
    )


def _write_d3d1_shape(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic Structure"
    preamble = [
        "PROFORMA INVOICE",
        "Unit Price (USD)",
        "BUYER:",
        "Synthetic Hotel",
        "Delivery date: 2026-12-31",
        "SELLER: Synthetic Supplier",
        "Sales Person: Synthetic Agent",
        "Synthetic note 2",
        "Synthetic note 3",
        "Synthetic note 4",
        "Synthetic note 5",
        "Synthetic note 6",
    ]
    for row, text in enumerate(preamble, start=1):
        sheet.cell(row=row, column=1, value=text)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    sheet.append(
        ["No.", "Item", "Size", "Specification", "Qty", "Remark", "Unit", "Code"]
    )
    for number in range(1, 61):
        item = "Duvet Cover" if number <= 3 else "Pillow Case"
        sheet.append(
            [number, item, "200*240", "Synthetic white cotton", "12", "", "PCS", ""]
        )

    sheet.cell(row=74, column=1, value="TOTAL")
    sheet.cell(row=76, column=1, value="Synthetic auxiliary schedule")
    for row in range(78, 84):
        sheet.cell(row=row, column=1, value=row - 77)
        sheet.cell(row=row, column=2, value=f"Synthetic schedule item {row - 77}")
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    sheet.cell(row=85, column=1, value="Synthetic secondary schedule")
    for row in range(86, 92):
        sheet.cell(row=row, column=1, value=row - 85)
        sheet.cell(row=row, column=2, value=f"Synthetic check item {row - 85}")
        if row <= 87:
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    sheet["K95"] = "=1+1"
    sheet["XFD95"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(path)


def _write_second_table(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic Ambiguous"
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet Cover", "200*240", "Synthetic white cotton", "12"])
    sheet.append(["TOTAL", "", "", "", "12"])
    sheet.append([])
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet Cover", "180*220", "Synthetic blue cotton", "6"])
    workbook.save(path)


def _standard_record_rows(path: Path) -> tuple[int, ...]:
    workbook = load_workbook(path, data_only=True, keep_links=False)
    try:
        table = parse_table(workbook.active)
        return tuple(
            item.excel_row_number for item in extract_raw_items(table, workbook.active)
        )
    finally:
        workbook.close()


def test_standard_geometry_reduces_d3d1_shape_to_three_unique_v2_units(
    tmp_path: Path,
) -> None:
    path = tmp_path / "synthetic-structure.xlsx"
    _write_d3d1_shape(path)

    raw = load_workbook(path, data_only=False, keep_links=False)
    try:
        assert raw.active.max_row == 95
        assert raw.active.max_column == 16_384
        assert len(raw.active.merged_cells.ranges) == 20
    finally:
        raw.close()
    assert _standard_record_rows(path) == (14, 15, 16)

    preprocessed = preprocess_workbook(path)
    units = build_v2_extraction_units(preprocessed)
    evidence = {
        item.evidence_id: item for unit in units for item in unit.evidence_catalog
    }
    shadow = build_deterministic_python_shadow(
        path,
        preprocessed,
        target_records=[unit.target for unit in units],
        evidence_catalog=tuple(evidence.values()),
    )

    assert preprocessed.sheets[0].used_range == "A1:K95"
    assert preprocessed.structure_status == "locally_resolved"
    assert preprocessed.structure_diagnostics.to_dict() == {
        "standard_parsed_record_count": 60,
        "standard_selected_record_count": 3,
        "heuristic_record_count": 60,
        "aligned_record_count": 3,
        "auxiliary_numbered_row_count": 12,
        "standard_aligned_sheet_count": 1,
        "heuristic_sheet_count": 0,
        "possible_secondary_table_count": 0,
        "evidence_mapping_failure_count": 0,
    }
    assert len(preprocessed.records) == len(units) == len(shadow) == 3
    assert [unit.target.source_row for unit in units] == [14, 15, 16]
    assert len({unit.extraction_unit_id for unit in units}) == 3
    assert len({unit.target.source_record_id for unit in units}) == 3
    assert all(unit.target.scope_id == "s1:scope-1" for unit in units)
    assert all(
        len(unit.target.evidence_ids) == len(set(unit.target.evidence_ids))
        for unit in units
    )


def test_d3d1_shape_runs_three_fake_extractions_without_layout(tmp_path: Path) -> None:
    path = tmp_path / "synthetic-structure.xlsx"
    _write_d3d1_shape(path)
    provider = FakeV2CandidateProvider({"candidates": []})
    dependencies = _dependencies(provider)

    result = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime",
        publish_root=tmp_path / "publish",
        dependencies=dependencies,
        client_idempotency_key="synthetic-client",
        business_key="synthetic-business",
        on_stage=lambda *_args: None,
    )

    assert result.structure_status == "locally_resolved"
    assert result.total_chunks == 3
    assert provider.structure_call_count == 0
    assert provider.extraction_call_count == 3
    assert provider.network_call_count == 0
    assert dependencies.dictionary_validator.calls == 1
    assert dependencies.material_matcher.calls == 1


def test_explicit_second_order_table_stays_ambiguous(tmp_path: Path) -> None:
    path = tmp_path / "second-table.xlsx"
    _write_second_table(path)

    preprocessed = preprocess_workbook(path)

    assert preprocessed.structure_status == "ambiguous"
    assert preprocessed.structure_diagnostics.possible_secondary_table_count == 1
    assert len(preprocessed.blocks) == 2

    provider = FakeV2CandidateProvider({"candidates": []})
    result = run_v2_offline_orchestration(preprocessed, provider, ())
    assert result.batch.reasons == ("structure_unresolved",)
    assert result.extraction_units == result.outcomes == ()
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == provider.network_call_count == 0


def test_unmappable_standard_coordinates_stay_ambiguous(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "unmappable.xlsx"
    _write_d3d1_shape(path)
    from bedding_order_parser.ai_full_order import preprocessing

    monkeypatch.setattr(
        preprocessing,
        "derive_standard_sheet_geometry",
        lambda _sheet: StandardSheetGeometry(
            stable=True,
            header_rows=(13,),
            record_rows=(999,),
            parsed_record_count=1,
            reason_code="synthetic_unmappable",
        ),
    )

    result = preprocess_workbook(path)

    assert result.structure_status == "ambiguous"
    assert result.structure_diagnostics.evidence_mapping_failure_count == 1


def test_corrupt_local_identity_fails_before_provider(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "synthetic-structure.xlsx"
    _write_d3d1_shape(path)
    real = service_module.preprocess_workbook(path)
    bad_record = replace(real.records[0], source_row=999)
    bad = replace(
        real,
        records=(bad_record, *real.records[1:]),
        structure_status="ambiguous",
    )
    monkeypatch.setattr(service_module, "preprocess_workbook", lambda _path: bad)
    provider = FakeV2CandidateProvider({"candidates": []})

    with pytest.raises(AIEnhancedJobPause) as raised:
        run_ai_enhanced_v2_job(
            path,
            runtime_root=tmp_path / "runtime",
            publish_root=tmp_path / "publish",
            dependencies=_dependencies(provider),
            client_idempotency_key="bad-mapping",
            business_key="bad-mapping",
            on_stage=lambda *_args: None,
        )

    assert raised.value.code == "AI_V2_STRUCTURE_MANIFEST_INVALID"
    assert provider.structure_call_count == provider.network_call_count == 0


def test_formal_provider_receives_versioned_context_and_continues_after_safe_apply(
    tmp_path: Path,
) -> None:
    path = tmp_path / "second-table.xlsx"
    _write_second_table(path)
    preprocessed = preprocess_workbook(path)
    manifest = build_structure_manifest(
        preprocessed,
        __import__(
            "bedding_order_parser.ai_full_order.orchestration",
            fromlist=["build_chunk_manifest"],
        ).build_chunk_manifest(preprocessed),
    )
    assert manifest["manifest_version"] == STRUCTURE_MANIFEST_VERSION
    assert manifest["unresolved_sheets"]
    assert manifest["unresolved_sheets"][0]["candidate_options"]

    transport = _RecordingTransport()
    provider = VolcengineArkFullOrderProvider(
        LLMSettings(
            enabled=True,
            provider="volcengine_ark",
            model="offline-layout",
            base_url="https://ark.invalid/api/v3",
            api_key="offline-test-key",
            max_retries=0,
        ),
        transport=transport,
        sleep=lambda _seconds: None,
    )

    result = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime",
        publish_root=tmp_path / "publish",
        dependencies=_dependencies(provider),
        client_idempotency_key="layout-provider",
        business_key="layout-provider",
        on_stage=lambda *_args: None,
    )

    assert result.layout_call_count == 1
    assert result.http_attempt_count == 3
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == 2
    assert len(transport.requests) == 3
    sent = json.loads(transport.requests[0].body)
    prompt_manifest = json.loads(sent["input"][1]["content"][0]["text"])
    assert prompt_manifest["unresolved_sheets"]
    assert prompt_manifest["known_chunks"] == []
    assert sent["store"] is False
    assert "stream" not in sent
    assert sent["tools"][0]["name"] == FULL_ORDER_LAYOUT_FUNCTION
    assert sent["tools"][0]["strict"] is True


def test_structure_provider_failure_has_a_distinct_safe_code(tmp_path: Path) -> None:
    path = tmp_path / "second-table.xlsx"
    _write_second_table(path)
    provider = _FailingStructureProvider({"candidates": []})

    with pytest.raises(AIEnhancedJobPause) as raised:
        run_ai_enhanced_v2_job(
            path,
            runtime_root=tmp_path / "runtime",
            publish_root=tmp_path / "publish",
            dependencies=_dependencies(provider),
            client_idempotency_key="provider-failure",
            business_key="provider-failure",
            on_stage=lambda *_args: None,
        )

    assert raised.value.code == "AI_V2_STRUCTURE_PROVIDER_FAILED"
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == provider.network_call_count == 0
