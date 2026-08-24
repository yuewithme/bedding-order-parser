from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from bedding_order_parser.ai_full_order.fake_provider import FakeFullOrderProvider
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook


def write_clear_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet["A1"] = "PROFORMA INVOICE"
    sheet.merge_cells("A1:E1")
    sheet.append(["No.", "Items", "", "", "Qty"])
    sheet.append(["", "Description", "Size", "Specification", ""])
    sheet.append(["1", "Duvet Cover", "200*240", "White cotton", "12", None, None, "=1+1"])
    sheet.append(["2", "Hidden Item", "", "", "5"])
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:D2")
    sheet.merge_cells("E2:E3")
    sheet.row_dimensions[5].hidden = True
    sheet.column_dimensions["F"].hidden = True
    sheet["F4"] = "DO NOT SEND"
    sheet["Z100"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    hidden = workbook.create_sheet("Hidden Notes")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Private note"
    workbook.save(path)


def write_ambiguous_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Unstructured free-form sheet"
    workbook.save(path)


def test_preprocesses_dual_view_sparse_cells_merges_hidden_content_and_local_line(tmp_path: Path) -> None:
    path = tmp_path / "clear.xlsx"
    write_clear_workbook(path)
    provider = FakeFullOrderProvider()

    result = preprocess_workbook(path, structure_resolver=provider)

    assert result.structure_status == "locally_resolved"
    assert result.structure_resolution_requested is False
    assert provider.structure_call_count == 0
    assert len(result.blocks) == 1
    assert len(result.records) == 1
    assert result.records[0].local_line_number == "s1:4"
    assert result.sheets[0].used_range == "A1:H5"
    assert result.sheets[0].hidden_rows == (5,)
    assert result.sheets[0].hidden_columns == ("F",)
    assert result.sheets[1].included is False
    assert result.sheets[1].visibility == "hidden"
    assert result.sheets[1].cells == ()

    cells = {cell.cell_id: cell for cell in result.sheets[0].cells}
    assert cells["s1!C2"].merged_anchor == "s1!B2"
    assert cells["s1!H4"].formula_text == "=1+1"
    assert "s1!F4" not in cells
    assert all("Hidden Item" not in item.original_text for item in result.evidence_catalog)
    assert all("DO NOT SEND" not in item.original_text for item in result.evidence_catalog)
    request = result.to_request_dict()
    assert request["record_count"] == 1
    assert "line_number" not in request["records"][0]


def test_ambiguous_structure_only_then_requests_the_offline_resolver(tmp_path: Path) -> None:
    path = tmp_path / "ambiguous.xlsx"
    write_ambiguous_workbook(path)
    provider = FakeFullOrderProvider()

    result = preprocess_workbook(path, structure_resolver=provider)

    assert result.structure_status == "ambiguous"
    assert result.structure_resolution_requested is True
    assert provider.structure_call_count == 1
    assert provider.network_call_count == 0
