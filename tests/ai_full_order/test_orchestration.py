from __future__ import annotations

import copy
import socket
from dataclasses import replace
from pathlib import Path

from openpyxl import Workbook

from bedding_order_parser.ai_full_order.contracts import AI_BUSINESS_FIELD_NAMES
from bedding_order_parser.ai_full_order.fake_provider import FakeFullOrderProvider
from bedding_order_parser.ai_full_order.orchestration import (
    BatchStatus,
    aggregate_batch,
    build_chunk_manifest,
    formal_line_number_from_request,
    run_offline_orchestration,
)
from bedding_order_parser.ai_full_order.preprocessing import PreprocessedWorkbook, preprocess_workbook
from bedding_order_parser.ai_full_order.resolution import (
    FieldResolutionError,
    PythonFieldCandidate,
    PythonShadowRecord,
    adapt_python_shadow_records,
)
from bedding_order_parser.excel.table_parser import parse_table
from bedding_order_parser.extraction.item_extractor import extract_raw_items


def write_single_order_book(path: Path, *, line_number: str = "1") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet.append(["Title"])
    sheet.append([""])
    sheet.append(["No.", "Item", "Specification", "Qty"])
    sheet.append([line_number, "Duvet Cover", "White cotton", "12"])
    workbook.save(path)


def write_two_scope_book(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    _add_sheet(first, "PI-A", "1")
    second = workbook.create_sheet("PI-B")
    _add_sheet(second, "PI-B", "2")
    workbook.save(path)


def write_ambiguous_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Unstructured note"
    workbook.save(path)


def _add_sheet(sheet, title: str, line_number: str) -> None:
    sheet.title = title
    sheet.append(["No.", "Item", "Specification", "Qty"])
    sheet.append([line_number, "Duvet Cover", "White cotton", "12"])


def shadow_for(preprocessed: PreprocessedWorkbook) -> tuple[PythonShadowRecord, ...]:
    formal_records = []
    request = preprocessed.to_request_dict()
    evidence = request["evidence_catalog"]
    for record in request["records"]:
        formal_records.append(
            {
                **{field: "" for field in AI_BUSINESS_FIELD_NAMES},
                "行号": formal_line_number_from_request(record, evidence),
            }
        )
    return adapt_python_shadow_records(preprocessed, formal_records)


def direct_shadow_for(
    preprocessed: PreprocessedWorkbook,
    field_name: str,
    value: str,
) -> tuple[PythonShadowRecord, ...]:
    shadow = list(shadow_for(preprocessed))
    evidence_id = _first_record_evidence(preprocessed)
    fields = dict(shadow[0].fields)
    fields[field_name] = PythonFieldCandidate(
        value=value,
        evidence_ids=(evidence_id,),
        status="extracted",
    )
    shadow[0] = replace(shadow[0], fields=fields)
    return tuple(shadow)


def _first_record_evidence(preprocessed: PreprocessedWorkbook) -> str:
    record = preprocessed.records[0]
    return next(eid for eid in record.evidence_ids if ":B" in eid)


def test_locally_resolved_chunks_skip_structure_recognition_and_stay_offline(
    tmp_path: Path,
    monkeypatch,
) -> None:
    path = tmp_path / "single.xlsx"
    write_single_order_book(path)
    preprocessed = preprocess_workbook(path)
    provider = FakeFullOrderProvider()

    def explode(*_args, **_kwargs):
        raise AssertionError("network access is forbidden")

    monkeypatch.setattr(socket, "create_connection", explode)
    result = run_offline_orchestration(preprocessed, provider, shadow_for(preprocessed))

    assert result.batch.ready_for_downstream
    assert result.structure_recognition_calls == 0
    assert provider.structure_call_count == 0
    assert provider.extraction_call_count == 1
    assert result.network_calls == 0
    assert all("物料编码" not in record.business_fields() for record in result.batch.records)
    assert all("相似分数" not in record.business_fields() for record in result.batch.records)


def test_ambiguous_structure_calls_fake_structure_recognition_once(tmp_path: Path) -> None:
    path = tmp_path / "ambiguous.xlsx"
    write_ambiguous_book(path)
    preprocessed = preprocess_workbook(path)
    provider = FakeFullOrderProvider()

    result = run_offline_orchestration(preprocessed, provider, ())

    assert preprocessed.structure_status == "ambiguous"
    assert result.structure_recognition_calls == 1
    assert provider.extraction_call_count == 0
    assert provider.network_call_count == 0


def test_chunk_ids_are_stable(tmp_path: Path) -> None:
    path = tmp_path / "stable.xlsx"
    write_two_scope_book(path)
    first = preprocess_workbook(path)
    second = preprocess_workbook(path)

    assert [item.chunk_id for item in build_chunk_manifest(first)] == [
        item.chunk_id for item in build_chunk_manifest(second)
    ]


def test_reversed_chunk_execution_does_not_change_final_record_order(tmp_path: Path) -> None:
    path = tmp_path / "two-scope.xlsx"
    write_two_scope_book(path)
    preprocessed = preprocess_workbook(path)
    provider = FakeFullOrderProvider()
    manifest = build_chunk_manifest(preprocessed)
    reversed_ids = [item.chunk_id for item in reversed(manifest)]

    result = run_offline_orchestration(
        preprocessed,
        provider,
        shadow_for(preprocessed),
        chunk_order=reversed_ids,
    )

    assert result.batch.ready_for_downstream
    assert [record.line_number for record in result.batch.records] == ["1", "2"]
    assert [outcome.manifest.chunk_id for outcome in result.chunk_outcomes] == reversed_ids


def test_duplicate_record_identity_is_isolated(tmp_path: Path) -> None:
    path = tmp_path / "duplicates.xlsx"
    write_two_scope_book(path)
    preprocessed = preprocess_workbook(path)
    duplicate_records = list(preprocessed.records)
    duplicate_records[1] = replace(
        duplicate_records[1],
        source_record_id=duplicate_records[0].source_record_id,
    )
    duplicated = replace(preprocessed, records=tuple(duplicate_records))

    result = run_offline_orchestration(
        duplicated,
        FakeFullOrderProvider(),
        shadow_for(duplicated),
    )

    assert result.batch.status is BatchStatus.ISOLATED
    assert "duplicate_record_identity" in result.batch.reasons


def test_missing_chunk_cannot_be_ready(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"
    write_two_scope_book(path)
    preprocessed = preprocess_workbook(path)
    result = run_offline_orchestration(
        preprocessed,
        FakeFullOrderProvider(),
        shadow_for(preprocessed),
    )

    partial = aggregate_batch(result.manifest, result.chunk_outcomes[:1])

    assert partial.status is BatchStatus.ISOLATED
    assert "missing_chunks" in partial.reasons
    assert "record_count_mismatch" in partial.reasons


def test_cross_scope_and_forged_evidence_stay_isolated(tmp_path: Path) -> None:
    path = tmp_path / "cross.xlsx"
    write_two_scope_book(path)
    preprocessed = preprocess_workbook(path)
    cross = run_offline_orchestration(
        preprocessed,
        FakeFullOrderProvider("cross_scope"),
        shadow_for(preprocessed),
    )

    assert cross.batch.status is BatchStatus.ISOLATED
    assert "schema_or_evidence_failure" in cross.batch.reasons

    forged = run_offline_orchestration(
        preprocessed,
        FakeFullOrderProvider("forged_cell"),
        shadow_for(preprocessed),
    )
    assert forged.batch.status is BatchStatus.ISOLATED
    assert "schema_or_evidence_failure" in forged.batch.reasons


def test_formal_line_number_matches_standard_mode_semantics(tmp_path: Path) -> None:
    path = tmp_path / "line-number.xlsx"
    write_single_order_book(path, line_number="25")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet.append(["Title"])
    sheet.append([""])
    sheet.append(["No.", "Item", "Specification", "Qty"])
    sheet.append(["25", "Duvet Cover", "White cotton", "12"])

    standard_items = extract_raw_items(parse_table(sheet), sheet)
    preprocessed = preprocess_workbook(path)
    request = preprocessed.to_request_dict()

    assert standard_items[0].line_number == "25"
    assert preprocessed.records[0].source_row == 4
    assert preprocessed.records[0].local_line_number == "s1:4"
    assert formal_line_number_from_request(request["records"][0], request["evidence_catalog"]) == "25"


def test_shadow_identity_mismatch_is_isolated(tmp_path: Path) -> None:
    path = tmp_path / "mismatch.xlsx"
    write_single_order_book(path)
    preprocessed = preprocess_workbook(path)
    shadow = list(shadow_for(preprocessed))
    shadow[0] = replace(shadow[0], source_record_id="wrong")

    result = run_offline_orchestration(preprocessed, FakeFullOrderProvider(), tuple(shadow))

    assert result.batch.status is BatchStatus.ISOLATED
    assert "schema_or_evidence_failure" in result.batch.reasons


def test_record_count_mismatch_is_rejected_before_resolution(tmp_path: Path) -> None:
    path = tmp_path / "count.xlsx"
    write_single_order_book(path)
    preprocessed = preprocess_workbook(path)
    request = preprocessed.to_request_dict()
    output = FakeFullOrderProvider().extract(request)
    output["record_count"] = 99

    try:
        from bedding_order_parser.ai_full_order.resolution import resolve_records

        resolve_records(output, request=request, python_shadow=shadow_for(preprocessed))
    except FieldResolutionError as exc:
        assert "record_count" in str(exc)
    else:
        raise AssertionError("record_count mismatch should fail")
