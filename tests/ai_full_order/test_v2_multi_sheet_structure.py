from __future__ import annotations

import copy
import json
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order.downstream import (
    MaterialMatchOutput,
    MaterialSelection,
)
from bedding_order_parser.ai_full_order.fake_provider import FakeV2CandidateProvider
from bedding_order_parser.ai_full_order.orchestration import build_chunk_manifest
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook
from bedding_order_parser.ai_full_order.structure_manifest import (
    MAX_STRUCTURE_EXCERPT_CHARS,
    MAX_STRUCTURE_EXCERPTS,
    STRUCTURE_CONTEXT_VERSION,
    StructureManifestAdapterError,
    build_structure_manifest,
    provider_structure_payload,
)
from bedding_order_parser.ai_full_order.structure_resolution import (
    LAYOUT_CONTRACT_VERSION,
    StructureDecisionValidationError,
    apply_structure_decision,
    validate_layout_output_shape,
)
from bedding_order_parser.web.ai_full_order_service import (
    AIEnhancedDependencies,
    AIEnhancedJobPause,
    run_ai_enhanced_v2_job,
)


class _Dictionary:
    def __init__(self) -> None:
        self.calls = 0

    def validate(self, records, _evidence):
        self.calls += 1
        return {
            "validation_version": "1.0",
            "mode": "validation_only",
            "status": "completed",
            "records": [{"行号": item.line_number} for item in records],
        }


class _Matcher:
    def __init__(self) -> None:
        self.calls = 0

    def match(self, records, resolved):
        self.calls += 1
        selections = {
            item.source_record_id: MaterialSelection(item.source_record_id, "", 0.0)
            for item in resolved
        }
        return MaterialMatchOutput(
            selections=selections,
            candidates_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "records": [
                    {"行号": item.values["行号"], "candidates": []}
                    for item in records
                ],
            },
            summary_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "accuracy_statement": "相似分数不是准确率，候选只用于人工复核。",
            },
        )


class _UsageProvider(FakeV2CandidateProvider):
    def __init__(self, layout_payload):
        super().__init__({"candidates": []}, layout_payload=layout_payload)
        self.http_attempt_count = 100
        self.usage_summary = {
            "input_tokens": 1000,
            "output_tokens": 500,
            "total_tokens": 1500,
        }

    def resolve_structure(self, manifest):
        result = super().resolve_structure(manifest)
        self.http_attempt_count += 1
        self._add_usage(10, 2)
        return result

    def extract_v2(self, request):
        result = super().extract_v2(request)
        self.http_attempt_count += 1
        self._add_usage(5, 1)
        return result

    def _add_usage(self, input_tokens: int, output_tokens: int) -> None:
        self.usage_summary["input_tokens"] += input_tokens
        self.usage_summary["output_tokens"] += output_tokens
        self.usage_summary["total_tokens"] += input_tokens + output_tokens


def _dependencies(provider):
    dictionary = _Dictionary()
    matcher = _Matcher()
    return (
        AIEnhancedDependencies(
            provider=provider,
            dictionary_validator=dictionary,
            material_matcher=matcher,
        ),
        dictionary,
        matcher,
    )


def _write_known_plus_auxiliary(path: Path) -> None:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "Orders"
    orders.append(["No.", "Item", "Size", "Specification", "Qty"])
    for number in range(1, 4):
        orders.append([str(number), "Duvet", "200*240", "White cotton", str(number)])
    auxiliary = workbook.create_sheet("Notes")
    auxiliary["A1"] = "Packing schedule"
    auxiliary["A2"] = "Color references"
    auxiliary["B2"] = "White"
    auxiliary["A3"] = "Bank account"
    auxiliary["B3"] = "sensitive-synthetic-value"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "must-not-leave-local"
    workbook.save(path)


def _write_known_plus_second_order_candidate(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Orders A"
    first.append(["No.", "Item", "Size", "Specification", "Qty"])
    first.append(["1", "Duvet", "200*240", "White cotton", "2"])
    second = workbook.create_sheet("Orders B")
    second.append(["No.", "Item", "Size", "Specification", "Qty"])
    second.append(["1", "Pillow", "50*80", "Blue cotton", "4"])
    second.append(["TOTAL", "", "", "", "4"])
    second.append([])
    second.append(["No.", "Item", "Size", "Specification", "Qty"])
    second.append(["1", "Sheet", "180*200", "White cotton", "1"])
    workbook.save(path)


def _context(path: Path):
    preprocessed = preprocess_workbook(path)
    manifest = build_structure_manifest(preprocessed, build_chunk_manifest(preprocessed))
    return preprocessed, manifest


def _decision(manifest, *, role: str, candidate_index: int = 0):
    sheet = manifest["unresolved_sheets"][0]
    candidate_id = ""
    reason = "insufficient_structure"
    status = "ambiguous"
    if role != "unresolved":
        candidate = sheet["candidate_options"][candidate_index]
        candidate_id = candidate["candidate_id"]
        reason = (
            "selected_local_order_candidate"
            if role == "order"
            else "auxiliary_non_order_content"
        )
        status = "resolved"
    return {
        "layout_contract_version": LAYOUT_CONTRACT_VERSION,
        "status": status,
        "decisions": [{
            "sheet_id": sheet["sheet_id"],
            "role": role,
            "candidate_id": candidate_id,
            "reason": reason,
        }],
    }


def test_known_and_unresolved_sheets_coexist_in_bounded_safe_context(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    preprocessed, manifest = _context(path)
    payload = provider_structure_payload(manifest)

    assert preprocessed.structure_status == "ambiguous"
    assert [item["sheet_id"] for item in payload["known_chunks"]] == ["s1"]
    assert [item["sheet_id"] for item in payload["unresolved_sheets"]] == ["s2"]
    assert payload["unresolved_sheets"][0]["candidate_options"][0]["role"] == "auxiliary"
    assert payload["structure_context_version"] == STRUCTURE_CONTEXT_VERSION
    serialized = json.dumps(payload, ensure_ascii=False)
    assert "must-not-leave-local" not in serialized
    assert "sensitive-synthetic-value" not in serialized
    assert "C:\\" not in serialized
    assert "Authorization" not in serialized
    assert "api_key" not in serialized
    excerpts = payload["unresolved_sheets"][0]["excerpts"]
    assert len(excerpts) <= MAX_STRUCTURE_EXCERPTS
    assert all(len(item["text"]) <= MAX_STRUCTURE_EXCERPT_CHARS for item in excerpts)


def test_auxiliary_decision_keeps_three_known_records_and_runs_full_chain(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    provider = FakeV2CandidateProvider(
        {"candidates": []}, layout_payload=_decision(manifest, role="auxiliary")
    )
    dependencies, dictionary, matcher = _dependencies(provider)

    result = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime",
        publish_root=tmp_path / "publish",
        dependencies=dependencies,
        client_idempotency_key="aux-client",
        business_key="aux-business",
        on_stage=lambda *_args: None,
    )

    assert result.total_chunks == 3
    assert result.structure_summary["validation_status"] == "applied"
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == 3
    assert dictionary.calls == matcher.calls == 1
    assert len(result.bundle.paths) == 5


def test_unresolved_decision_stops_before_extract_and_downstream(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    provider = FakeV2CandidateProvider(
        {"candidates": []}, layout_payload=_decision(manifest, role="unresolved")
    )
    dependencies, dictionary, matcher = _dependencies(provider)

    with pytest.raises(AIEnhancedJobPause) as raised:
        run_ai_enhanced_v2_job(
            path,
            runtime_root=tmp_path / "runtime",
            publish_root=tmp_path / "publish",
            dependencies=dependencies,
            client_idempotency_key="unresolved-client",
            business_key="unresolved-business",
            on_stage=lambda *_args: None,
        )

    assert raised.value.code == "AI_V2_STRUCTURE_UNRESOLVED"
    assert raised.value.structure_summary["validation_status"] == "unresolved"
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == dictionary.calls == matcher.calls == 0
    assert not (tmp_path / "publish").exists()


def test_second_order_sheet_requires_and_applies_local_order_candidate(tmp_path: Path) -> None:
    path = tmp_path / "two-orders.xlsx"
    _write_known_plus_second_order_candidate(path)
    preprocessed, manifest = _context(path)
    sheet = manifest["unresolved_sheets"][0]
    assert sheet["sheet_id"] == "s2"
    assert sheet["candidate_options"][0]["role"] == "order"
    applied = apply_structure_decision(
        preprocessed, manifest, _decision(manifest, role="order")
    )

    assert applied.resolved is True
    assert {record.sheet_id for record in applied.preprocessed.records} == {"s1", "s2"}
    assert any(block.sheet_id == "s2" for block in applied.preprocessed.blocks)


@pytest.mark.parametrize(
    "mutation",
    [
        "unknown_sheet",
        "unknown_candidate",
        "cross_sheet_candidate",
        "duplicate_decision",
        "missing_decision",
        "invalid_role",
        "invalid_reason",
        "extra_field",
    ],
)
def test_invalid_layout_decisions_are_hard_rejected(tmp_path: Path, mutation: str) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    preprocessed, manifest = _context(path)
    output = _decision(manifest, role="auxiliary")
    if mutation == "unknown_sheet":
        output["decisions"][0]["sheet_id"] = "s999"
    elif mutation == "unknown_candidate":
        output["decisions"][0]["candidate_id"] = "layout-candidate:s2:unknown"
    elif mutation == "cross_sheet_candidate":
        candidate = preprocessed.layout_candidates[0]
        preprocessed = replace(
            preprocessed,
            layout_candidates=(replace(candidate, sheet_id="s1"),),
        )
    elif mutation == "duplicate_decision":
        output["decisions"].append(copy.deepcopy(output["decisions"][0]))
    elif mutation == "missing_decision":
        output["decisions"] = []
    elif mutation == "invalid_role":
        output["decisions"][0]["role"] = "ignored"
    elif mutation == "invalid_reason":
        output["decisions"][0]["reason"] = "model_says_so"
    elif mutation == "extra_field":
        output["decisions"][0]["explanation"] = "free text"

    with pytest.raises(StructureDecisionValidationError):
        apply_structure_decision(preprocessed, manifest, output)


def test_layout_output_schema_rejects_old_status_only_contract() -> None:
    with pytest.raises(StructureDecisionValidationError):
        validate_layout_output_shape({"status": "resolved"})


def test_invalid_candidate_stops_job_before_extract_downstream_or_publish(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    output = _decision(manifest, role="auxiliary")
    output["decisions"][0]["candidate_id"] = "layout-candidate:s2:unknown"
    provider = FakeV2CandidateProvider({"candidates": []}, layout_payload=output)
    dependencies, dictionary, matcher = _dependencies(provider)

    with pytest.raises(AIEnhancedJobPause) as raised:
        run_ai_enhanced_v2_job(
            path,
            runtime_root=tmp_path / "runtime",
            publish_root=tmp_path / "publish",
            dependencies=dependencies,
            client_idempotency_key="invalid-client",
            business_key="invalid-business",
            on_stage=lambda *_args: None,
        )

    assert raised.value.code == "AI_V2_STRUCTURE_DECISION_INVALID"
    assert provider.structure_call_count == 1
    assert provider.extraction_call_count == dictionary.calls == matcher.calls == 0
    assert not (tmp_path / "publish").exists()


def test_manifest_source_or_context_tampering_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    bad_source = dict(manifest)
    bad_source["source_file_sha256"] = "0" * 64
    bad_context = dict(manifest)
    bad_context["context_sha256"] = "0" * 64

    with pytest.raises(StructureManifestAdapterError):
        provider_structure_payload(bad_source)
    with pytest.raises(StructureManifestAdapterError):
        provider_structure_payload(bad_context)


def test_resolved_structure_summary_replays_without_layout_call(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    provider = FakeV2CandidateProvider(
        {"candidates": []}, layout_payload=_decision(manifest, role="auxiliary")
    )
    dependencies, _dictionary, _matcher = _dependencies(provider)
    captured = {}
    first = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime-1",
        publish_root=tmp_path / "publish-1",
        dependencies=dependencies,
        client_idempotency_key="first-client",
        business_key="first-business",
        on_stage=lambda *_args: None,
        on_structure_summary=lambda value: captured.update(value),
    )
    layout_before = provider.structure_call_count
    second = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime-2",
        publish_root=tmp_path / "publish-2",
        dependencies=dependencies,
        client_idempotency_key="second-client",
        business_key="second-business",
        on_stage=lambda *_args: None,
        persisted_structure_summary=captured,
    )

    assert first.structure_summary == second.structure_summary
    assert provider.structure_call_count == layout_before == 1
    assert second.layout_call_count == 0

    changed_model = replace(dependencies, model_name="offline-test-v2")
    third = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime-3",
        publish_root=tmp_path / "publish-3",
        dependencies=changed_model,
        client_idempotency_key="third-client",
        business_key="third-business",
        on_stage=lambda *_args: None,
        persisted_structure_summary=captured,
    )
    assert provider.structure_call_count == 2
    assert third.layout_call_count == 1


def test_job_usage_is_delta_not_provider_session_total(tmp_path: Path) -> None:
    path = tmp_path / "known-aux.xlsx"
    _write_known_plus_auxiliary(path)
    _preprocessed, manifest = _context(path)
    provider = _UsageProvider(_decision(manifest, role="auxiliary"))
    dependencies, _dictionary, _matcher = _dependencies(provider)

    result = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime",
        publish_root=tmp_path / "publish",
        dependencies=dependencies,
        client_idempotency_key="usage-client",
        business_key="usage-business",
        on_stage=lambda *_args: None,
    )

    assert result.usage == {
        "input_tokens": 25,
        "output_tokens": 5,
        "total_tokens": 30,
    }
    assert result.http_attempt_count == 4
    assert provider.usage_summary["total_tokens"] == 1530


def test_local_structure_fast_path_still_has_zero_layout_calls(tmp_path: Path) -> None:
    path = tmp_path / "local.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet", "200*240", "White cotton", "2"])
    workbook.save(path)
    provider = FakeV2CandidateProvider({"candidates": []})
    dependencies, _dictionary, _matcher = _dependencies(provider)

    result = run_ai_enhanced_v2_job(
        path,
        runtime_root=tmp_path / "runtime",
        publish_root=tmp_path / "publish",
        dependencies=dependencies,
        client_idempotency_key="local-client",
        business_key="local-business",
        on_stage=lambda *_args: None,
    )

    assert result.structure_status == "locally_resolved"
    assert provider.structure_call_count == 0
    assert provider.extraction_call_count == 1
