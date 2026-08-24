from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import socket
import time
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Mapping
from urllib.error import HTTPError, URLError
from urllib.request import ProxyHandler, Request, build_opener

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order.orchestration import (
    build_chunk_manifest,
    build_v2_extraction_units,
)
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook
from bedding_order_parser.ai_full_order.structure_manifest import (
    LAYOUT_CONTRACT_VERSION,
    LAYOUT_PROMPT_VERSION,
    MAX_STRUCTURE_EXCERPT_CHARS,
    MAX_STRUCTURE_EXCERPTS,
    STRUCTURE_CONTEXT_VERSION,
    STRUCTURE_MANIFEST_VERSION,
    build_structure_manifest,
    provider_structure_payload,
    validate_structure_manifest,
)
from bedding_order_parser.ai_full_order.structure_resolution import (
    LAYOUT_OUTPUT_SCHEMA,
    apply_structure_decision,
)
from bedding_order_parser.ai_full_order.volcengine_ark import (
    FULL_ORDER_LAYOUT_FUNCTION,
    VolcengineArkFullOrderProvider,
)
from bedding_order_parser.llm.settings import (
    DEFAULT_ARK_BASE_URL,
    LLMSettings,
    VOLCENGINE_ARK_PROVIDER,
)
from bedding_order_parser.llm.transport import (
    JSONHTTPRequest,
    JSONHTTPResponse,
    JSONTransport,
    TransportConnectionError,
    TransportTimeout,
)


APPROVED_MODEL = "doubao-seed-2-0-lite-260428"
AUTHORIZATION_ENV = "ARK_LAYOUT_V2_D4A3E_AUTHORIZATION"
AUTHORIZATION_VALUE = "CONFIRMED_SINGLE_LAYOUT_CALL"
MAX_OUTPUT_TOKENS = 1024


class AcceptanceGateError(RuntimeError):
    """Stop before an unsafe or over-budget real structure request."""


class _SingleAttemptDirectTransport:
    def __init__(self) -> None:
        self.calls = 0
        self._opener = build_opener(ProxyHandler({}))

    def send(self, request: JSONHTTPRequest) -> JSONHTTPResponse:
        if self.calls:
            raise AcceptanceGateError("Real transport budget is already exhausted.")
        self.calls += 1
        started = time.monotonic()
        raw_request = Request(
            request.url,
            data=request.body,
            headers=dict(request.headers),
            method=request.method,
        )
        try:
            with self._opener.open(
                raw_request, timeout=request.timeout_seconds
            ) as response:
                return JSONHTTPResponse(
                    status_code=int(response.status),
                    headers=_normalized_headers(response.headers),
                    body=response.read(),
                    elapsed_ms=_elapsed_ms(started),
                )
        except HTTPError as exc:
            return JSONHTTPResponse(
                status_code=int(exc.code),
                headers=_normalized_headers(exc.headers),
                body=exc.read(),
                elapsed_ms=_elapsed_ms(started),
            )
        except (TimeoutError, socket.timeout) as exc:
            raise TransportTimeout("The bounded Ark layout request timed out.") from exc
        except URLError as exc:
            if isinstance(exc.reason, (TimeoutError, socket.timeout)):
                raise TransportTimeout("The bounded Ark layout request timed out.") from exc
            raise TransportConnectionError("The bounded Ark layout connection failed.") from exc
        except OSError as exc:
            raise TransportConnectionError("The bounded Ark layout connection failed.") from exc


class _BoundedLayoutTransport:
    """Validate the exact synthetic request before permitting one HTTP attempt."""

    def __init__(
        self,
        inner: JSONTransport,
        *,
        expected_payload: Mapping[str, Any],
        expected_model: str,
        expected_base_url: str,
        forbidden_secret: str,
    ) -> None:
        self.inner = inner
        self.expected_payload = dict(expected_payload)
        self.expected_model = expected_model
        self.expected_url = f"{expected_base_url.rstrip('/')}/responses"
        self.forbidden_secret = forbidden_secret
        self.http_attempts = 0
        self.status_class = "not_called"
        self.response_shape = "not_called"
        self.request_audit: dict[str, Any] = {}

    def send(self, request: JSONHTTPRequest) -> JSONHTTPResponse:
        if self.http_attempts:
            raise AcceptanceGateError("Layout acceptance permits one HTTP attempt only.")
        body = _validated_bounded_body(
            request,
            expected_url=self.expected_url,
            expected_model=self.expected_model,
            expected_payload=self.expected_payload,
            forbidden_secret=self.forbidden_secret,
        )
        body["max_output_tokens"] = MAX_OUTPUT_TOKENS
        bounded = JSONHTTPRequest(
            method=request.method,
            url=request.url,
            headers=request.headers,
            body=json.dumps(body, ensure_ascii=False, separators=(",", ":")).encode(
                "utf-8"
            ),
            timeout_seconds=request.timeout_seconds,
        )
        self.request_audit = {
            "responses_api": request.url.endswith("/responses"),
            "store": body["store"],
            "stream": bool(body.get("stream", False)),
            "strict_function": body["tools"][0]["strict"] is True,
            "function_name": body["tool_choice"]["name"],
            "max_output_tokens": body["max_output_tokens"],
            "synthetic_structure_only": True,
        }
        self.http_attempts = 1
        response = self.inner.send(bounded)
        self.status_class = f"{int(response.status_code) // 100}xx"
        self.response_shape = _response_shape(response)
        return response


class _StaticLayoutTransport:
    def __init__(self, expected_payload: Mapping[str, Any]) -> None:
        self.calls = 0
        self.expected_payload = expected_payload

    def send(self, request: JSONHTTPRequest) -> JSONHTTPResponse:
        self.calls += 1
        body = json.loads(request.body)
        context = json.loads(body["input"][1]["content"][0]["text"])
        assert context == self.expected_payload
        sheet = context["unresolved_sheets"][0]
        candidate = sheet["candidate_options"][0]
        response = {
            "id": "offline-layout-v2",
            "model": APPROVED_MODEL,
            "status": "completed",
            "output": [{
                "type": "function_call",
                "name": FULL_ORDER_LAYOUT_FUNCTION,
                "arguments": json.dumps({
                    "layout_contract_version": LAYOUT_CONTRACT_VERSION,
                    "status": "resolved",
                    "decisions": [{
                        "sheet_id": sheet["sheet_id"],
                        "role": "auxiliary",
                        "candidate_id": candidate["candidate_id"],
                        "reason": "auxiliary_non_order_content",
                    }],
                }),
            }],
            "usage": {
                "input_tokens": 120,
                "output_tokens": 18,
                "total_tokens": 138,
            },
        }
        return JSONHTTPResponse(
            200,
            {"x-request-id": "offline-layout-header"},
            json.dumps(response).encode("utf-8"),
            1,
        )


def synthetic_workbook_bytes() -> bytes:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "Synthetic Orders"
    orders.append(["No.", "Item", "Size", "Specification", "Qty"])
    for number in range(1, 4):
        orders.append(
            [str(number), "Sample Linen", "200*240", "Synthetic white", str(number)]
        )
    notes = workbook.create_sheet("Synthetic Notes")
    notes["A1"] = "Packing reference"
    notes["A2"] = "Color reference"
    notes["B2"] = "Synthetic white"
    hidden = workbook.create_sheet("Hidden Synthetic")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "hidden-synthetic-marker"
    raw = BytesIO()
    workbook.save(raw)
    workbook.close()
    return raw.getvalue()


def offline_preflight(content: bytes) -> dict[str, Any]:
    with TemporaryDirectory(prefix="bedding-d4a3e-preflight-") as raw:
        path = Path(raw) / "synthetic-layout.xlsx"
        path.write_bytes(content)
        preprocessed = preprocess_workbook(path)
        manifest = build_structure_manifest(
            preprocessed, build_chunk_manifest(preprocessed)
        )
        validate_structure_manifest(manifest)
        payload = provider_structure_payload(manifest)
        states = {item.sheet_id: item for item in preprocessed.sheet_states}
        if preprocessed.structure_status != "ambiguous":
            raise AcceptanceGateError("Synthetic workbook must require structure resolution.")
        if states["s1"].local_status != "confirmed_order":
            raise AcceptanceGateError("Synthetic s1 must be locally confirmed.")
        if states["s2"].local_status != "unresolved_order_candidate":
            raise AcceptanceGateError("Synthetic s2 must be locally unresolved.")
        if len(preprocessed.records) != 3:
            raise AcceptanceGateError("Synthetic s1 must contain exactly three records.")
        if [item["sheet_id"] for item in payload["known_chunks"]] != ["s1"]:
            raise AcceptanceGateError("Known structure context must contain only s1.")
        if [item["sheet_id"] for item in payload["unresolved_sheets"]] != ["s2"]:
            raise AcceptanceGateError("Unresolved structure context must contain only s2.")
        options = payload["unresolved_sheets"][0]["candidate_options"]
        if len(options) != 1 or options[0]["role"] != "auxiliary":
            raise AcceptanceGateError("Synthetic s2 must expose one auxiliary candidate.")
        _validate_safe_structure_payload(payload)
        return {
            "source_sha256": preprocessed.source_file_sha256,
            "context_sha256": manifest["context_sha256"],
            "manifest_version": manifest["manifest_version"],
            "structure_context_version": payload["structure_context_version"],
            "layout_contract_version": payload["layout_contract_version"],
            "layout_prompt_version": LAYOUT_PROMPT_VERSION,
            "known_sheet_ids": ["s1"],
            "unresolved_sheet_ids": ["s2"],
            "record_count": len(preprocessed.records),
            "auxiliary_candidate_id": options[0]["candidate_id"],
            "payload_security_valid": True,
            "hidden_sheet_excluded": True,
            "free_text_bounded": True,
            "provider_calls": 0,
            "http_attempts": 0,
        }


def run_acceptance(
    settings: LLMSettings,
    inner_transport: JSONTransport,
    *,
    require_user_authorization: bool,
) -> dict[str, Any]:
    configuration = _configuration_audit(
        settings, require_user_authorization=require_user_authorization
    )
    content = synthetic_workbook_bytes()
    preflight = offline_preflight(content)
    summary: dict[str, Any] = {}
    temporary_root: Path
    with TemporaryDirectory(prefix="bedding-d4a3e-real-") as raw:
        temporary_root = Path(raw)
        path = temporary_root / "synthetic-layout.xlsx"
        path.write_bytes(content)
        preprocessed = preprocess_workbook(path)
        manifest = build_structure_manifest(
            preprocessed, build_chunk_manifest(preprocessed)
        )
        safe_payload = provider_structure_payload(manifest)
        transport = _BoundedLayoutTransport(
            inner_transport,
            expected_payload=safe_payload,
            expected_model=settings.model,
            expected_base_url=settings.base_url,
            forbidden_secret=settings.api_key,
        )
        provider = VolcengineArkFullOrderProvider(settings, transport=transport)
        usage_before = dict(provider.usage_summary)
        output: Mapping[str, Any] | None = None
        application = None
        units = ()
        failure_category = ""
        try:
            output = provider.resolve_structure(manifest)
            application = apply_structure_decision(preprocessed, manifest, output)
            if application.resolved:
                units = build_v2_extraction_units(application.preprocessed)
        except BaseException as exc:
            failure_category = type(exc).__name__
        finally:
            telemetry = provider.latest_telemetry.to_dict()
            usage = {
                name: int(provider.usage_summary[name]) - int(usage_before[name])
                for name in ("input_tokens", "output_tokens", "total_tokens")
            }
            decision = (
                dict(output["decisions"][0])
                if output is not None and output.get("decisions")
                else {}
            )
            expected_candidate = preflight["auxiliary_candidate_id"]
            summary = {
                "configuration": configuration,
                "preflight": {
                    **preflight,
                    "source_sha256": _sha_prefix(preflight["source_sha256"]),
                    "context_sha256": _sha_prefix(preflight["context_sha256"]),
                    "auxiliary_candidate_id": expected_candidate,
                },
                "provider": {
                    "provider": provider.provider_name,
                    "model": provider.model_name,
                    "layout_logical_calls": provider.structure_call_count,
                    "extraction_logical_calls": provider.extraction_call_count,
                    "http_attempts": transport.http_attempts,
                    "retry_count": settings.max_retries,
                    "http_status_class": transport.status_class,
                    "response_shape": transport.response_shape,
                    "request_id": _redact_request_id(str(telemetry["request_id"])),
                    "latency_ms": int(telemetry["latency_ms"]),
                    "input_tokens": usage["input_tokens"],
                    "output_tokens": usage["output_tokens"],
                    "total_tokens": usage["total_tokens"],
                },
                "request": dict(transport.request_audit),
                "result": {
                    "failure_category": failure_category,
                    "status": str(output.get("status", "")) if output else "",
                    "sheet_id": str(decision.get("sheet_id", "")),
                    "role": str(decision.get("role", "")),
                    "candidate_id": str(decision.get("candidate_id", "")),
                    "candidate_from_request": bool(decision)
                    and decision.get("candidate_id") == expected_candidate,
                    "reason": str(decision.get("reason", "")),
                    "strict_contract_valid": output is not None,
                    "local_validator_valid": application is not None,
                    "local_apply_valid": bool(application and application.resolved),
                    "preserved_record_count": (
                        len(application.preprocessed.records) if application else 0
                    ),
                    "extraction_unit_count": len(units),
                },
                "safety": {
                    "synthetic_data_only": True,
                    "real_business_data_used": False,
                    "raw_request_saved": False,
                    "raw_response_saved": False,
                    "secret_saved": False,
                },
            }
            _assert_safe_summary(summary, settings.api_key)
            _assert_no_secret_or_raw_payload_files(temporary_root, settings.api_key)
    summary["cleanup"] = {
        "temporary_root_removed": not temporary_root.exists(),
        "synthetic_workbook_removed": not temporary_root.exists(),
    }
    return summary


def _configuration_audit(
    settings: LLMSettings, *, require_user_authorization: bool
) -> dict[str, Any]:
    if require_user_authorization and os.environ.get(AUTHORIZATION_ENV) != AUTHORIZATION_VALUE:
        raise AcceptanceGateError("Explicit one-call authorization marker is missing.")
    if settings.configuration_status() != "ready":
        raise AcceptanceGateError("Ark configuration is not ready.")
    if settings.provider != VOLCENGINE_ARK_PROVIDER:
        raise AcceptanceGateError("Unexpected provider configuration.")
    if settings.model != APPROVED_MODEL:
        raise AcceptanceGateError("Configured model is not approved for this Gate.")
    if settings.base_url.rstrip("/") != DEFAULT_ARK_BASE_URL:
        raise AcceptanceGateError("Configured Base URL is not approved for this Gate.")
    if settings.max_retries != 0:
        raise AcceptanceGateError("LLM_MAX_RETRIES must be zero.")
    return {
        "provider_ready": True,
        "api_key_present": settings.api_key_configured,
        "provider": settings.provider,
        "model": settings.model,
        "base_url_id": "ark-cn-beijing-api-v3",
        "responses_api": True,
        "store": False,
        "stream": False,
        "provider_retries": settings.max_retries,
        "transport_max_http_attempts": 1,
        "max_output_tokens": MAX_OUTPUT_TOKENS,
    }


def _validated_bounded_body(
    request: JSONHTTPRequest,
    *,
    expected_url: str,
    expected_model: str,
    expected_payload: Mapping[str, Any],
    forbidden_secret: str,
) -> dict[str, Any]:
    if request.method != "POST" or request.url != expected_url:
        raise AcceptanceGateError("Ark layout request endpoint is invalid.")
    body = json.loads(request.body)
    if body.get("model") != expected_model or body.get("store") is not False:
        raise AcceptanceGateError("Ark layout request settings are invalid.")
    if body.get("stream", False) is not False:
        raise AcceptanceGateError("Ark layout request must be non-streaming.")
    tools = body.get("tools")
    if not isinstance(tools, list) or len(tools) != 1:
        raise AcceptanceGateError("Ark layout request must contain exactly one tool.")
    tool = tools[0]
    if (
        tool.get("name") != FULL_ORDER_LAYOUT_FUNCTION
        or tool.get("strict") is not True
        or tool.get("parameters") != LAYOUT_OUTPUT_SCHEMA
        or body.get("tool_choice", {}).get("name") != FULL_ORDER_LAYOUT_FUNCTION
    ):
        raise AcceptanceGateError("Ark layout strict function contract is invalid.")
    try:
        payload = json.loads(body["input"][1]["content"][0]["text"])
    except (KeyError, IndexError, TypeError, json.JSONDecodeError) as exc:
        raise AcceptanceGateError("Ark layout context cannot be safely decoded.") from exc
    if payload != expected_payload:
        raise AcceptanceGateError("Ark layout context differs from local validated payload.")
    _validate_safe_structure_payload(payload)
    serialized = request.body.decode("utf-8")
    if forbidden_secret and forbidden_secret in serialized:
        raise AcceptanceGateError("API key appeared in the request body.")
    if "Authorization" in serialized or re.search(r"[A-Za-z]:\\", serialized):
        raise AcceptanceGateError("Forbidden request-body content was detected.")
    return body


def _validate_safe_structure_payload(payload: Mapping[str, Any]) -> None:
    if set(payload) != {
        "structure_context_version",
        "layout_contract_version",
        "known_chunks",
        "unresolved_sheets",
    }:
        raise AcceptanceGateError("Structure payload fields are invalid.")
    if (
        payload["structure_context_version"] != STRUCTURE_CONTEXT_VERSION
        or payload["layout_contract_version"] != LAYOUT_CONTRACT_VERSION
    ):
        raise AcceptanceGateError("Structure payload version is invalid.")
    if len(payload["known_chunks"]) != 1 or len(payload["unresolved_sheets"]) != 1:
        raise AcceptanceGateError("Structure payload must contain one known and one unresolved Sheet.")
    unresolved = payload["unresolved_sheets"][0]
    excerpts = unresolved.get("excerpts")
    if not isinstance(excerpts, list) or len(excerpts) > MAX_STRUCTURE_EXCERPTS:
        raise AcceptanceGateError("Structure excerpts exceed their count bound.")
    if any(
        not isinstance(item, Mapping)
        or len(str(item.get("text", ""))) > MAX_STRUCTURE_EXCERPT_CHARS
        for item in excerpts
    ):
        raise AcceptanceGateError("Structure excerpts exceed their length bound.")
    serialized = json.dumps(payload, ensure_ascii=False)
    forbidden = (
        "hidden-synthetic-marker",
        "PK\x03\x04",
        "Authorization",
        "api_key",
        "evidence_catalog",
        "formula_text",
    )
    if any(item in serialized for item in forbidden):
        raise AcceptanceGateError("Structure payload contains forbidden content.")
    if re.search(r"[A-Za-z]:\\", serialized):
        raise AcceptanceGateError("Structure payload contains a local path.")


def _response_shape(response: JSONHTTPResponse) -> str:
    try:
        payload = json.loads(response.body)
    except (UnicodeDecodeError, json.JSONDecodeError):
        return "non_json"
    output = payload.get("output") if isinstance(payload, Mapping) else None
    if isinstance(output, list) and any(
        isinstance(item, Mapping) and item.get("type") == "function_call"
        for item in output
    ):
        return "function_call"
    if isinstance(output, list) and any(
        isinstance(item, Mapping) and item.get("type") == "message"
        for item in output
    ):
        return "output_text"
    return "other_json"


def _assert_safe_summary(summary: Mapping[str, Any], api_key: str) -> None:
    serialized = json.dumps(summary, ensure_ascii=False)
    if api_key and api_key in serialized:
        raise AcceptanceGateError("API key appeared in the safe summary.")
    forbidden = ("Authorization", "system_prompt")
    if any(item in serialized for item in forbidden):
        raise AcceptanceGateError("Safe summary contains a forbidden field.")


def _assert_no_secret_or_raw_payload_files(root: Path, api_key: str) -> None:
    forbidden_names = ("raw_response", "provider_response", "request_body", "authorization")
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if any(name in path.name.casefold() for name in forbidden_names):
            raise AcceptanceGateError("A raw provider payload file was created.")
        content = path.read_bytes()
        if api_key and api_key.encode("utf-8") in content:
            raise AcceptanceGateError("An API key was persisted in the temporary root.")
        if b"Authorization" in content:
            raise AcceptanceGateError("An Authorization header was persisted.")


def _redact_request_id(value: str) -> str:
    if len(value) <= 8:
        return "present" if value else ""
    return f"{value[:4]}...{value[-4:]}"


def _sha_prefix(value: str) -> str:
    return f"{value[:12]}..."


def _normalized_headers(headers: object) -> dict[str, str]:
    items = getattr(headers, "items", None)
    if not callable(items):
        return {}
    return {str(key).lower(): str(value) for key, value in items()}


def _elapsed_ms(started: float) -> int:
    return max(0, round((time.monotonic() - started) * 1000))


def test_synthetic_multi_sheet_preflight_is_safe_and_complete() -> None:
    summary = offline_preflight(synthetic_workbook_bytes())
    assert summary["manifest_version"] == STRUCTURE_MANIFEST_VERSION
    assert summary["structure_context_version"] == STRUCTURE_CONTEXT_VERSION
    assert summary["layout_contract_version"] == LAYOUT_CONTRACT_VERSION
    assert summary["layout_prompt_version"] == LAYOUT_PROMPT_VERSION
    assert summary["known_sheet_ids"] == ["s1"]
    assert summary["unresolved_sheet_ids"] == ["s2"]
    assert summary["record_count"] == 3
    assert summary["provider_calls"] == summary["http_attempts"] == 0


def test_bounded_harness_applies_formal_provider_with_fake_transport() -> None:
    content = synthetic_workbook_bytes()
    with TemporaryDirectory(prefix="bedding-d4a3e-fake-context-") as raw:
        path = Path(raw) / "synthetic.xlsx"
        path.write_bytes(content)
        preprocessed = preprocess_workbook(path)
        manifest = build_structure_manifest(
            preprocessed, build_chunk_manifest(preprocessed)
        )
        payload = provider_structure_payload(manifest)
    settings = LLMSettings(
        enabled=True,
        provider=VOLCENGINE_ARK_PROVIDER,
        model=APPROVED_MODEL,
        base_url=DEFAULT_ARK_BASE_URL,
        api_key="offline-layout-secret",
        max_retries=0,
    )
    inner = _StaticLayoutTransport(payload)
    summary = run_acceptance(
        settings, inner, require_user_authorization=False
    )
    assert summary["provider"]["layout_logical_calls"] == 1
    assert summary["provider"]["extraction_logical_calls"] == 0
    assert summary["provider"]["http_attempts"] == 1
    assert summary["result"]["strict_contract_valid"] is True
    assert summary["result"]["candidate_from_request"] is True
    assert summary["result"]["local_apply_valid"] is True
    assert summary["result"]["extraction_unit_count"] == 3
    assert summary["cleanup"]["temporary_root_removed"] is True
    assert inner.calls == 1


def test_nonzero_retry_is_rejected_before_transport() -> None:
    content = synthetic_workbook_bytes()
    with TemporaryDirectory(prefix="bedding-d4a3e-retry-context-") as raw:
        path = Path(raw) / "synthetic.xlsx"
        path.write_bytes(content)
        preprocessed = preprocess_workbook(path)
        manifest = build_structure_manifest(
            preprocessed, build_chunk_manifest(preprocessed)
        )
        payload = provider_structure_payload(manifest)
    inner = _StaticLayoutTransport(payload)
    settings = LLMSettings(
        enabled=True,
        provider=VOLCENGINE_ARK_PROVIDER,
        model=APPROVED_MODEL,
        base_url=DEFAULT_ARK_BASE_URL,
        api_key="offline-layout-secret",
        max_retries=1,
    )
    with pytest.raises(AcceptanceGateError, match="zero"):
        run_acceptance(settings, inner, require_user_authorization=False)
    assert inner.calls == 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--real", action="store_true")
    args = parser.parse_args()
    if not args.real:
        raise SystemExit("Use pytest for offline validation or pass --real explicitly.")
    settings = LLMSettings.from_environment()
    summary = run_acceptance(
        settings,
        _SingleAttemptDirectTransport(),
        require_user_authorization=True,
    )
    print(json.dumps(summary, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
