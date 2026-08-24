from __future__ import annotations

import socket
from pathlib import Path

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order.contracts import AI_BUSINESS_FIELD_NAMES
from bedding_order_parser.ai_full_order.fake_provider import FakeV2CandidateProvider
from bedding_order_parser.ai_full_order.comparison import (
    V2ComparisonStatus,
    V2ReviewSeverity,
)
from bedding_order_parser.ai_full_order.field_policy import V2DecisionReason, V2DecisionStatus
from bedding_order_parser.ai_full_order.orchestration import (
    BatchStatus,
    build_v2_extraction_units,
    run_v2_offline_orchestration,
)
from bedding_order_parser.ai_full_order.preprocessing import EvidenceItem, preprocess_workbook
from bedding_order_parser.ai_full_order.python_shadow import build_deterministic_python_shadow
from bedding_order_parser.excel.workbook_reader import compute_sha256


def _write_simple(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    _metadata(sheet)
    sheet.append(["Bank account:", "SYNTHETIC-DO-NOT-SEND", "", "", ""])
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet Cover", "200*240", "100% cotton white", "12"])
    workbook.save(path)


def _write_multilevel(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    _metadata(sheet)
    sheet.append(["No.", "Linens Description", "", "", "Qty", "Remarks"])
    sheet.append(["", "Item", "After Wash Size", "Specification", "", "Remarks"])
    sheet.append(
        [
            "1",
            "Duvet Cover King",
            "180*270cm",
            "300TC cotton white, Bag with Hand Holes, no flange",
            "300",
            "Delivery Size",
        ]
    )
    sheet.merge_cells("A5:A6")
    sheet.merge_cells("B5:D5")
    sheet.merge_cells("E5:E6")
    sheet.merge_cells("F5:F6")
    workbook.save(path)


def _write_semantic(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    _metadata(sheet)
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(
        [
            "1",
            "Duvet Cover",
            "200*240",
            "100% cotton white, Carton packed, bag style without embroidery",
            "12",
        ]
    )
    workbook.save(path)


def _metadata(sheet) -> None:
    sheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)", ""])
    sheet.append(["BUYER:", "", "", "", "", ""])
    sheet.append(["Test Hotel", "", "", "", "Contact Person: Aaron Lee", ""])
    sheet.append(["Delivery date:", "2026-09-30", "", "", "", ""])


def _shadow_context(path: Path):
    preprocessed = preprocess_workbook(path)
    units = build_v2_extraction_units(preprocessed)
    evidence: dict[str, EvidenceItem] = {}
    for unit in units:
        evidence.update({item.evidence_id: item for item in unit.evidence_catalog})
    shadow = build_deterministic_python_shadow(
        path,
        preprocessed,
        target_records=[unit.target for unit in units],
        evidence_catalog=tuple(evidence.values()),
    )
    return preprocessed, units, shadow


@pytest.mark.parametrize(
    "writer",
    [_write_simple, _write_multilevel, _write_semantic],
    ids=["simple", "multilevel_inheritance", "semantic_description"],
)
def test_three_synthetic_fixtures_build_nonempty_shadow_and_ready_canonical_batch(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    writer,
) -> None:
    path = tmp_path / f"{writer.__name__}.xlsx"
    writer(path)
    before = compute_sha256(path)
    preprocessed, units, shadow = _shadow_context(path)
    provider = FakeV2CandidateProvider({"candidates": []})

    monkeypatch.setattr(
        socket,
        "create_connection",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("network forbidden")),
    )
    result = run_v2_offline_orchestration(preprocessed, provider, shadow)

    assert compute_sha256(path) == before
    assert preprocessed.structure_status == "locally_resolved"
    assert result.structure_recognition_calls == 0
    assert provider.structure_call_count == 0
    assert result.extraction_calls == len(units) == 1
    assert result.network_calls == provider.network_call_count == 0
    assert result.batch.ready_for_downstream
    assert result.batch.status is BatchStatus.READY
    assert len(provider.requests[0]["target"]) == 6
    assert set(provider.requests[0]["target"]["evidence_ids"]) == {
        item["evidence_id"] for item in provider.requests[0]["evidence_catalog"]
    }
    assert len(provider.requests[0].get("records", [])) == 0
    assert "SYNTHETIC-DO-NOT-SEND" not in str(provider.requests[0])

    direct_shadow_fields = {
        name
        for name, candidate in shadow[0].fields.items()
        if candidate.has_direct_evidence
    }
    assert {"客户", "币种", "业务员", "数量", "计划发货日期"} <= direct_shadow_fields
    allowed_evidence = set(units[0].target.evidence_ids)
    assert all(
        set(candidate.evidence_ids) <= allowed_evidence
        for candidate in shadow[0].fields.values()
    )
    canonical = result.batch.records[0]
    assert tuple(canonical.business_fields()) == AI_BUSINESS_FIELD_NAMES
    assert all(isinstance(value, str) for value in canonical.business_fields().values())
    assert canonical.line_number == "1"
    assert "行号" not in canonical.business_fields()
    assert "物料编码" not in canonical.business_fields()
    assert "相似分数" not in canonical.business_fields()


def test_semantic_ai_candidate_is_accepted_when_python_has_no_direct_candidate(
    tmp_path: Path,
) -> None:
    path = tmp_path / "semantic.xlsx"
    _write_semantic(path)
    preprocessed, units, shadow = _shadow_context(path)
    evidence = next(
        item
        for item in units[0].evidence_catalog
        if "Carton packed" in item.original_text
    )
    provider = FakeV2CandidateProvider(
        {
            "candidates": [
                {
                    "field_name": "包装方式",
                    "candidate_value": "纸箱包装",
                    "evidence_references": [evidence.evidence_id],
                    "interpretation": "semantic",
                    "supporting_quote": "Carton packed",
                }
            ]
        }
    )

    result = run_v2_offline_orchestration(preprocessed, provider, shadow)

    decision = result.batch.records[0].decisions["包装方式"]
    assert result.batch.ready_for_downstream
    assert decision.value == "纸箱包装"
    assert decision.reason_code is V2DecisionReason.AI_SEMANTIC_SELECTED
    assert decision.ai_candidate is not None
    assert decision.ai_candidate.quote_span is not None


def test_candidate_content_issue_is_reviewable_without_blocking_batch(tmp_path: Path) -> None:
    path = tmp_path / "ordinary-issue.xlsx"
    _write_semantic(path)
    preprocessed, units, shadow = _shadow_context(path)
    evidence = next(
        item for item in units[0].evidence_catalog if "Carton packed" in item.original_text
    )
    provider = FakeV2CandidateProvider(
        {
            "candidates": [
                {
                    "field_name": "包装方式",
                    "candidate_value": "invented packaging",
                    "evidence_references": [evidence.evidence_id],
                    "interpretation": "direct",
                    "supporting_quote": "",
                }
            ]
        }
    )

    result = run_v2_offline_orchestration(preprocessed, provider, shadow)

    decision = result.batch.records[0].decisions["包装方式"]
    assert result.batch.technical_ready
    assert result.batch.review_required
    assert decision.status is V2DecisionStatus.MISSING
    assert decision.candidate_issue_code == "direct_candidate_untraceable"
    assert decision.value == ""


def test_high_review_direct_conflict_keeps_ai_value_and_batch_ready(tmp_path: Path) -> None:
    path = tmp_path / "high-risk-conflict.xlsx"
    _write_simple(path)
    preprocessed, units, shadow = _shadow_context(path)
    salesperson_evidence = next(
        item for item in units[0].evidence_catalog if "Aaron Lee" in item.original_text
    )
    provider = FakeV2CandidateProvider(
        {
            "candidates": [
                {
                    "field_name": "客户",
                    "candidate_value": "Aaron Lee",
                    "evidence_references": [salesperson_evidence.evidence_id],
                    "interpretation": "direct",
                    "supporting_quote": "",
                }
            ]
        }
    )

    result = run_v2_offline_orchestration(preprocessed, provider, shadow)

    assert result.batch.status is BatchStatus.READY
    assert result.batch.technical_ready
    assert result.batch.review_required
    assert result.batch.high_review_count >= 1
    decision = result.batch.records[0].decisions["客户"]
    assert decision.status is V2DecisionStatus.AI_SELECTED
    assert decision.value == "Aaron Lee"
    assert decision.comparison_status is V2ComparisonStatus.DIFFERENT
    assert decision.review_required is True
    assert decision.review_severity is V2ReviewSeverity.HIGH
    assert decision.blocking is False


def test_v2_extraction_unit_identity_is_stable_and_single_target(tmp_path: Path) -> None:
    path = tmp_path / "stable.xlsx"
    _write_simple(path)

    first = build_v2_extraction_units(preprocess_workbook(path))
    second = build_v2_extraction_units(preprocess_workbook(path))

    assert [item.extraction_unit_id for item in first] == [
        item.extraction_unit_id for item in second
    ]
    assert all(item.target.record_local_id for item in first)
    assert all(item.target.evidence_ids for item in first)


@pytest.mark.parametrize(
    "payload",
    [
        {"candidates": [{"field_name": "颜色", "candidate_value": "invented", "evidence_references": ["unknown"], "interpretation": "direct", "supporting_quote": ""}]},
        {"candidates": [{"field_name": "颜色", "candidate_value": "white", "evidence_references": ["unknown"], "interpretation": "direct", "supporting_quote": ""}, {"field_name": "颜色", "candidate_value": "blue", "evidence_references": ["unknown-2"], "interpretation": "direct", "supporting_quote": ""}]},
    ],
    ids=["unknown_evidence", "duplicate_field"],
)
def test_hard_v2_failures_reject_the_unit_and_batch(tmp_path: Path, payload) -> None:
    path = tmp_path / "hard-failure.xlsx"
    _write_simple(path)
    preprocessed, _units, shadow = _shadow_context(path)

    result = run_v2_offline_orchestration(
        preprocessed,
        FakeV2CandidateProvider(payload),
        shadow,
    )

    assert result.batch.status is BatchStatus.ISOLATED
    assert "hard_contract_or_resolution_failure" in result.batch.reasons
    assert result.batch.records == ()
