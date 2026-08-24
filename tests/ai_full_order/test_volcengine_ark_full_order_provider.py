from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from bedding_order_parser.ai_full_order.downstream import MaterialMatchOutput, MaterialSelection
from bedding_order_parser.ai_full_order.fake_provider import FakeFullOrderProvider
from bedding_order_parser.ai_full_order.preprocessing import preprocess_workbook
from bedding_order_parser.ai_full_order.orchestration import (
    build_chunk_manifest,
    build_v2_extraction_request,
    build_v2_extraction_units,
)
from bedding_order_parser.ai_full_order.structure_manifest import build_structure_manifest
from bedding_order_parser.ai_full_order.volcengine_ark import (
    FULL_ORDER_EXTRACTION_FUNCTION,
    FULL_ORDER_LAYOUT_FUNCTION,
    FULL_ORDER_V2_EXTRACTION_FUNCTION,
    FULL_ORDER_V2_PROMPT_VERSION,
    VolcengineArkFullOrderProvider,
)
from bedding_order_parser.ai_full_order.contracts import FullOrderContractError
from bedding_order_parser.llm.settings import LLMSettings
from bedding_order_parser.llm.transport import JSONHTTPRequest, JSONHTTPResponse
from bedding_order_parser.web.services import JobService


SECRET = "ark-d1-test-secret-never-log"


class RecordingTransport:
    def __init__(self, responder) -> None:
        self.responder = responder
        self.requests: list[JSONHTTPRequest] = []

    def send(self, request: JSONHTTPRequest) -> JSONHTTPResponse:
        self.requests.append(request)
        return self.responder(request)


class DeferredExecutor:
    def submit(self, _function, *_args) -> None:
        return None

    def shutdown(self, **_kwargs) -> None:
        return None


def mark_legacy_v1(service: JobService, job_id: str) -> None:
    stored = service._read_job(job_id)
    stored["ai_contract_version"] = "1.0"
    stored["ai_contract_source"] = "explicit_legacy_v1_test"
    service._write_job(stored)


class FakeDictionaryValidator:
    def __init__(self) -> None:
        self.calls = 0

    def validate(self, records, evidence):
        self.calls += 1
        return {
            "validation_version": "1.0",
            "mode": "validation_only",
            "status": "completed",
            "records": [{"行号": item.line_number} for item in records],
        }


class FakeMaterialMatcher:
    def __init__(self) -> None:
        self.calls = 0

    def match(self, records, resolved):
        self.calls += 1
        selections = {
            item.source_record_id: MaterialSelection(item.source_record_id, "MAT-D1", 0.75)
            for item in resolved
        }
        return MaterialMatchOutput(
            selections=selections,
            candidates_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "records": [{"行号": item.values["行号"], "candidates": []} for item in records],
            },
            summary_payload={
                "mode": "manual_review_only",
                "record_count": len(records),
                "accuracy_statement": "相似分数不是准确率，候选只用于人工复核。",
            },
        )


def settings(*, enabled: bool = True, retries: int = 0) -> LLMSettings:
    return LLMSettings(
        enabled=enabled,
        provider="volcengine_ark",
        model="doubao-d1-test",
        base_url="https://ark.test/api/v3",
        api_key=SECRET,
        max_retries=retries,
    )


def workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "PI"
    workbook.active.append(["No.", "Item", "Specification", "Qty"])
    workbook.active.append(["1", "Duvet Cover", "White cotton", "12"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def v2_desktop_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PI"
    sheet.append(["", "PROFORMA INVOICE", "", "", "Unit Price (USD)"])
    sheet.append(["BUYER:", "", "", "", ""])
    sheet.append(["Synthetic Hotel", "", "", "", "Contact Person: Aaron Lee"])
    sheet.append(["Delivery date:", "2026-09-30", "", "", ""])
    sheet.append(["No.", "Item", "Size", "Specification", "Qty"])
    sheet.append(["1", "Duvet Cover", "200*240", "100% cotton white", "12"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def response(payload: dict, *, status_code: int = 200) -> JSONHTTPResponse:
    body = {
        "id": "resp-d1-001",
        "model": "doubao-d1-test",
        "status": "completed",
        "output": [{
            "type": "function_call",
            "name": FULL_ORDER_EXTRACTION_FUNCTION,
            "arguments": json.dumps(payload, ensure_ascii=False),
        }],
        "usage": {"input_tokens": 13, "output_tokens": 7, "total_tokens": 20},
    }
    return JSONHTTPResponse(status_code, {"x-request-id": "header-d1"}, json.dumps(body, ensure_ascii=False).encode("utf-8"), 1)


def successful_transport(request: JSONHTTPRequest) -> JSONHTTPResponse:
    body = json.loads(request.body)
    source = json.loads(body["input"][1]["content"][0]["text"])
    return response(FakeFullOrderProvider().extract(source))


def test_full_order_provider_uses_strict_responses_contract_and_safe_metadata(tmp_path: Path) -> None:
    path = tmp_path / "fixture.xlsx"
    path.write_bytes(workbook_bytes())
    preprocessed = preprocess_workbook(path)
    request = preprocessed.to_request_dict()
    transport = RecordingTransport(successful_transport)
    provider = VolcengineArkFullOrderProvider(settings(), transport=transport, sleep=lambda _seconds: None)

    result = provider.extract(request)

    assert result["provider"] == "volcengine_ark"
    assert result["model"] == "doubao-d1-test"
    assert result["request_id"] == "resp-d1-001"
    assert result["usage"] == {"input_tokens": 13, "output_tokens": 7, "total_tokens": 20}
    assert result["attempt_count"] == 1
    assert provider.latest_telemetry.to_dict()["request_id"] == "resp-d1-001"
    assert provider.http_attempt_count == 1
    sent = json.loads(transport.requests[0].body)
    source = json.loads(sent["input"][1]["content"][0]["text"])
    assert sent["store"] is False
    assert sent["tools"][0]["strict"] is True
    assert sent["tools"][0]["parameters"]["additionalProperties"] is False
    assert sent["tool_choice"]["name"] == FULL_ORDER_EXTRACTION_FUNCTION
    assert source == request
    serialized = json.dumps(sent, ensure_ascii=False)
    assert SECRET not in serialized
    assert "job_id" not in serialized
    assert "C:\\" not in serialized


def test_v2_provider_uses_independent_function_schema_and_local_transport_metadata(
    tmp_path: Path,
) -> None:
    path = tmp_path / "fixture-v2.xlsx"
    path.write_bytes(workbook_bytes())
    unit = build_v2_extraction_units(preprocess_workbook(path))[0]
    request = build_v2_extraction_request(unit)

    def v2_transport(http_request: JSONHTTPRequest) -> JSONHTTPResponse:
        body = json.loads(http_request.body)
        assert body["store"] is False
        assert body.get("stream", False) is False
        assert body["tool_choice"]["name"] == FULL_ORDER_V2_EXTRACTION_FUNCTION
        assert body["tools"][0]["parameters"] == {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "candidates": body["tools"][0]["parameters"]["properties"]["candidates"]
            },
            "required": ["candidates"],
        }
        source = json.loads(body["input"][1]["content"][0]["text"])
        assert source == request
        response_body = {
            "id": "resp-v2-001",
            "model": "doubao-d1-test",
            "status": "completed",
            "output": [
                {
                    "type": "function_call",
                    "name": FULL_ORDER_V2_EXTRACTION_FUNCTION,
                    "arguments": json.dumps({"candidates": []}, ensure_ascii=False),
                }
            ],
            "usage": {"input_tokens": 5, "output_tokens": 2, "total_tokens": 7},
        }
        return JSONHTTPResponse(
            200,
            {"x-request-id": "header-v2"},
            json.dumps(response_body, ensure_ascii=False).encode("utf-8"),
            1,
        )

    transport = RecordingTransport(v2_transport)
    provider = VolcengineArkFullOrderProvider(
        settings(), transport=transport, sleep=lambda _seconds: None
    )

    output = provider.extract_v2(request)

    assert output == {"candidates": []}
    assert FULL_ORDER_V2_PROMPT_VERSION == "2.0"
    assert provider.latest_telemetry.request_id == "resp-v2-001"
    assert provider.latest_telemetry.total_tokens == 7
    assert provider.extraction_call_count == 1
    assert provider.http_attempt_count == 1
    assert "provider" not in output and "usage" not in output


def test_desktop_new_v2_job_uses_formal_provider_with_fake_transport(tmp_path: Path) -> None:
    def v2_transport(http_request: JSONHTTPRequest) -> JSONHTTPResponse:
        body = json.loads(http_request.body)
        assert body["tool_choice"]["name"] == FULL_ORDER_V2_EXTRACTION_FUNCTION
        response_body = {
            "id": "resp-v2-desktop",
            "model": "doubao-d1-test",
            "status": "completed",
            "output": [{
                "type": "function_call",
                "name": FULL_ORDER_V2_EXTRACTION_FUNCTION,
                "arguments": json.dumps({"candidates": []}, ensure_ascii=False),
            }],
            "usage": {"input_tokens": 8, "output_tokens": 2, "total_tokens": 10},
        }
        return JSONHTTPResponse(
            200,
            {"x-request-id": "header-v2-desktop"},
            json.dumps(response_body, ensure_ascii=False).encode("utf-8"),
            1,
        )

    transport = RecordingTransport(v2_transport)
    dictionary = FakeDictionaryValidator()
    matcher = FakeMaterialMatcher()
    service = JobService(
        tmp_path / "web-v2",
        store_path=tmp_path / "material-v2.sqlite3",
        index_dir=tmp_path / "index-v2",
        executor=DeferredExecutor(),
        ai_enhanced_settings=settings(),
        ai_enhanced_transport=transport,
        ai_enhanced_dictionary_validator=dictionary,
        ai_enhanced_material_matcher=matcher,
    )
    assert service.ai_enhanced_preflight()["ready"] is True
    assert transport.requests == []
    job = service.create_job(
        "v2-desktop.xlsx", v2_desktop_workbook_bytes(), parse_mode="ai_enhanced"
    )

    service._run_job(job["id"])
    completed = service.get_job(job["id"])

    assert completed["status"] == "completed"
    assert completed["ai_contract_version"] == "2.0"
    assert completed["ai_execution"]["contract_version"] == "2.0"
    assert completed["ai_execution"]["logical_call_count"] == 1
    assert completed["ai_execution"]["http_attempt_count"] == 1
    assert completed["ai_execution"]["token_summary"]["total_tokens"] == 10
    assert completed["has_complete_five_results"] is True
    assert len(transport.requests) == 1
    assert dictionary.calls == matcher.calls == 1


def test_missing_configuration_and_preflight_never_call_transport(tmp_path: Path) -> None:
    transport = RecordingTransport(successful_transport)
    dictionary = FakeDictionaryValidator()
    matcher = FakeMaterialMatcher()
    service = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=DeferredExecutor(),
        ai_enhanced_settings=settings(enabled=False),
        ai_enhanced_transport=transport,
        ai_enhanced_dictionary_validator=dictionary,
        ai_enhanced_material_matcher=matcher,
    )

    assert service.ai_enhanced_preflight()["ready"] is False
    assert transport.requests == []


def test_structure_recognition_uses_versioned_safe_context(tmp_path: Path) -> None:
    path = tmp_path / "ambiguous.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["No.", "Item", "Qty"])
    sheet.append(["1", "Duvet", "2"])
    sheet.append(["TOTAL", "", "2"])
    sheet.append([])
    sheet.append(["No.", "Item", "Qty"])
    sheet.append(["1", "Pillow", "3"])
    workbook.save(path)
    preprocessed = preprocess_workbook(path)
    manifest = build_structure_manifest(preprocessed, build_chunk_manifest(preprocessed))

    def layout_transport(request: JSONHTTPRequest) -> JSONHTTPResponse:
        sent = json.loads(request.body)
        context = json.loads(sent["input"][1]["content"][0]["text"])
        sheet_context = context["unresolved_sheets"][0]
        option = sheet_context["candidate_options"][0]
        body = {
            "id": "resp-layout",
            "status": "completed",
            "output": [{
                "type": "function_call",
                "name": FULL_ORDER_LAYOUT_FUNCTION,
                "arguments": json.dumps({
                    "layout_contract_version": "2.0",
                    "status": "resolved",
                    "decisions": [{
                        "sheet_id": sheet_context["sheet_id"],
                        "role": option["role"],
                        "candidate_id": option["candidate_id"],
                        "reason": "selected_local_order_candidate",
                    }],
                }),
            }],
        }
        return JSONHTTPResponse(200, {}, json.dumps(body).encode("utf-8"), 1)

    transport = RecordingTransport(layout_transport)
    provider = VolcengineArkFullOrderProvider(settings(), transport=transport, sleep=lambda _seconds: None)

    assert provider.resolve_structure(manifest)["status"] == "resolved"

    sent = json.loads(transport.requests[0].body)
    source = json.loads(sent["input"][1]["content"][0]["text"])
    assert source["layout_contract_version"] == "2.0"
    assert source["unresolved_sheets"]
    assert sent["tool_choice"]["name"] == FULL_ORDER_LAYOUT_FUNCTION


def test_retryable_ark_response_is_bounded_and_counted(tmp_path: Path) -> None:
    path = tmp_path / "fixture.xlsx"
    path.write_bytes(workbook_bytes())
    request = preprocess_workbook(path).to_request_dict()
    rate_limited = JSONHTTPResponse(
        429, {"retry-after": "0"}, b'{"error":{"type":"rate_limit"}}', 1
    )
    calls = iter([rate_limited, successful_transport])

    def retry_transport(http_request: JSONHTTPRequest) -> JSONHTTPResponse:
        next_item = next(calls)
        return next_item(http_request) if callable(next_item) else next_item

    provider = VolcengineArkFullOrderProvider(
        settings(retries=1), transport=RecordingTransport(retry_transport), sleep=lambda _seconds: None
    )

    assert provider.extract(request)["attempt_count"] == 2
    assert provider.http_attempt_count == 2


def test_output_text_compatibility_keeps_the_same_strict_contract(tmp_path: Path) -> None:
    path = tmp_path / "fixture.xlsx"
    path.write_bytes(workbook_bytes())
    request = preprocess_workbook(path).to_request_dict()

    def output_text_transport(_request: JSONHTTPRequest) -> JSONHTTPResponse:
        payload = FakeFullOrderProvider().extract(request)
        body = {
            "id": "resp-d1-text",
            "status": "completed",
            "output_text": json.dumps(payload, ensure_ascii=False),
            "usage": {"input_tokens": "2", "output_tokens": "3"},
        }
        return JSONHTTPResponse(200, {}, json.dumps(body, ensure_ascii=False).encode("utf-8"), 1)

    result = VolcengineArkFullOrderProvider(
        settings(), transport=RecordingTransport(output_text_transport), sleep=lambda _seconds: None
    ).extract(request)

    assert result["request_id"] == "resp-d1-text"
    assert result["usage"]["total_tokens"] == 5


def test_strict_contract_rejects_model_extra_fields(tmp_path: Path) -> None:
    path = tmp_path / "fixture.xlsx"
    path.write_bytes(workbook_bytes())
    request = preprocess_workbook(path).to_request_dict()

    def invalid_transport(_request: JSONHTTPRequest) -> JSONHTTPResponse:
        payload = FakeFullOrderProvider().extract(request)
        payload["records"][0]["fields"]["物料编码"] = {}
        return response(payload)

    provider = VolcengineArkFullOrderProvider(
        settings(), transport=RecordingTransport(invalid_transport), sleep=lambda _seconds: None
    )

    with pytest.raises(FullOrderContractError, match="extra fields"):
        provider.extract(request)

    assert provider.latest_contract_diagnostic == {
        "stage": "forbidden_fields",
        "category": "extra_fields",
        "path": "$.records[].fields",
        "forbidden_fields": ["物料编码"],
        "extra_field_count": 1,
    }


def test_provider_reports_safe_request_and_response_parse_diagnostics(tmp_path: Path) -> None:
    provider = VolcengineArkFullOrderProvider(
        settings(), transport=RecordingTransport(successful_transport), sleep=lambda _seconds: None
    )
    with pytest.raises(FullOrderContractError):
        provider.extract({"untrusted_request_key": "must-not-appear"})
    assert provider.latest_contract_diagnostic["stage"] == "request_validation"
    assert "untrusted_request_key" not in json.dumps(provider.latest_contract_diagnostic, ensure_ascii=False)

    path = tmp_path / "fixture.xlsx"
    path.write_bytes(workbook_bytes())
    request = preprocess_workbook(path).to_request_dict()

    def malformed_transport(_request: JSONHTTPRequest) -> JSONHTTPResponse:
        body = {"id": "resp-d2d-malformed", "status": "completed", "usage": {}}
        return JSONHTTPResponse(200, {}, json.dumps(body).encode("utf-8"), 1)

    provider = VolcengineArkFullOrderProvider(
        settings(), transport=RecordingTransport(malformed_transport), sleep=lambda _seconds: None
    )
    with pytest.raises(FullOrderContractError, match="valid strict JSON"):
        provider.extract(request)
    assert provider.latest_contract_diagnostic == {
        "stage": "response_parsing",
        "category": "response_parse",
    }


def test_desktop_constructs_ready_ark_provider_and_runs_offline_success(tmp_path: Path) -> None:
    transport = RecordingTransport(successful_transport)
    dictionary = FakeDictionaryValidator()
    matcher = FakeMaterialMatcher()
    service = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=DeferredExecutor(),
        ai_enhanced_settings=settings(),
        ai_enhanced_transport=transport,
        ai_enhanced_dictionary_validator=dictionary,
        ai_enhanced_material_matcher=matcher,
    )
    assert service.ai_enhanced_preflight()["ready"] is True
    assert transport.requests == []
    job = service.create_job("offline.xlsx", workbook_bytes(), parse_mode="ai_enhanced")
    mark_legacy_v1(service, job["id"])

    service._run_job(job["id"])

    completed = service.get_job(job["id"])
    assert completed["status"] == "completed"
    assert completed["has_complete_five_results"] is True
    assert completed["ai_execution"]["provider"] == "volcengine_ark"
    assert completed["ai_execution"]["request_id"] == "resp-d1-001"
    assert completed["ai_execution"]["http_attempt_count"] == 1
    assert completed["ai_execution"]["token_summary"]["total_tokens"] == 20
    assert len(transport.requests) == 1
    assert dictionary.calls == matcher.calls == 1


def test_desktop_ark_deterministic_failure_waits_without_publication(tmp_path: Path) -> None:
    def bad_request(_request: JSONHTTPRequest) -> JSONHTTPResponse:
        return JSONHTTPResponse(400, {"x-request-id": "bad-d1"}, b'{"error":{"type":"invalid_request"}}', 1)

    transport = RecordingTransport(bad_request)
    service = JobService(
        tmp_path / "web",
        store_path=tmp_path / "material.sqlite3",
        index_dir=tmp_path / "index",
        executor=DeferredExecutor(),
        ai_enhanced_settings=settings(),
        ai_enhanced_transport=transport,
        ai_enhanced_dictionary_validator=FakeDictionaryValidator(),
        ai_enhanced_material_matcher=FakeMaterialMatcher(),
    )
    job = service.create_job("failure.xlsx", workbook_bytes(), parse_mode="ai_enhanced")
    mark_legacy_v1(service, job["id"])

    service._run_job(job["id"])

    failed = service.get_job(job["id"])
    assert failed["status"] == "awaiting_user_decision"
    assert failed["has_complete_five_results"] is False
    assert failed["effective_parse_mode"] == "ai_enhanced"
    assert failed["ai_execution"]["safe_error_code"] == "AI_SCHEMA_OR_EVIDENCE_FAILED"
    assert failed["ai_execution"]["logical_call_count"] == 1
    assert failed["ai_execution"]["http_attempt_count"] == 1
    assert len(transport.requests) == 1
