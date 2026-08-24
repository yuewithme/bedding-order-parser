"""Read existing Excel dictionaries into preview dataclasses."""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from bedding_order_parser.exceptions import BeddingOrderParserError
from bedding_order_parser.dictionaries.models import (
    DictionaryBundle,
    DictionarySource,
    FabricRow,
    RuleRow,
    StyleRow,
)


APPROVED_RULES_SHA256 = (
    "8d527595f671b63762a15b1f5aa89004df4e773f68e776c824c37d57dece3c7c"
)
APPROVED_STYLES_SHA256 = (
    "75faab06a151ee8f9d6d9dcb28ca4679414f4008fb86ae5d88acf5d0ee60660c"
)

RULE_SHEET_NAME = "被套 提取规则"
FABRIC_SHEET_NAME = "面料类价格表"
STYLE_SHEET_NAME = "Sheet1"

RULE_HEADERS = ["字段名", "可能值", "关键描述", "默认值规则", "补充说明"]
FABRIC_HEADERS = ["面料品类", "面料", "颜色", "涤棉成份", "密度"]
STYLE_HEADERS = [
    "被套款式",
    "飞边 (Flange)",
    "系带 (Tie)",
    "拉链 (Zipper)",
    "是否口袋式 (Has Pocket)",
    "是否迎宾式 (Is Welcome Style)",
    "其他款式结构 (Other Structure)",
    "备注尺寸 (Dimensions)",
]


class DictionaryLoadError(BeddingOrderParserError):
    """Raised when a dictionary workbook cannot be loaded as approved."""


@dataclass(frozen=True)
class _CellValue:
    value: str
    source_cell: str


def compute_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    return "\n".join(lines).strip()


def load_dictionary_bundle(
    rules_path: Path,
    styles_path: Path,
    *,
    expected_rules_sha256: str = APPROVED_RULES_SHA256,
    expected_styles_sha256: str = APPROVED_STYLES_SHA256,
) -> DictionaryBundle:
    rules_resolved = _validate_dictionary_path(rules_path)
    styles_resolved = _validate_dictionary_path(styles_path)
    rules_sha256 = _validate_sha256(rules_resolved, expected_rules_sha256)
    styles_sha256 = _validate_sha256(styles_resolved, expected_styles_sha256)

    rules_workbook = _load_workbook(rules_resolved)
    styles_workbook = _load_workbook(styles_resolved)
    try:
        rule_sheet = _require_sheet(rules_workbook, RULE_SHEET_NAME, rules_resolved)
        fabric_sheet = _require_sheet(rules_workbook, FABRIC_SHEET_NAME, rules_resolved)
        style_sheet = _require_sheet(styles_workbook, STYLE_SHEET_NAME, styles_resolved)

        _validate_headers(rule_sheet, RULE_HEADERS, RULE_SHEET_NAME)
        _validate_headers(fabric_sheet, FABRIC_HEADERS, FABRIC_SHEET_NAME)
        _validate_headers(style_sheet, STYLE_HEADERS, STYLE_SHEET_NAME)

        rules = _read_rule_rows(rule_sheet)
        fabrics = _read_fabric_rows(fabric_sheet)
        styles = _read_style_rows(style_sheet)
    finally:
        rules_workbook.close()
        styles_workbook.close()

    sources = [
        DictionarySource(rules_resolved.name, rules_sha256, RULE_SHEET_NAME),
        DictionarySource(rules_resolved.name, rules_sha256, FABRIC_SHEET_NAME),
        DictionarySource(styles_resolved.name, styles_sha256, STYLE_SHEET_NAME),
    ]
    summary = {
        "rule_rows": len(rules),
        "fabric_rows": len(fabrics),
        "style_rows": len(styles),
    }
    return DictionaryBundle(
        sources=sources,
        rules=rules,
        fabrics=fabrics,
        styles=styles,
        summary=summary,
    )


def _validate_dictionary_path(path: Path) -> Path:
    resolved = path.expanduser().resolve()
    if not resolved.exists():
        raise DictionaryLoadError(f"Dictionary file does not exist: {path}")
    if not resolved.is_file():
        raise DictionaryLoadError(f"Dictionary path is not a file: {path}")
    if resolved.suffix.lower() != ".xlsx":
        raise DictionaryLoadError(f"Dictionary file must be .xlsx: {resolved.name}")
    return resolved


def _validate_sha256(path: Path, expected_sha256: str) -> str:
    actual_sha256 = compute_sha256(path)
    if actual_sha256.casefold() != expected_sha256.casefold():
        raise DictionaryLoadError(
            "Dictionary SHA-256 mismatch for "
            f"{path.name}: expected {expected_sha256}, actual {actual_sha256}"
        )
    return actual_sha256


def _load_workbook(path: Path):
    try:
        return load_workbook(
            path,
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl raises several reader-specific errors.
        raise DictionaryLoadError(f"Failed to open dictionary workbook: {path.name}") from exc


def _require_sheet(workbook, sheet_name: str, path: Path):
    if sheet_name not in workbook.sheetnames:
        raise DictionaryLoadError(f"Missing sheet {sheet_name!r} in {path.name}")
    return workbook[sheet_name]


def _validate_headers(worksheet, expected_headers: list[str], sheet_name: str) -> None:
    actual_headers = [
        _cell_text(worksheet, 1, column_index).value
        for column_index in range(1, len(expected_headers) + 1)
    ]
    mismatches = [
        (expected, actual)
        for expected, actual in zip(expected_headers, actual_headers, strict=True)
        if expected.casefold() not in actual.casefold()
    ]
    if mismatches:
        raise DictionaryLoadError(
            f"Unexpected headers in {sheet_name}: expected {expected_headers}, "
            f"actual {actual_headers}"
        )


def _read_rule_rows(worksheet) -> list[RuleRow]:
    rows = []
    for row_index in range(2, 37):
        values = _row_values(worksheet, row_index, RULE_HEADERS)
        rows.append(
            RuleRow(
                source_row=row_index,
                source_cells={header: cell.source_cell for header, cell in values.items()},
                field_name=values[RULE_HEADERS[0]].value,
                standard_value=values[RULE_HEADERS[1]].value,
                rule_description=values[RULE_HEADERS[2]].value,
                default_rule=values[RULE_HEADERS[3]].value,
                notes=values[RULE_HEADERS[4]].value,
                raw_values={header: cell.value for header, cell in values.items()},
            )
        )
    return rows


def _read_fabric_rows(worksheet) -> list[FabricRow]:
    rows = []
    for row_index in range(2, 77):
        values = _row_values(worksheet, row_index, FABRIC_HEADERS)
        rows.append(
            FabricRow(
                source_row=row_index,
                fabric_family=values[FABRIC_HEADERS[0]].value,
                fabric_standard=values[FABRIC_HEADERS[1]].value,
                color_standard=values[FABRIC_HEADERS[2]].value,
                composition_raw=values[FABRIC_HEADERS[3]].value,
                density=values[FABRIC_HEADERS[4]].value,
                raw_values={header: cell.value for header, cell in values.items()},
            )
        )
    return rows


def _read_style_rows(worksheet) -> list[StyleRow]:
    rows = []
    for row_index in range(2, 107):
        values = _row_values(worksheet, row_index, STYLE_HEADERS)
        rows.append(
            StyleRow(
                source_row=row_index,
                standard_name=values[STYLE_HEADERS[0]].value,
                flange=values[STYLE_HEADERS[1]].value,
                tie=values[STYLE_HEADERS[2]].value,
                zipper=values[STYLE_HEADERS[3]].value,
                has_pocket=values[STYLE_HEADERS[4]].value,
                is_welcome_style=values[STYLE_HEADERS[5]].value,
                other_structure=values[STYLE_HEADERS[6]].value,
                dimensions=values[STYLE_HEADERS[7]].value,
                raw_values={header: cell.value for header, cell in values.items()},
            )
        )
    return rows


def _row_values(worksheet, row_index: int, headers: list[str]) -> dict[str, _CellValue]:
    return {
        header: _cell_text(worksheet, row_index, column_index)
        for column_index, header in enumerate(headers, start=1)
    }


def _cell_text(worksheet, row_index: int, column_index: int) -> _CellValue:
    cell = worksheet.cell(row=row_index, column=column_index)
    if cell.value is not None:
        return _CellValue(normalize_text(cell.value), cell.coordinate)
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            source = worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            )
            return _CellValue(normalize_text(source.value), source.coordinate)
    return _CellValue("", f"{get_column_letter(column_index)}{row_index}")


def bounded_nonempty_bounds(
    worksheet,
    *,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
) -> dict[str, int]:
    rows: list[int] = []
    columns: list[int] = []
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_column, max_column + 1):
            if _cell_text(worksheet, row_index, column_index).value:
                rows.append(row_index)
                columns.append(column_index)
    if not rows:
        return {}
    return {
        "min_row": min(rows),
        "max_row": max(rows),
        "min_column": min(columns),
        "max_column": max(columns),
    }
