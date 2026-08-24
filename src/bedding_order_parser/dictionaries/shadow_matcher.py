"""Independent dictionary shadow comparison against official Gate 2D output."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from bedding_order_parser.dictionaries.loader import compute_sha256, load_dictionary_bundle
from bedding_order_parser.dictionaries.models import DictionaryBundle, FabricRow, StyleRow
from bedding_order_parser.dictionaries.shadow_models import (
    SHADOW_FIELDS,
    SHADOW_STATUSES,
    ShadowFieldComparison,
    ShadowFileReport,
    ShadowRecord,
    ShadowReport,
    ShadowStatus,
)
from bedding_order_parser.exceptions import BeddingOrderParserError


class ShadowCompareError(BeddingOrderParserError):
    """Raised when the independent shadow comparison cannot be built."""


@dataclass(frozen=True)
class _Evidence:
    source_text: str
    source_cells: list[str]
    python_value: str
    python_status: str


@dataclass(frozen=True)
class _SizeParse:
    raw_first_dimension: float
    raw_second_dimension: float
    normalized_length: float
    normalized_width: float
    unit: str
    structural_extension_cm: float

    def candidate_text(self) -> str:
        suffix = (
            f"+{_format_number(self.structural_extension_cm)}cm"
            if self.structural_extension_cm
            else ""
        )
        return (
            f"{_format_number(self.normalized_length)}*"
            f"{_format_number(self.normalized_width)}{suffix}"
        )


def build_shadow_report(
    *,
    input_dir: Path,
    results_dir: Path,
    reports_dir: Path,
    rules_path: Path,
    styles_path: Path,
) -> ShadowReport:
    """Build a read-only shadow report without modifying official outputs."""

    input_resolved = _require_dir(input_dir, "input directory")
    results_resolved = _require_dir(results_dir, "results directory")
    reports_resolved = _require_dir(reports_dir, "reports directory")
    bundle = load_dictionary_bundle(rules_path, styles_path)

    source_files = {path.name: path for path in input_resolved.glob("*.xlsx")}
    file_reports: list[ShadowFileReport] = []

    for parse_report_path in sorted(reports_resolved.glob("*.json")):
        parse_report = _read_json(parse_report_path)
        input_name = str(parse_report.get("input", {}).get("file_name", ""))
        if input_name not in source_files:
            raise ShadowCompareError(
                f"Parse report input has no matching source workbook: {input_name}"
            )
        source_path = source_files[input_name]
        result_path = _resolve_result_path(parse_report, results_resolved)
        official_records = _read_json(result_path)
        parse_records = parse_report.get("records", [])
        if len(official_records) != len(parse_records):
            raise ShadowCompareError(
                f"Record count mismatch for {input_name}: "
                f"{len(official_records)} official vs {len(parse_records)} report"
            )

        workbook = load_workbook(
            source_path,
            read_only=False,
            data_only=True,
            keep_links=False,
        )
        try:
            records: list[ShadowRecord] = []
            for official_record, parse_record in zip(
                official_records,
                parse_records,
                strict=True,
            ):
                field_results = {
                    field_name: compare_shadow_field(
                        bundle,
                        field_name,
                        _evidence_for_field(
                            workbook,
                            parse_record,
                            official_record,
                            field_name,
                        ),
                    )
                    for field_name in SHADOW_FIELDS
                }
                records.append(
                    ShadowRecord(
                        line_number=str(official_record.get("行号", "")),
                        fields=field_results,
                    )
                )
        finally:
            workbook.close()

        file_reports.append(
            ShadowFileReport(
                source_file=source_path.name,
                source_sha256=compute_sha256(source_path),
                result_json=result_path.name,
                result_json_sha256=compute_sha256(result_path),
                parse_report_json=parse_report_path.name,
                parse_report_sha256=compute_sha256(parse_report_path),
                records=records,
            )
        )

    return ShadowReport(summary=_build_summary(file_reports), files=file_reports)


def compare_shadow_field(
    bundle: DictionaryBundle,
    field_name: str,
    evidence: _Evidence,
) -> ShadowFieldComparison:
    text = _normalize_text(evidence.source_text)
    python_value = _normalize_text(evidence.python_value)
    if not text:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "source_not_provided",
            reason=f"No source evidence provided; Python status is {evidence.python_status}.",
        )

    if field_name == "币种":
        return _compare_currency(field_name, evidence)
    if field_name == "物料名称":
        return _compare_product_name(bundle, field_name, evidence)
    if field_name == "规格":
        return _compare_size(field_name, evidence)
    if field_name == "颜色":
        return _compare_color(bundle, field_name, evidence)
    if field_name == "面料":
        return _compare_fabric(bundle, field_name, evidence)
    if field_name == "面料-涤棉成分":
        return _compare_composition(field_name, evidence)
    if field_name == "款式":
        return _compare_style(bundle, field_name, evidence)
    if field_name == "尺寸类型":
        return _compare_size_type(field_name, evidence)
    if field_name == "行备注":
        return _compare_line_note(field_name, evidence)
    if field_name == "是否绣花":
        return _compare_embroidery(field_name, evidence)

    return _comparison(
        field_name,
        evidence,
        [],
        [],
        "dictionary_no_match",
        reason="Unsupported shadow field.",
    )


def _compare_currency(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    lowered = _audit_text(evidence.source_text)
    rules = {
        "美元": [r"\busd\b", r"\bus\$", r"\$"],
        "人民币": [r"\brmb\b", r"\bcny\b", "人民币", "¥"],
        "日元": [r"\bjpy\b", r"\byen\b", "円"],
        "欧元": [r"\beur\b", "€"],
    }
    candidates = [
        standard
        for standard, patterns in rules.items()
        if any(re.search(pattern, lowered, flags=re.IGNORECASE) for pattern in patterns)
    ]
    return _single_candidate_result(
        field_name,
        evidence,
        candidates,
        [f"currency.{candidate}" for candidate in candidates],
    )


def _compare_product_name(
    bundle: DictionaryBundle,
    field_name: str,
    evidence: _Evidence,
) -> ShadowFieldComparison:
    lowered = _audit_text(evidence.source_text)
    candidates = []
    matched_rules = []
    for row in bundle.rules:
        row_text = _audit_text(" ".join(row.raw_values.values()))
        if "duvet cover" in lowered and "duvet cover" in row_text:
            candidates.append(row.standard_value or "被套")
            matched_rules.append(f"PI单提取规则.xlsx:{row.source_row}")
    if not candidates and re.search(r"duvet\s*cover|quilt\s*cover|被套", lowered):
        candidates = ["被套"]
        matched_rules = ["product_category.duvet_cover"]
    if candidates and any(candidate and candidate in evidence.python_value for candidate in candidates):
        return _comparison(
            field_name,
            evidence,
            _unique(candidates),
            matched_rules,
            "equivalent_match",
            matched_components=["product_category"],
            reason="Source product category is compatible with official material name.",
        )
    return _single_candidate_result(field_name, evidence, candidates, matched_rules)


def _compare_size(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    source_size = _parse_size(evidence.source_text, raw_order="width_length")
    if not source_size:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "dictionary_no_match",
            missing_components=["length", "width"],
            reason="No parseable size expression was found in source evidence.",
        )
    official_size = _parse_size(evidence.python_value, raw_order="length_width")
    candidate = source_size.candidate_text()
    official_candidate = official_size.candidate_text() if official_size else ""
    candidates = [candidate]
    if candidate == official_candidate or candidate == _normalize_size_literal(evidence.python_value):
        status: ShadowStatus = (
            "exact_match"
            if _size_text_exactly_matches(evidence.source_text, evidence.python_value)
            else "equivalent_match"
        )
        return _comparison(
            field_name,
            evidence,
            candidates,
            ["size.unit_order_and_structural_extension"],
            status,
            matched_components=["length", "width", "structural_extension_cm"],
            reason="Source size converts to the official normalized size.",
        )

    if (
        official_size
        and _same_number(source_size.normalized_length, official_size.normalized_length)
        and _same_number(source_size.normalized_width, official_size.normalized_width)
        and official_size.structural_extension_cm > 0
        and source_size.structural_extension_cm != official_size.structural_extension_cm
    ):
        missing = (
            ["structural_extension_cm"]
            if source_size.structural_extension_cm == 0
            else ["structural_extension_cm_ambiguous"]
        )
        return _comparison(
            field_name,
            evidence,
            candidates,
            ["size.unit_order_and_structural_extension"],
            "partial_match",
            matched_components=["length", "width"],
            missing_components=missing,
            reason=(
                "Source dimensions match the official size, but structural extension "
                "evidence was not provided, not safely extractable, or ambiguous."
            ),
        )
    return _comparison(
        field_name,
        evidence,
        candidates,
        ["size.unit_order_and_structural_extension"],
        "conflict",
        conflicting_components=["dimensions"],
        reason="Source size does not match official normalized size.",
    )


def _compare_color(
    bundle: DictionaryBundle,
    field_name: str,
    evidence: _Evidence,
) -> ShadowFieldComparison:
    candidates = _color_candidates(evidence.source_text)
    matched_rules = [f"color.{candidate}" for candidate in candidates]
    for fabric in bundle.fabrics:
        standard = _standard_color(fabric.color_standard)
        if standard and standard in candidates:
            matched_rules.append(f"面料类价格表:{fabric.source_row}")
    return _single_candidate_result(field_name, evidence, candidates, matched_rules)


def _compare_fabric(
    bundle: DictionaryBundle,
    field_name: str,
    evidence: _Evidence,
) -> ShadowFieldComparison:
    components = _fabric_components(evidence.source_text)
    if not components:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "dictionary_no_match",
            reason="No fabric category, density, or composition evidence was found.",
        )

    matched = _matching_fabric_rows(bundle.fabrics, components)
    detailed_candidates = _unique(
        row.fabric_standard for row in matched if row.fabric_standard
    )
    candidates = _unique(
        _fabric_projection(_fabric_row_components(row))
        for row in matched
        if _fabric_projection(_fabric_row_components(row))
    )
    matched_rules = [f"面料类价格表:{row.source_row}" for row in matched]
    official_components = _fabric_components(evidence.python_value)
    comparable_keys = ("category", "density", "composition")
    matched_components = [
        key
        for key, value in components.items()
        if key in comparable_keys
        and value
        and value == official_components.get(key)
    ]
    conflicting_components = [
        key
        for key, value in components.items()
        if key in comparable_keys
        and value
        and official_components.get(key)
        and value != official_components.get(key)
    ]
    if conflicting_components:
        return _comparison(
            field_name,
            evidence,
            candidates or [_fabric_projection(components)],
            matched_rules or ["fabric.component_rules"],
            "conflict",
            detailed_candidates=detailed_candidates,
            matched_components=matched_components,
            conflicting_components=conflicting_components,
            reason="Explicit source fabric components conflict with official output.",
        )
    if not matched:
        return _comparison(
            field_name,
            evidence,
            [_fabric_projection(components)],
            ["fabric.component_rules"],
            "dictionary_no_match",
            matched_components=matched_components,
            missing_components=[
                key
                for key in comparable_keys
                if components.get(key) and key not in matched_components
            ],
            reason="Source fabric evidence was parsed but no dictionary row matched it.",
        )
    if len(candidates) > 1:
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "ambiguous",
            detailed_candidates=detailed_candidates,
            matched_components=matched_components,
            reason="Multiple business-distinct fabric projections remain after filtering.",
        )
    candidate_components = _fabric_components(candidates[0]) if candidates else {}
    candidate_conflicts = [
        key
        for key in comparable_keys
        if candidate_components.get(key)
        and official_components.get(key)
        and candidate_components[key] != official_components[key]
    ]
    if candidate_conflicts:
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "conflict",
            detailed_candidates=detailed_candidates,
            matched_components=matched_components,
            conflicting_components=candidate_conflicts,
            reason="The unique dictionary projection conflicts with official fabric components.",
        )
    comparable_source = {
        key: value for key, value in components.items() if key in comparable_keys
    }
    comparable_official = {
        key: value for key, value in official_components.items() if key in comparable_keys
    }
    if comparable_source and all(
        comparable_official.get(key) in (None, value)
        for key, value in comparable_source.items()
    ):
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "equivalent_match",
            detailed_candidates=detailed_candidates,
            matched_components=sorted(
                key
                for key in comparable_source
                if comparable_official.get(key) == comparable_source[key]
            ),
            reason=(
                "The source projection is compatible with official output; detailed "
                "dictionary rows remain available without creating false ambiguity."
            ),
        )
    if matched_components:
        return _comparison(
            field_name,
            evidence,
            candidates or [_compose_fabric_candidate(components)],
            matched_rules or ["fabric.component_rules"],
            "partial_match",
            detailed_candidates=detailed_candidates,
            matched_components=matched_components,
            missing_components=[
                key for key in comparable_source if key not in matched_components
            ],
            reason="Only part of the source fabric evidence is present in the official value.",
        )
    return _comparison(
        field_name,
        evidence,
        candidates,
        matched_rules,
        "partial_match",
        detailed_candidates=detailed_candidates,
        missing_components=list(comparable_source),
        reason="A unique dictionary projection exists, but official evidence is insufficient.",
    )


def _compare_composition(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    candidate = _composition_candidate(evidence.source_text)
    if not candidate:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "dictionary_no_match",
            reason="No composition expression was found in source evidence.",
        )
    official = _composition_candidate(evidence.python_value)
    if candidate == official:
        return _comparison(
            field_name,
            evidence,
            [_composition_label(candidate)],
            ["composition.normalization"],
            "equivalent_match",
            matched_components=["composition"],
            reason="Source composition is equivalent to the official normalized value.",
        )
    return _comparison(
        field_name,
        evidence,
        [_composition_label(candidate)],
        ["composition.normalization"],
        "conflict",
        conflicting_components=["composition"],
        reason="Source composition conflicts with official normalized value.",
    )


def _compare_style(
    bundle: DictionaryBundle,
    field_name: str,
    evidence: _Evidence,
) -> ShadowFieldComparison:
    components = _style_components(evidence.source_text)
    if not components:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "source_not_provided",
            reason="The source row does not provide a supported style component.",
        )
    matched = _matching_style_rows(bundle.styles, components)
    canonical_style = _style_projection_label(components)
    canonical_rows = [
        row for row in matched if row.standard_name == canonical_style
    ]
    if canonical_rows:
        matched = canonical_rows
    candidates = _unique(row.standard_name for row in matched if row.standard_name)
    matched_rules = [f"款式表_structured.xlsx:{row.source_row}" for row in matched]
    if len(candidates) > 1:
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "ambiguous",
            matched_components=sorted(components),
            reason="Multiple business-distinct styles fully match the source components.",
        )
    if candidates and candidates[0] == evidence.python_value:
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "equivalent_match",
            matched_components=sorted(components),
            reason="Source style components map to the same standard style as official output.",
        )
    if candidates and _style_candidate_compatible(candidates[0], evidence.python_value):
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "equivalent_match",
            matched_components=sorted(components),
            reason="Dictionary style features are compatible with official output.",
        )
    if candidates:
        source_standard = _style_standard_components(evidence.python_value)
        shared_keys = set(components) & set(source_standard)
        conflicting = [
            key
            for key in sorted(shared_keys)
            if components[key] != source_standard[key]
        ]
        if not evidence.python_value or not conflicting:
            return _comparison(
                field_name,
                evidence,
                candidates,
                matched_rules,
                "partial_match",
                matched_components=sorted(components),
                missing_components=["official_standard_style"],
                reason=(
                    "Dictionary style components match source evidence, but the official "
                    "style is empty or not directly comparable."
                ),
            )
        return _comparison(
            field_name,
            evidence,
            candidates,
            matched_rules,
            "conflict",
            matched_components=sorted(components),
            conflicting_components=conflicting,
            reason="Explicit style components conflict with official output.",
        )
    if not _style_evidence_sufficient(components):
        return _comparison(
            field_name,
            evidence,
            [_style_label(components)],
            ["style.feature_rules"],
            "partial_match",
            matched_components=sorted(components),
            missing_components=["style_form"],
            reason="Only partial style evidence is present; no unique style is inferred.",
        )
    return _comparison(
        field_name,
        evidence,
        [_style_label(components)],
        ["style.feature_rules"],
        "dictionary_no_match",
        matched_components=sorted(components),
        reason="Style evidence was parsed but no dictionary row matched it.",
    )


def _compare_size_type(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    lowered = _audit_text(evidence.source_text)
    candidates = []
    if "after wash" in lowered or "washed size" in lowered or "洗涤" in lowered:
        candidates.append("洗涤尺寸")
    if "before wash" in lowered or "delivery size" in lowered or "finished size" in lowered:
        candidates.append("交货尺寸")
    return _single_candidate_result(field_name, evidence, candidates, ["size_type.keywords"])


def _compare_line_note(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    lowered = _audit_text(evidence.source_text)
    candidates: list[str] = []
    if "hand hole" in lowered or "手洞" in lowered:
        candidates.append("含手洞")
    if "overlap" in lowered or "重叠" in lowered:
        candidates.append("重叠片")
    if "piping" in lowered or "牙条" in lowered:
        candidates.append("牙条")
    if "embroidery" in lowered or "绣" in lowered:
        candidates.append("绣花")
    if not candidates:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "dictionary_no_match",
            reason="No supported craft note keyword was found.",
        )
    matched = [candidate for candidate in candidates if candidate in evidence.python_value]
    missing = [candidate for candidate in candidates if candidate not in evidence.python_value]
    if missing and matched:
        status: ShadowStatus = "partial_match"
    elif missing:
        status = "conflict"
    else:
        status = "exact_match" if evidence.python_value in candidates else "partial_match"
    return _comparison(
        field_name,
        evidence,
        candidates,
        ["line_note.craft_keywords"],
        status,
        matched_components=matched,
        missing_components=missing,
        conflicting_components=missing if status == "conflict" else [],
        reason="Craft note keywords were compared against official row note.",
    )


def _compare_embroidery(field_name: str, evidence: _Evidence) -> ShadowFieldComparison:
    lowered = _audit_text(evidence.source_text)
    candidates = []
    if re.search(r"no\s+embroider|without\s+embroider|无绣|不绣", lowered):
        candidates.append("N")
    elif "embroidery" in lowered or "embroider" in lowered or "绣" in lowered:
        candidates.append("Y")
    return _single_candidate_result(field_name, evidence, candidates, ["embroidery.keywords"])


def _single_candidate_result(
    field_name: str,
    evidence: _Evidence,
    candidates: list[str],
    matched_rules: list[str],
) -> ShadowFieldComparison:
    unique_candidates = _unique(candidate for candidate in candidates if candidate)
    if not unique_candidates:
        return _comparison(
            field_name,
            evidence,
            [],
            [],
            "dictionary_no_match",
            reason="No dictionary candidate matched source evidence.",
        )
    if len(unique_candidates) > 1:
        return _comparison(
            field_name,
            evidence,
            unique_candidates,
            matched_rules,
            "ambiguous",
            reason="Multiple dictionary candidates matched source evidence.",
        )
    candidate = unique_candidates[0]
    if candidate == evidence.python_value:
        return _comparison(
            field_name,
            evidence,
            unique_candidates,
            matched_rules,
            "exact_match",
            matched_components=[field_name],
            reason="Dictionary candidate exactly matches official output.",
        )
    if _audit_text(candidate) == _audit_text(evidence.python_value):
        return _comparison(
            field_name,
            evidence,
            unique_candidates,
            matched_rules,
            "equivalent_match",
            matched_components=[field_name],
            reason="Dictionary candidate is text-equivalent to official output.",
        )
    return _comparison(
        field_name,
        evidence,
        unique_candidates,
        matched_rules,
        "conflict",
        conflicting_components=[field_name],
        reason="Dictionary candidate conflicts with official output.",
    )


def _comparison(
    field_name: str,
    evidence: _Evidence,
    candidates: list[str],
    matched_rules: list[str],
    status: ShadowStatus,
    *,
    detailed_candidates: list[str] | None = None,
    matched_components: list[str] | None = None,
    missing_components: list[str] | None = None,
    conflicting_components: list[str] | None = None,
    reason: str = "",
) -> ShadowFieldComparison:
    return ShadowFieldComparison(
        field_name=field_name,
        source_text=evidence.source_text,
        source_cells=evidence.source_cells,
        python_value=evidence.python_value,
        python_status=evidence.python_status,
        dictionary_candidates=_unique(candidates),
        matched_rules=_unique(matched_rules),
        comparison_status=status,
        detailed_candidates=_unique(detailed_candidates or []),
        matched_components=matched_components or [],
        missing_components=missing_components or [],
        conflicting_components=conflicting_components or [],
        reason=reason,
    )


def _evidence_for_field(workbook, parse_record: dict[str, Any], official_record: dict[str, Any], field: str) -> _Evidence:
    diagnostic = parse_record.get("fields", {}).get(field, {})
    source = diagnostic.get("source", {}) if isinstance(diagnostic, dict) else {}
    sheet_name = source.get("sheet") or workbook.sheetnames[0]
    cells = [cell for cell in source.get("cells", []) if isinstance(cell, str)]
    if field == "规格":
        cells = _size_source_cells_with_same_row_context(parse_record, cells)
    source_text = _read_source_text(workbook, sheet_name, cells)
    return _Evidence(
        source_text=source_text,
        source_cells=cells,
        python_value=_normalize_text(official_record.get(field, "")),
        python_status=_normalize_text(diagnostic.get("status", "")),
    )


def _read_source_text(workbook, sheet_name: str, cells: list[str]) -> str:
    if not cells:
        return ""
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    values = []
    for reference in cells:
        for cell_ref in _expand_reference(reference):
            value = _merged_cell_value(worksheet, cell_ref)
            if value:
                values.append(value)
    return " | ".join(_unique(values))


def _size_source_cells_with_same_row_context(
    parse_record: dict[str, Any],
    size_cells: list[str],
) -> list[str]:
    rows = _cell_rows(size_cells)
    if not rows:
        return size_cells
    same_row_cells = list(size_cells)
    fields = parse_record.get("fields", {})
    for field_name in ("款式", "行备注", "颜色"):
        source = fields.get(field_name, {}).get("source", {})
        for cell in source.get("cells", []):
            if isinstance(cell, str) and _cell_rows([cell]) & rows and cell not in same_row_cells:
                same_row_cells.append(cell)
    return same_row_cells


def _cell_rows(cells: list[str]) -> set[int]:
    rows: set[int] = set()
    for cell in cells:
        if ":" in cell:
            _, min_row, _, max_row = range_boundaries(cell)
            rows.update(range(min_row, max_row + 1))
            continue
        match = re.search(r"(\d+)$", cell)
        if match:
            rows.add(int(match.group(1)))
    return rows


def _expand_reference(reference: str) -> list[str]:
    if ":" not in reference:
        return [reference]
    min_col, min_row, max_col, max_row = range_boundaries(reference)
    cells = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append(worksheet_cell_name(row, col))
    return cells


def worksheet_cell_name(row: int, column: int) -> str:
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"


def _merged_cell_value(worksheet, coordinate: str) -> str:
    cell = worksheet[coordinate]
    if cell.value is not None:
        return _normalize_text(cell.value)
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            source = worksheet.cell(merged_range.min_row, merged_range.min_col)
            return _normalize_text(source.value)
    return ""


def _resolve_result_path(parse_report: dict[str, Any], results_dir: Path) -> Path:
    configured = parse_report.get("outputs", {}).get("result_json")
    if configured:
        configured_path = Path(str(configured))
        if configured_path.exists():
            return configured_path
        candidate = results_dir / configured_path.name
        if candidate.exists():
            return candidate
    input_name = str(parse_report.get("input", {}).get("file_name", ""))
    fallback = results_dir / f"{Path(input_name).stem}_gate2d.json"
    if fallback.exists():
        return fallback
    raise ShadowCompareError(f"Cannot resolve official result JSON for {input_name}")


def _build_summary(file_reports: list[ShadowFileReport]) -> dict[str, Any]:
    field_counts: dict[str, dict[str, int]] = {
        field: {status: 0 for status in SHADOW_STATUSES} for field in SHADOW_FIELDS
    }
    status_records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    record_count = 0
    for file_report in file_reports:
        for record in file_report.records:
            record_count += 1
            for field_name, comparison in record.fields.items():
                field_counts[field_name][comparison.comparison_status] += 1
                if comparison.comparison_status in {
                    "conflict",
                    "ambiguous",
                    "dictionary_more_specific",
                }:
                    status_records[comparison.comparison_status].append(
                        {
                            "source_file": file_report.source_file,
                            "line_number": record.line_number,
                            "field_name": field_name,
                            "python_value": comparison.python_value,
                            "dictionary_candidates": comparison.dictionary_candidates,
                            "reason": comparison.reason,
                        }
                    )
    totals = Counter()
    for counts in field_counts.values():
        totals.update(counts)
    return {
        "file_count": len(file_reports),
        "record_count": record_count,
        "field_count": record_count * len(SHADOW_FIELDS),
        "shadow_fields": list(SHADOW_FIELDS),
        "status_totals": {status: totals[status] for status in SHADOW_STATUSES},
        "field_status_counts": field_counts,
        "conflict_records": status_records["conflict"],
        "ambiguous_records": status_records["ambiguous"],
        "dictionary_more_specific_records": status_records["dictionary_more_specific"],
    }


def _require_dir(path: Path, label: str) -> Path:
    resolved = path.expanduser().resolve()
    if not resolved.exists() or not resolved.is_dir():
        raise ShadowCompareError(f"Missing {label}: {path}")
    return resolved


def _read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ShadowCompareError(f"Failed to read JSON: {path}") from exc


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", "\n").split()).strip()


def _audit_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _normalize_text(value))
    return re.sub(r"\s+", " ", text.casefold()).strip()


def _unique(values) -> list[str]:
    seen = set()
    result = []
    for value in values:
        text = _normalize_text(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result


def _extract_size_candidate(text: str) -> str:
    parsed = _parse_size(text, raw_order="width_length")
    return parsed.candidate_text() if parsed else ""


def _parse_size(text: str, *, raw_order: str) -> _SizeParse | None:
    lowered = _audit_text(text)
    if not lowered:
        return None
    pattern = re.compile(
        r"(?:(?P<label1>[wl])\s*[:=]?\s*)?(?P<a>\d+(?:\.\d+)?)\s*"
        r"(?P<unit1>cm|mm|inch|in|\")?\s*[x*×]\s*"
        r"(?:(?P<label2>[wl])\s*[:=]?\s*)?(?P<b>\d+(?:\.\d+)?)\s*"
        r"(?P<unit2>cm|mm|inch|in|\")?",
        re.IGNORECASE,
    )
    match = pattern.search(lowered)
    if not match:
        return None
    unit = _size_unit(match.group("unit1") or match.group("unit2"))
    a = _size_to_cm(float(match.group("a")), unit)
    b = _size_to_cm(float(match.group("b")), unit)
    label1 = match.group("label1")
    label2 = match.group("label2")
    if label1 == "w" or label2 == "l" or raw_order == "width_length":
        length, width = b, a
    else:
        length, width = a, b
    return _SizeParse(
        raw_first_dimension=a,
        raw_second_dimension=b,
        normalized_length=length,
        normalized_width=width,
        unit=unit,
        structural_extension_cm=_extract_structural_extension_cm(lowered),
    )


def _size_unit(unit: str | None) -> str:
    normalized = (unit or "cm").casefold()
    if normalized in {'"'}:
        return "inch"
    if normalized == "in":
        return "inch"
    return normalized


def _size_to_cm(value: float, unit: str | None) -> float:
    normalized = _size_unit(unit)
    if normalized == "mm":
        return value / 10
    if normalized == "inch":
        return value * 2.54
    return value


def _extract_structural_extension_cm(text: str) -> float:
    extension_terms = (
        r"inner\s+flap|overlapping\s+piece|opening\s+extension|"
        r"overlap|flap|重叠片?|内叠片|开口(?:延伸|加长)"
    )
    patterns = [
        re.compile(
            rf"(?P<term>{extension_terms})\D{{0,24}}"
            r"(?P<number>\d+(?:\.\d+)?)\s*(?P<unit>mm|cm|inch|in|\")?",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?P<number>\d+(?:\.\d+)?)\s*(?P<unit>mm|cm|inch|in|\")?"
            rf"\D{{0,24}}(?P<term>{extension_terms})",
            re.IGNORECASE,
        ),
        re.compile(
            r"\+\s*(?P<number>\d+(?:\.\d+)?)\s*(?P<unit>mm|cm|inch|in|\")?",
            re.IGNORECASE,
        ),
    ]
    for pattern in patterns:
        for match in pattern.finditer(text):
            if _extension_match_is_excluded(text, match):
                continue
            return _size_to_cm(float(match.group("number")), match.group("unit"))
    return 0


def _extension_match_is_excluded(text: str, match: re.Match[str]) -> bool:
    window = text[max(0, match.start() - 35) : match.end() + 35]
    number_position = match.start("number") - max(0, match.start() - 35)
    if "|" in text[match.start() : match.start("number")]:
        return True
    if re.search(r"\b(?:no|without)\s+(?:inner\s+)?flap\b|无内叠片|无叠片", window, flags=re.IGNORECASE):
        return True
    near_number = window[max(0, number_position - 18) : number_position + 30]
    if re.search(r"hand\s*(?:hole|hold)s?|location|loc\b", near_number, flags=re.IGNORECASE):
        return True
    excluded = (
        r"hand\s*(?:hole|hold)s?|location|loc\b|tc\b|thread\s*count|"
        r"\d+\s*s\s*[x*]\s*\d+\s*s|gsm|percent|%|色线|飞边|flange"
    )
    if re.search(excluded, window, flags=re.IGNORECASE):
        term = match.groupdict().get("term", "")
        if term and re.search(r"flap|overlap|opening|重叠|内叠|开口", term, flags=re.IGNORECASE):
            hand_window = re.search(r"hand\s*(?:hole|hold)s?", window, flags=re.IGNORECASE)
            if hand_window and abs(hand_window.start() - (match.start() - max(0, match.start() - 35))) < 16:
                return True
            return False
        return True
    return False


def _format_number(value: float) -> str:
    rounded = round(value, 2)
    if rounded.is_integer():
        return str(int(rounded))
    return f"{rounded:.2f}".rstrip("0").rstrip(".")


def _normalize_size_literal(value: str) -> str:
    return _audit_text(value).replace(" ", "").replace("×", "*")


def _size_text_exactly_matches(source_text: str, official_text: str) -> bool:
    source = _parse_size(source_text, raw_order="length_width")
    official = _parse_size(official_text, raw_order="length_width")
    return bool(source and official and source.candidate_text() == official.candidate_text())


def _same_number(left: float, right: float) -> bool:
    return abs(left - right) < 0.01


def _same_size_reversed(left: str, right: str) -> bool:
    left_match = re.match(r"(\d+(?:\.\d+)?)\*(\d+(?:\.\d+)?)(.*)", left or "")
    right_match = re.match(r"(\d+(?:\.\d+)?)\*(\d+(?:\.\d+)?)(.*)", right or "")
    return bool(
        left_match
        and right_match
        and left_match.group(1) == right_match.group(2)
        and left_match.group(2) == right_match.group(1)
        and left_match.group(3) == right_match.group(3)
    )


def _color_candidates(text: str) -> list[str]:
    lowered = _audit_text(text)
    candidates = []
    if re.search(r"\bwhite\b|\bplain white\b|漂白|本白", lowered):
        candidates.append("漂白色")
    if re.search(r"\bgrey\b|\bgray\b|灰", lowered):
        candidates.append("灰色")
    if re.search(r"\bblue\b|蓝", lowered):
        candidates.append("蓝色")
    if re.search(r"\bbeige\b|米", lowered):
        candidates.append("米色")
    return _unique(candidates)


def _standard_color(text: str) -> str:
    candidates = _color_candidates(text)
    return candidates[0] if candidates else ""


def _fabric_components(text: str) -> dict[str, str]:
    lowered = _audit_text(text)
    components: dict[str, str] = {}
    if re.search(r"sateen\s+stripes?|stripes?\s+sateen|striped?|缎条|条纹", lowered):
        components["category"] = "缎条"
    elif re.search(r"sateen|satin|贡缎|缎纹", lowered):
        components["category"] = "贡缎"
    elif re.search(r"percale|平纹|平布", lowered):
        components["category"] = "平布"
    elif re.search(r"twill|斜纹", lowered):
        components["category"] = "斜纹"
    elif re.search(r"\bplain\s+(?:fabric|woven|weave)\b", lowered):
        components["category"] = "平布"
    density_matches = re.findall(
        r"\b(?:t|tc)\s*[:=\-]?\s*(\d{2,4})\b"
        r"|\b(\d{2,4})\s*(?:tc|t|thread\s*count)\b",
        lowered,
    )
    density_values = [
        int(left or right)
        for left, right in density_matches
        if int(left or right) >= 100
    ]
    if density_values:
        components["density"] = f"T{density_values[-1]}"
    composition = _composition_candidate(
        lowered,
        allow_coton=True,
        allow_extended_cotton=True,
    )
    if composition:
        components["composition"] = _composition_label(composition)
    yarn_count = _yarn_count(lowered)
    if yarn_count:
        components["yarn_count"] = yarn_count
    construction = _construction_density(lowered)
    if construction:
        components["construction"] = construction
    weave = _weave_type(lowered)
    if weave:
        components["weave"] = weave
    stripe_width = _stripe_width(lowered)
    if stripe_width:
        components["stripe_width"] = stripe_width
    color = _standard_color(lowered)
    if color:
        components["color"] = color
    return components


def _compose_fabric_candidate(components: dict[str, str]) -> str:
    return "/".join(value for value in components.values() if value)


def _fabric_projection(components: dict[str, str]) -> str:
    return "/".join(
        components[key]
        for key in ("category", "density", "composition")
        if components.get(key)
    )


def _fabric_row_components(row: FabricRow) -> dict[str, str]:
    components = _fabric_components(f"{row.fabric_family} {row.fabric_standard}")
    composition = _composition_candidate(row.composition_raw, allow_coton=True)
    if composition:
        components["composition"] = _composition_label(composition)
    elif "composition" in components:
        del components["composition"]
    density_match = re.search(r"\bT(\d{2,4})\b", row.density, flags=re.IGNORECASE)
    if density_match:
        components["density"] = f"T{int(density_match.group(1))}"
    color = _standard_color(row.color_standard)
    if color:
        components["color"] = color
    return components


def _matching_fabric_rows(
    rows: list[FabricRow],
    components: dict[str, str],
) -> list[FabricRow]:
    matched = list(rows)
    row_components = {row.source_row: _fabric_row_components(row) for row in rows}
    filter_order = (
        "category",
        "composition",
        "density",
        "yarn_count",
        "construction",
        "weave",
        "stripe_width",
        "color",
    )
    for key in filter_order:
        if not components.get(key):
            continue
        matched = [
            row
            for row in matched
            if row_components[row.source_row].get(key) == components[key]
        ]
        if not matched:
            break
    return matched


def _yarn_count(text: str) -> str:
    match = re.search(
        r"\b(?:j?c|t)?\s*(\d{2,3})\s*s(?:\s*/\s*2)?"
        r"\s*[x*]\s*(?:j?c|t)?\s*(\d{2,3})\s*s(?:\s*/\s*2)?\b",
        text,
    )
    if not match:
        return ""
    return f"{int(match.group(1))}S*{int(match.group(2))}S"


def _construction_density(text: str) -> str:
    for match in re.finditer(r"\b(\d{2,3})\s*[x*]\s*\(?(\d{2,3})", text):
        left = int(match.group(1))
        right = int(match.group(2))
        suffix = text[match.end() : match.end() + 5]
        prefix = text[max(0, match.start() - 8) : match.start()]
        if re.match(r"\s*(?:cm|mm|inch|in\b|\")", suffix):
            continue
        if re.search(r"(?:duvet|size|[wl])\s*$", prefix):
            continue
        if 80 <= left <= 250 and 50 <= right <= 250:
            return f"{left}*{right}"
    return ""


def _weave_type(text: str) -> str:
    if re.search(r"sateen|satin|贡缎|缎纹", text):
        return "缎纹"
    if re.search(r"twill|斜纹", text):
        return "斜纹"
    if re.search(r"percale|plain\s+(?:fabric|woven|weave)|平纹|平布", text):
        return "平纹"
    return ""


def _stripe_width(text: str) -> str:
    matches = list(
        re.finditer(
            r"(\d+(?:\.\d+)?)\s*(mm|cm).{0,20}?(?:sateen\s+)?stripes?"
            r"|(?:sateen\s+)?stripes?.{0,20}?(\d+(?:\.\d+)?)\s*(mm|cm)"
            r"|(\d+(?:\.\d+)?)\s*(mm|cm)\s*缎条"
            r"|缎条.{0,20}?(\d+(?:\.\d+)?)\s*(mm|cm)",
            text,
        )
    )
    if not matches:
        return ""
    match = matches[0]
    number = next(
        float(group)
        for index, group in enumerate(match.groups())
        if index in (0, 2, 4, 6) and group is not None
    )
    unit = next(
        group
        for index, group in enumerate(match.groups())
        if index in (1, 3, 5, 7) and group is not None
    )
    centimeters = number / 10 if unit == "mm" else number
    return f"{_format_number(centimeters)}cm"


def _composition_candidate(
    text: str,
    *,
    allow_coton: bool = False,
    allow_extended_cotton: bool = False,
) -> tuple[int, int] | None:
    lowered = _audit_text(text)
    cotton_words = r"(?:cotton|coton|棉)" if allow_coton else r"(?:cotton|棉)"
    direct_cotton = rf"100\s*%?\s*{cotton_words}"
    extended_cotton = (
        rf"100\s*%?\s*.{{0,20}}?{cotton_words}"
        if allow_extended_cotton
        else direct_cotton
    )
    if re.search(rf"{extended_cotton}|100c\b|100%c\b", lowered):
        return (100, 0)
    c_t = re.search(r"c\s*(\d{1,3})\s*/\s*t\s*(\d{1,3})", lowered)
    if c_t:
        return (int(c_t.group(1)), int(c_t.group(2)))
    t_c = re.search(r"t\s*(\d{1,3})\s*/\s*c\s*(\d{1,3})", lowered)
    if t_c:
        return (int(t_c.group(2)), int(t_c.group(1)))
    cotton_poly = re.search(
        rf"(\d{{1,3}})\s*%?\s*{cotton_words}.{{0,20}}?"
        r"(\d{1,3})\s*%?\s*(?:poly(?:ester)?|涤)",
        lowered,
    )
    if cotton_poly:
        return (int(cotton_poly.group(1)), int(cotton_poly.group(2)))
    pair = re.search(r"\b(\d{1,3})\s*/\s*(\d{1,3})\b", lowered)
    if pair and (
        "cotton" in lowered
        or (allow_coton and "coton" in lowered)
        or "棉" in lowered
        or "poly" in lowered
        or "涤" in lowered
    ):
        return (int(pair.group(1)), int(pair.group(2)))
    return None


def _composition_label(candidate: tuple[int, int]) -> str:
    cotton, polyester = candidate
    if polyester == 0:
        return "100C"
    return f"C{cotton}/T{polyester}"


def _style_components(text: str) -> dict[str, str]:
    lowered = _audit_text(text)
    components: dict[str, str] = {}
    has_bag = bool(
        re.search(
            r"\bbag\b|\bbag\s+(?:style|model|type)\b|\bopen\s+bag\b"
            r"|\bpocket\b|口袋式|口袋",
            lowered,
        )
    )
    has_positive_hand_hole = bool(
        re.search(r"\bhands?\s+holes?\b|\bhand\s+holes?\b|\bhand\s+holds?\b|手洞", lowered)
    ) and not bool(
        re.search(
            r"\b(?:no|without)\s+hands?\s+holes?\b"
            r"|\b(?:no|without)\s+hand\s+(?:holes?|holds?)\b|无手洞",
            lowered,
        )
    )
    has_no_flange = bool(
        re.search(r"\b(?:no|without)\s+(?:flange|falnge)\b|无飞边", lowered)
    )
    has_three_side_flange = bool(
        re.search(
            r"\b(?:3|three)\s*sides?\s+(?:with\s+)?(?:flange|falnge)\b"
            r"|\b(?:flange|falnge)\s+(?:at|for|on)\s+(?:3|three)\s+sides?\b",
            lowered,
        )
    )
    has_flange = bool(re.search(r"\b(?:flange|falnge)\b|飞边", lowered)) and not has_no_flange
    has_tail_flange = bool(
        re.search(r"\b5\s*cm\s*(?:flange|hem)\b|\binternal\s+fold\b", lowered)
    )
    has_no_tie = bool(re.search(r"\b(?:no|without)\s+ties?\b|无系带", lowered))
    has_tie = bool(re.search(r"\bties?\b|系带", lowered)) and not has_no_tie
    has_zipper = bool(re.search(r"\bzip(?:per)?\b|拉链", lowered))
    has_no_flap = bool(
        re.search(r"\b(?:no|without)\s+(?:inner\s+)?flap\b|无内叠片|无叠片", lowered)
    )
    positive_flap_text = re.sub(
        r"\b(?:no|without)\s+(?:inner\s+)?flap\b|无内叠片|无叠片",
        " ",
        lowered,
    )
    has_flap = bool(
        re.search(
            r"\b(?:inner\s+)?flap\b|overlap(?:ping)?|内叠片|重叠片|叠边",
            positive_flap_text,
        )
    )
    has_bottom_opening = bool(
        re.search(
            r"\bbottom\s+opening\b"
            r"|\bopen(?:ing)?\s+(?:at|on)\s+(?:the\s+)?bottom\b"
            r"|底部开口|被尾开口",
            lowered,
        )
    )
    has_envelope = bool(re.search(r"\benvelope(?:\s+style)?\b|信封", lowered)) or (
        has_flap and not has_bag
    )

    if has_bag:
        components["form"] = "bag"
    elif has_envelope:
        components["form"] = "envelope"
    elif has_bottom_opening and has_no_flap:
        components["form"] = "open_bottom_no_flap"

    if "form" in components:
        if components["form"] == "envelope":
            components["flange"] = "none"
        elif has_three_side_flange:
            components["flange"] = "three"
        elif has_tail_flange or has_flange:
            components["flange"] = "tail_single"
        else:
            components["flange"] = "none"
        components["tie"] = "yes" if has_tie else "no"
        components["zipper"] = "yes" if has_zipper else "no"
        components["pocket"] = "yes" if has_bag else "no"
        components["welcome"] = "yes" if has_positive_hand_hole else "no"
        components["double_opening"] = (
            "yes"
            if has_bag and components["flange"] in {"tail_single", "three"}
            else "no"
        )
    if has_bottom_opening:
        components["bottom_opening"] = "yes"
    if has_flap:
        components["flap"] = "yes"
    elif has_no_flap:
        components["flap"] = "no"
    if has_positive_hand_hole:
        components["hand_hole"] = "yes"
    if re.search(r"\bpiping\b|牙条", lowered):
        components["piping"] = "yes"
    return components


def _style_row_components(row: StyleRow) -> dict[str, str]:
    text = _audit_text(f"{row.standard_name} {row.other_structure}")
    components: dict[str, str] = {}
    if "口袋" in text or row.has_pocket in {"是", "有口袋"}:
        components["form"] = "bag"
        components["pocket"] = "yes"
    elif "信封" in text or "平口" in text:
        components["form"] = "envelope"
        components["pocket"] = "no"
    elif row.has_pocket in {"否", "无口袋"}:
        components["pocket"] = "no"
    if "无飞边" in row.flange or "无飞边" in text:
        components["flange"] = "none"
    elif "四飞边" in row.flange or "四飞边" in text:
        components["flange"] = "four"
    elif "三飞边" in row.flange or "三飞边" in text:
        components["flange"] = "three"
    elif "两飞边" in row.flange or "两飞边" in text:
        components["flange"] = "two"
    elif "被头单飞边" in text:
        components["flange"] = "head_single"
    elif "被尾单飞边" in text:
        components["flange"] = "tail_single"
    elif "单飞边" in text:
        components["flange"] = "single"
    elif "飞边" in row.flange or "飞边" in text:
        components["flange"] = "tail_single"
    if "有系带" in row.tie or "加系带" in text:
        components["tie"] = "yes"
    elif "无系带" in row.tie or "无系带" in text:
        components["tie"] = "no"
    if "有拉链" in row.zipper or "拉链" in text:
        components["zipper"] = "yes"
    elif "无拉链" in row.zipper or row.zipper:
        components["zipper"] = "no"
    if row.is_welcome_style in {"是", "迎宾式"} or "迎宾" in text:
        components["welcome"] = "yes"
    elif row.is_welcome_style in {"否", "非迎宾"} or row.is_welcome_style:
        components["welcome"] = "no"
    components["double_opening"] = (
        "yes" if re.search(r"双层口|双叠边", text) else "no"
    )
    if "牙条" in text:
        components["piping"] = "yes"
    if "魔术贴" in text:
        components["magic_tape"] = "yes"
    if "纽扣" in text:
        components["button"] = "yes"
    if "压舌" in text:
        components["tongue"] = "yes"
    if "距边开口" in text or "反面距" in text:
        components["offset_opening"] = "yes"
    if "勾角" in text:
        components["hooked_corner"] = "yes"
    if "四角系带" in text:
        components["corner_tie"] = "yes"
    if components.get("form") == "bag" and "信封" in text:
        components["envelope_hybrid"] = "yes"
    return components


def _matching_style_rows(
    rows: list[StyleRow],
    components: dict[str, str],
) -> list[StyleRow]:
    core_keys = (
        "form",
        "flange",
        "tie",
        "zipper",
        "pocket",
        "welcome",
        "double_opening",
    )
    unsupported_extras = (
        "piping",
        "magic_tape",
        "button",
        "tongue",
        "offset_opening",
        "hooked_corner",
        "corner_tie",
        "envelope_hybrid",
    )
    matched = []
    for row in rows:
        row_components = _style_row_components(row)
        if any(
            components.get(key) is not None
            and row_components.get(key) != components[key]
            for key in core_keys
            if key in components
        ):
            continue
        if any(
            row_components.get(key) == "yes" and components.get(key) != "yes"
            for key in unsupported_extras
        ):
            continue
        if components.get("form") == "open_bottom_no_flap":
            continue
        if components.get("form") and row_components.get("form") != components["form"]:
            continue
        if "flange" in components and row_components.get("flange") != components["flange"]:
            continue
        if "form" in components and all(
            key in row_components for key in ("tie", "zipper", "pocket", "welcome")
        ):
            matched.append(row)
    return matched


def _style_candidate_compatible(candidate: str, official: str) -> bool:
    candidate_components = _style_standard_components(candidate)
    official_components = _style_standard_components(official)
    comparable_keys = (
        "form",
        "flange",
        "tie",
        "zipper",
        "pocket",
        "welcome",
        "double_opening",
    )
    shared = [
        key
        for key in comparable_keys
        if key in candidate_components and key in official_components
    ]
    return bool(
        shared
        and all(
            candidate_components[key] == official_components[key] for key in shared
        )
    )


def _style_standard_components(text: str) -> dict[str, str]:
    lowered = _audit_text(text)
    components: dict[str, str] = {}
    if "口袋" in lowered:
        components["form"] = "bag"
        components["pocket"] = "yes"
    elif "信封" in lowered or "平口" in lowered:
        components["form"] = "envelope"
        components["pocket"] = "no"
    if "无飞边" in lowered:
        components["flange"] = "none"
    elif "三飞边" in lowered:
        components["flange"] = "three"
    elif "飞边" in lowered:
        components["flange"] = "tail_single"
    components["tie"] = "yes" if re.search(r"有系带|加系带", lowered) else "no"
    components["zipper"] = "yes" if "拉链" in lowered else "no"
    components["welcome"] = "yes" if "迎宾" in lowered else "no"
    components["double_opening"] = (
        "yes" if re.search(r"双层口|双叠边", lowered) else "no"
    )
    return components


def _style_projection_label(components: dict[str, str]) -> str:
    form = components.get("form")
    flange = components.get("flange")
    tie = "有系带" if components.get("tie") == "yes" else "无系带"
    welcome = "迎宾" if components.get("welcome") == "yes" else ""
    if form == "bag":
        if flange == "three":
            return f"三飞边双层口叠边口袋{tie}{welcome}式"
        if flange == "tail_single":
            return f"被尾单飞边双层口叠边口袋{tie}{welcome}式"
        if flange == "none":
            return f"无飞边口袋{tie}{welcome}式"
    if form == "envelope" and flange == "none":
        return f"无飞边平口信封{welcome}式"
    return ""


def _style_evidence_sufficient(components: dict[str, str]) -> bool:
    return components.get("form") in {"bag", "envelope", "open_bottom_no_flap"}


def _style_label(components: dict[str, str]) -> str:
    return "/".join(f"{key}={components[key]}" for key in sorted(components))
