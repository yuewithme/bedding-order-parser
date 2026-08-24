"""Read synthetic or approved workbooks into sparse, scoped AI evidence."""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
from pathlib import Path
import re
from typing import Any, Protocol

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bedding_order_parser.ai_full_order.contracts import (
    ParseMode,
    SCHEMA_VERSION,
    build_source_record_id,
    normalize_evidence_text,
)
from bedding_order_parser.excel.workbook_reader import compute_sha256
from bedding_order_parser.ai_full_order.standard_geometry import (
    derive_standard_sheet_geometry,
)


PREPROCESSOR_VERSION = "1.2"


HEADER_WORDS = (
    "no",
    "item",
    "items",
    "description",
    "specification",
    "size",
    "qty",
    "quantity",
    "序号",
    "品名",
    "描述",
    "规格",
    "尺寸",
    "数量",
)
NUMBERED_ROW = re.compile(r"\d+(?:\.0)?$")


class StructureResolver(Protocol):
    def resolve_structure(self, manifest: dict[str, Any]) -> object: ...


@dataclass(frozen=True)
class SparseCell:
    cell_id: str
    sheet_id: str
    sheet_name: str
    reference: str
    display_text: str
    formula_text: str
    value_type: str
    number_format: str
    merged_anchor: str

    def to_dict(self) -> dict[str, str]:
        return {
            "cell_id": self.cell_id,
            "sheet_id": self.sheet_id,
            "sheet_name": self.sheet_name,
            "reference": self.reference,
            "display_text": self.display_text,
            "formula_text": self.formula_text,
            "value_type": self.value_type,
            "number_format": self.number_format,
            "merged_anchor": self.merged_anchor,
        }


@dataclass(frozen=True)
class EvidenceItem:
    evidence_id: str
    scope_id: str
    sheet_id: str
    sheet_name: str
    cell_range: str
    original_text: str
    normalized_text: str

    def to_dict(self) -> dict[str, str]:
        return {
            "evidence_id": self.evidence_id,
            "scope_id": self.scope_id,
            "sheet_id": self.sheet_id,
            "sheet_name": self.sheet_name,
            "cell_range": self.cell_range,
            "original_text": self.original_text,
            "normalized_text": self.normalized_text,
        }


@dataclass(frozen=True)
class LocalRecord:
    record_local_id: str
    source_record_id: str
    scope_id: str
    sheet_id: str
    source_row: int
    local_line_number: str
    evidence_ids: tuple[str, ...]

    def to_request_dict(self) -> dict[str, Any]:
        return {
            "record_local_id": self.record_local_id,
            "source_record_id": self.source_record_id,
            "scope_id": self.scope_id,
            "sheet_id": self.sheet_id,
            "source_row": self.source_row,
            "evidence_ids": list(self.evidence_ids),
        }


@dataclass(frozen=True)
class OrderBlock:
    block_id: str
    scope_id: str
    sheet_id: str
    sheet_name: str
    cell_range: str
    header_evidence_ids: tuple[str, ...]
    record_local_ids: tuple[str, ...]

    def to_request_dict(self) -> dict[str, Any]:
        return {
            "block_id": self.block_id,
            "scope_id": self.scope_id,
            "sheet_id": self.sheet_id,
            "sheet_name": self.sheet_name,
            "cell_range": self.cell_range,
            "header_evidence_ids": list(self.header_evidence_ids),
            "record_local_ids": list(self.record_local_ids),
        }


@dataclass(frozen=True)
class SheetStructure:
    sheet_id: str
    sheet_name: str
    visibility: str
    included: bool
    used_range: str
    hidden_rows: tuple[int, ...]
    hidden_columns: tuple[str, ...]
    cells: tuple[SparseCell, ...]


@dataclass(frozen=True)
class LocalStructureCandidate:
    """A locally constructed option that layout AI may only select by ID."""

    candidate_id: str
    sheet_id: str
    role: str
    block_ids: tuple[str, ...]
    record_local_ids: tuple[str, ...]
    cell_ranges: tuple[str, ...]


@dataclass(frozen=True)
class SheetStructureState:
    """Independent local structure facts for one included workbook sheet."""

    sheet_id: str
    local_status: str
    standard_geometry_status: str
    heuristic_status: str
    known_block_ids: tuple[str, ...]
    known_record_local_ids: tuple[str, ...]
    candidate_ids: tuple[str, ...]
    resolution_required: bool


@dataclass(frozen=True)
class StructureDiagnostics:
    standard_parsed_record_count: int = 0
    standard_selected_record_count: int = 0
    heuristic_record_count: int = 0
    aligned_record_count: int = 0
    auxiliary_numbered_row_count: int = 0
    standard_aligned_sheet_count: int = 0
    heuristic_sheet_count: int = 0
    possible_secondary_table_count: int = 0
    evidence_mapping_failure_count: int = 0

    def to_dict(self) -> dict[str, int]:
        return {
            "standard_parsed_record_count": self.standard_parsed_record_count,
            "standard_selected_record_count": self.standard_selected_record_count,
            "heuristic_record_count": self.heuristic_record_count,
            "aligned_record_count": self.aligned_record_count,
            "auxiliary_numbered_row_count": self.auxiliary_numbered_row_count,
            "standard_aligned_sheet_count": self.standard_aligned_sheet_count,
            "heuristic_sheet_count": self.heuristic_sheet_count,
            "possible_secondary_table_count": self.possible_secondary_table_count,
            "evidence_mapping_failure_count": self.evidence_mapping_failure_count,
        }


@dataclass(frozen=True)
class PreprocessedWorkbook:
    source_file_sha256: str
    sheets: tuple[SheetStructure, ...]
    blocks: tuple[OrderBlock, ...]
    records: tuple[LocalRecord, ...]
    evidence_catalog: tuple[EvidenceItem, ...]
    structure_status: str
    structure_resolution_requested: bool
    structure_diagnostics: StructureDiagnostics = StructureDiagnostics()
    sheet_states: tuple[SheetStructureState, ...] = ()
    layout_candidates: tuple[LocalStructureCandidate, ...] = ()

    def to_request_dict(self) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "parse_mode": ParseMode.AI_ENHANCED.value,
            "source_file_sha256": self.source_file_sha256,
            "request_chunk_id": f"local-{self.source_file_sha256[:12]}",
            "structure_status": self.structure_status,
            "blocks": [block.to_request_dict() for block in self.blocks],
            "records": [record.to_request_dict() for record in self.records],
            "record_count": len(self.records),
            "evidence_catalog": [item.to_dict() for item in self.evidence_catalog],
        }


def preprocess_workbook(
    path: str | Path,
    *,
    structure_resolver: StructureResolver | None = None,
) -> PreprocessedWorkbook:
    """Build local evidence without parsing a business result or calling a provider."""
    workbook_path = Path(path).expanduser().resolve()
    before_hash = compute_sha256(workbook_path)
    formula_book = load_workbook(workbook_path, data_only=False, read_only=False, keep_links=False)
    display_book = load_workbook(workbook_path, data_only=True, read_only=False, keep_links=False)
    try:
        sheets: list[SheetStructure] = []
        blocks: list[OrderBlock] = []
        records: list[LocalRecord] = []
        evidence_catalog: list[EvidenceItem] = []
        sheet_states: list[SheetStructureState] = []
        layout_candidates: list[LocalStructureCandidate] = []
        unresolved_structure = False
        diagnostics = StructureDiagnostics()
        for ordinal, formula_sheet in enumerate(formula_book.worksheets, start=1):
            display_sheet = display_book[formula_sheet.title]
            sheet_id = f"s{ordinal}"
            structure, visible_cells = _build_sheet_structure(sheet_id, formula_sheet, display_sheet)
            sheets.append(structure)
            if not structure.included:
                sheet_states.append(
                    SheetStructureState(
                        sheet_id=sheet_id,
                        local_status="ignored_by_local_contract",
                        standard_geometry_status="not_included",
                        heuristic_status="not_applicable",
                        known_block_ids=(),
                        known_record_local_ids=(),
                        candidate_ids=(),
                        resolution_required=False,
                    )
                )
                continue
            if not visible_cells:
                sheet_states.append(
                    SheetStructureState(
                        sheet_id=sheet_id,
                        local_status="ignored_by_local_contract",
                        standard_geometry_status="empty_visible_sheet",
                        heuristic_status="no_content",
                        known_block_ids=(),
                        known_record_local_ids=(),
                        candidate_ids=(),
                        resolution_required=False,
                    )
                )
                continue
            heuristic_blocks, heuristic_records, heuristic_evidence, heuristic_ambiguous = _build_blocks(
                source_file_sha256=before_hash,
                sheet=structure,
                cells=visible_cells,
            )
            standard_geometry = derive_standard_sheet_geometry(display_sheet)
            aligned = None
            if standard_geometry.stable:
                aligned = _build_standard_aligned_block(
                    source_file_sha256=before_hash,
                    sheet=structure,
                    cells=visible_cells,
                    header_rows=standard_geometry.header_rows,
                    record_rows=standard_geometry.record_rows,
                )
            if aligned is not None:
                sheet_blocks, sheet_records, sheet_evidence = aligned
                ambiguous = False
            else:
                sheet_blocks = heuristic_blocks
                sheet_records = heuristic_records
                sheet_evidence = heuristic_evidence
                ambiguous = True if standard_geometry.stable else heuristic_ambiguous
                if standard_geometry.has_explicit_secondary_table:
                    ambiguous = True
            blocks.extend(sheet_blocks)
            records.extend(sheet_records)
            evidence_catalog.extend(sheet_evidence)
            unresolved_structure = unresolved_structure or ambiguous
            if not ambiguous and sheet_blocks:
                sheet_states.append(
                    SheetStructureState(
                        sheet_id=sheet_id,
                        local_status="confirmed_order",
                        standard_geometry_status=standard_geometry.reason_code,
                        heuristic_status=(
                            "confirmed" if aligned is None else "superseded_by_standard"
                        ),
                        known_block_ids=tuple(block.block_id for block in sheet_blocks),
                        known_record_local_ids=tuple(
                            record.record_local_id for record in sheet_records
                        ),
                        candidate_ids=(),
                        resolution_required=False,
                    )
                )
            else:
                candidates = _layout_candidates_for_sheet(
                    source_file_sha256=before_hash,
                    sheet=structure,
                    blocks=sheet_blocks,
                    records=sheet_records,
                    standard_geometry_status=standard_geometry.reason_code,
                )
                layout_candidates.extend(candidates)
                sheet_states.append(
                    SheetStructureState(
                        sheet_id=sheet_id,
                        local_status="unresolved_order_candidate",
                        standard_geometry_status=standard_geometry.reason_code,
                        heuristic_status=(
                            "candidate_available" if sheet_blocks else "no_record_candidate"
                        ),
                        known_block_ids=(),
                        known_record_local_ids=(),
                        candidate_ids=tuple(item.candidate_id for item in candidates),
                        resolution_required=True,
                    )
                )
            diagnostics = StructureDiagnostics(
                standard_parsed_record_count=(
                    diagnostics.standard_parsed_record_count
                    + standard_geometry.parsed_record_count
                ),
                standard_selected_record_count=(
                    diagnostics.standard_selected_record_count
                    + len(standard_geometry.record_rows)
                ),
                heuristic_record_count=(
                    diagnostics.heuristic_record_count + len(heuristic_records)
                ),
                aligned_record_count=(
                    diagnostics.aligned_record_count
                    + (len(sheet_records) if aligned is not None else 0)
                ),
                auxiliary_numbered_row_count=(
                    diagnostics.auxiliary_numbered_row_count
                    + len(standard_geometry.auxiliary_numbered_rows)
                ),
                standard_aligned_sheet_count=(
                    diagnostics.standard_aligned_sheet_count
                    + int(aligned is not None)
                ),
                heuristic_sheet_count=(
                    diagnostics.heuristic_sheet_count
                    + int(aligned is None and bool(heuristic_blocks))
                ),
                possible_secondary_table_count=(
                    diagnostics.possible_secondary_table_count
                    + int(standard_geometry.has_explicit_secondary_table)
                ),
                evidence_mapping_failure_count=(
                    diagnostics.evidence_mapping_failure_count
                    + int(standard_geometry.stable and aligned is None)
                ),
            )
        after_hash = compute_sha256(workbook_path)
        if after_hash != before_hash:
            raise ValueError("Workbook SHA-256 changed during preprocessing.")
        status = "locally_resolved" if blocks and not unresolved_structure else "ambiguous"
        result = PreprocessedWorkbook(
            source_file_sha256=before_hash,
            sheets=tuple(sheets),
            blocks=tuple(blocks),
            records=tuple(records),
            evidence_catalog=tuple(evidence_catalog),
            structure_status=status,
            structure_resolution_requested=False,
            structure_diagnostics=diagnostics,
            sheet_states=tuple(sheet_states),
            layout_candidates=tuple(layout_candidates),
        )
        if status == "ambiguous" and structure_resolver is not None:
            structure_resolver.resolve_structure(result.to_request_dict())
            result = PreprocessedWorkbook(
                source_file_sha256=result.source_file_sha256,
                sheets=result.sheets,
                blocks=result.blocks,
                records=result.records,
                evidence_catalog=result.evidence_catalog,
                structure_status=result.structure_status,
                structure_resolution_requested=True,
                structure_diagnostics=result.structure_diagnostics,
                sheet_states=result.sheet_states,
                layout_candidates=result.layout_candidates,
            )
        return result
    finally:
        formula_book.close()
        display_book.close()


def _build_sheet_structure(
    sheet_id: str,
    formula_sheet: Worksheet,
    display_sheet: Worksheet,
) -> tuple[SheetStructure, dict[tuple[int, int], SparseCell]]:
    visibility = str(formula_sheet.sheet_state)
    used = _actual_used_bounds(formula_sheet, display_sheet)
    if used is None:
        return (
            SheetStructure(sheet_id, formula_sheet.title, visibility, visibility == "visible", "", (), (), ()),
            {},
        )
    min_row, min_col, max_row, max_col = used
    if visibility != "visible":
        return (
            SheetStructure(
                sheet_id, formula_sheet.title, visibility, False,
                _range_text(min_row, min_col, max_row, max_col), (), (), (),
            ),
            {},
        )
    hidden_rows = tuple(row for row in range(min_row, max_row + 1) if formula_sheet.row_dimensions[row].hidden)
    hidden_columns = tuple(
        get_column_letter(column)
        for column in range(min_col, max_col + 1)
        if formula_sheet.column_dimensions[get_column_letter(column)].hidden
    )
    anchors = _merged_anchors(formula_sheet)
    cells: list[SparseCell] = []
    cell_map: dict[tuple[int, int], SparseCell] = {}
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            if row in hidden_rows or get_column_letter(column) in hidden_columns:
                continue
            anchor = anchors.get((row, column))
            formula_cell = formula_sheet.cell(row=row, column=column)
            display_cell = display_sheet.cell(row=row, column=column)
            is_anchor = anchor in {None, (row, column)}
            formula_text = _clean(formula_cell.value) if formula_cell.data_type == "f" and is_anchor else ""
            display_text = _clean(display_cell.value) if is_anchor else ""
            if not display_text and not formula_text and anchor is None:
                continue
            reference = formula_cell.coordinate
            cell = SparseCell(
                cell_id=f"{sheet_id}!{reference}",
                sheet_id=sheet_id,
                sheet_name=formula_sheet.title,
                reference=reference,
                display_text=display_text,
                formula_text=formula_text,
                value_type="formula" if formula_text else str(formula_cell.data_type or "string"),
                number_format=str(formula_cell.number_format or "General"),
                merged_anchor="" if is_anchor else f"{sheet_id}!{formula_sheet.cell(row=anchor[0], column=anchor[1]).coordinate}",
            )
            cells.append(cell)
            cell_map[(row, column)] = cell
    return (
        SheetStructure(
            sheet_id=sheet_id,
            sheet_name=formula_sheet.title,
            visibility=visibility,
            included=visibility == "visible",
            used_range=_range_text(min_row, min_col, max_row, max_col),
            hidden_rows=hidden_rows,
            hidden_columns=hidden_columns,
            cells=tuple(cells),
        ),
        cell_map,
    )


def _actual_used_bounds(formula_sheet: Worksheet, display_sheet: Worksheet) -> tuple[int, int, int, int] | None:
    coordinates: list[tuple[int, int]] = []
    for row in range(1, formula_sheet.max_row + 1):
        for column in range(1, formula_sheet.max_column + 1):
            formula_value = formula_sheet.cell(row=row, column=column).value
            display_value = display_sheet.cell(row=row, column=column).value
            if formula_value is not None or display_value is not None:
                coordinates.append((row, column))
    for merged in formula_sheet.merged_cells.ranges:
        anchor = formula_sheet.cell(row=merged.min_row, column=merged.min_col).value
        if anchor is not None:
            coordinates.extend(
                [(merged.min_row, merged.min_col), (merged.max_row, merged.max_col)]
            )
    if not coordinates:
        return None
    rows, columns = zip(*coordinates, strict=True)
    return min(rows), min(columns), max(rows), max(columns)


def _merged_anchors(sheet: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    for merged in sheet.merged_cells.ranges:
        anchor = (merged.min_row, merged.min_col)
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                anchors[(row, column)] = anchor
    return anchors


def _build_blocks(
    *,
    source_file_sha256: str,
    sheet: SheetStructure,
    cells: dict[tuple[int, int], SparseCell],
) -> tuple[list[OrderBlock], list[LocalRecord], list[EvidenceItem], bool]:
    if not cells:
        return [], [], [], True
    rows = _rows_from_cells(cells)
    data_groups = _numbered_row_groups(rows)
    blocks: list[OrderBlock] = []
    records: list[LocalRecord] = []
    evidence: list[EvidenceItem] = []
    ambiguous = not data_groups
    for ordinal, data_rows in enumerate(data_groups, start=1):
        header_rows = _header_rows_before(rows, data_rows[0])
        if not header_rows:
            ambiguous = True
            continue
        scope_id = f"{sheet.sheet_id}:scope-{ordinal}"
        block_id = f"{sheet.sheet_id}:block-{ordinal}"
        header_evidence = _evidence_for_rows(scope_id, sheet, cells, header_rows, evidence)
        record_ids: list[str] = []
        for record_ordinal, source_row in enumerate(data_rows, start=1):
            record_evidence = list(header_evidence)
            record_evidence.extend(_evidence_for_rows(scope_id, sheet, cells, [source_row], evidence))
            record_local_id = f"{scope_id}:record-{record_ordinal}"
            record_ids.append(record_local_id)
            records.append(
                LocalRecord(
                    record_local_id=record_local_id,
                    source_record_id=build_source_record_id(
                        source_file_sha256=source_file_sha256,
                        sheet_id=sheet.sheet_id,
                        scope_id=scope_id,
                        source_row=source_row,
                        evidence_ids=record_evidence,
                    ),
                    scope_id=scope_id,
                    sheet_id=sheet.sheet_id,
                    source_row=source_row,
                    local_line_number=f"{sheet.sheet_id}:{source_row}",
                    evidence_ids=tuple(record_evidence),
                )
            )
        min_row = min([*header_rows, *data_rows])
        max_row = max([*header_rows, *data_rows])
        columns = [column for row, column in cells if min_row <= row <= max_row]
        blocks.append(
            OrderBlock(
                block_id=block_id,
                scope_id=scope_id,
                sheet_id=sheet.sheet_id,
                sheet_name=sheet.sheet_name,
                cell_range=_range_text(min_row, min(columns), max_row, max(columns)),
                header_evidence_ids=tuple(header_evidence),
                record_local_ids=tuple(record_ids),
            )
        )
    return blocks, records, evidence, ambiguous


def _build_standard_aligned_block(
    *,
    source_file_sha256: str,
    sheet: SheetStructure,
    cells: dict[tuple[int, int], SparseCell],
    header_rows: tuple[int, ...],
    record_rows: tuple[int, ...],
) -> tuple[list[OrderBlock], list[LocalRecord], list[EvidenceItem]] | None:
    if not header_rows or not record_rows or len(set(record_rows)) != len(record_rows):
        return None
    if any(not _has_line_number_evidence(cells, row) for row in record_rows):
        return None

    scope_id = f"{sheet.sheet_id}:scope-1"
    block_id = f"{sheet.sheet_id}:block-1"
    evidence: list[EvidenceItem] = []
    header_evidence = _evidence_for_rows(
        scope_id, sheet, cells, list(header_rows), evidence
    )
    if not header_evidence:
        return None

    records: list[LocalRecord] = []
    record_ids: list[str] = []
    for ordinal, source_row in enumerate(record_rows, start=1):
        record_evidence = list(header_evidence)
        record_evidence.extend(
            _evidence_for_rows(scope_id, sheet, cells, [source_row], evidence)
        )
        if not record_evidence:
            return None
        record_local_id = f"{scope_id}:record-{ordinal}"
        record_ids.append(record_local_id)
        records.append(
            LocalRecord(
                record_local_id=record_local_id,
                source_record_id=build_source_record_id(
                    source_file_sha256=source_file_sha256,
                    sheet_id=sheet.sheet_id,
                    scope_id=scope_id,
                    source_row=source_row,
                    evidence_ids=record_evidence,
                ),
                scope_id=scope_id,
                sheet_id=sheet.sheet_id,
                source_row=source_row,
                local_line_number=f"{sheet.sheet_id}:{source_row}",
                evidence_ids=tuple(record_evidence),
            )
        )

    min_row = min(*header_rows, *record_rows)
    max_row = max(*header_rows, *record_rows)
    columns = [column for row, column in cells if min_row <= row <= max_row]
    if not columns:
        return None
    block = OrderBlock(
        block_id=block_id,
        scope_id=scope_id,
        sheet_id=sheet.sheet_id,
        sheet_name=sheet.sheet_name,
        cell_range=_range_text(min_row, min(columns), max_row, max(columns)),
        header_evidence_ids=tuple(header_evidence),
        record_local_ids=tuple(record_ids),
    )
    return [block], records, evidence


def _layout_candidates_for_sheet(
    *,
    source_file_sha256: str,
    sheet: SheetStructure,
    blocks: list[OrderBlock],
    records: list[LocalRecord],
    standard_geometry_status: str,
) -> tuple[LocalStructureCandidate, ...]:
    if blocks:
        return (
            _layout_candidate(
                source_file_sha256=source_file_sha256,
                sheet_id=sheet.sheet_id,
                role="order",
                block_ids=tuple(block.block_id for block in blocks),
                record_local_ids=tuple(record.record_local_id for record in records),
                cell_ranges=tuple(block.cell_range for block in blocks),
            ),
        )
    if standard_geometry_status == "standard_table_unresolved":
        return (
            _layout_candidate(
                source_file_sha256=source_file_sha256,
                sheet_id=sheet.sheet_id,
                role="auxiliary",
                block_ids=(),
                record_local_ids=(),
                cell_ranges=(sheet.used_range,),
            ),
        )
    return ()


def _layout_candidate(
    *,
    source_file_sha256: str,
    sheet_id: str,
    role: str,
    block_ids: tuple[str, ...],
    record_local_ids: tuple[str, ...],
    cell_ranges: tuple[str, ...],
) -> LocalStructureCandidate:
    candidate_id = layout_candidate_id(
        source_file_sha256=source_file_sha256,
        sheet_id=sheet_id,
        role=role,
        block_ids=block_ids,
        record_local_ids=record_local_ids,
        cell_ranges=cell_ranges,
    )
    return LocalStructureCandidate(
        candidate_id=candidate_id,
        sheet_id=sheet_id,
        role=role,
        block_ids=block_ids,
        record_local_ids=record_local_ids,
        cell_ranges=cell_ranges,
    )


def layout_candidate_id(
    *,
    source_file_sha256: str,
    sheet_id: str,
    role: str,
    block_ids: tuple[str, ...],
    record_local_ids: tuple[str, ...],
    cell_ranges: tuple[str, ...],
) -> str:
    canonical = "\n".join(
        (
            PREPROCESSOR_VERSION,
            source_file_sha256,
            sheet_id,
            role,
            *block_ids,
            *record_local_ids,
            *cell_ranges,
        )
    )
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:20]
    return f"layout-candidate:{sheet_id}:{digest}"


def _has_line_number_evidence(
    cells: dict[tuple[int, int], SparseCell], source_row: int
) -> bool:
    cell = cells.get((source_row, 1))
    if cell is None or cell.merged_anchor:
        return False
    return NUMBERED_ROW.fullmatch((cell.display_text or cell.formula_text).strip()) is not None


def _rows_from_cells(cells: dict[tuple[int, int], SparseCell]) -> dict[int, list[str]]:
    rows: dict[int, list[str]] = {}
    for (row, column), cell in cells.items():
        if cell.merged_anchor:
            continue
        text = cell.display_text or cell.formula_text
        if text:
            rows.setdefault(row, []).append(text)
    return rows


def _numbered_row_groups(rows: dict[int, list[str]]) -> list[list[int]]:
    numbered = [row for row in sorted(rows) if _is_numbered(rows[row])]
    groups: list[list[int]] = []
    for row in numbered:
        if groups and row == groups[-1][-1] + 1:
            groups[-1].append(row)
        else:
            groups.append([row])
    return groups


def _is_numbered(values: list[str]) -> bool:
    first = next((value.strip() for value in values if value.strip()), "")
    return NUMBERED_ROW.fullmatch(first) is not None


def _header_rows_before(rows: dict[int, list[str]], first_data_row: int) -> list[int]:
    candidates: list[int] = []
    for row in range(first_data_row - 1, max(0, first_data_row - 4), -1):
        if _header_score(rows.get(row, [])) >= 2:
            candidates.append(row)
        elif candidates:
            break
    return list(reversed(candidates))


def _header_score(values: list[str]) -> int:
    content = " ".join(values).casefold()
    return sum(1 for word in HEADER_WORDS if word in content)


def _evidence_for_rows(
    scope_id: str,
    sheet: SheetStructure,
    cells: dict[tuple[int, int], SparseCell],
    rows: list[int],
    catalog: list[EvidenceItem],
) -> list[str]:
    existing = {item.evidence_id for item in catalog}
    ids: list[str] = []
    for (row, _column), cell in sorted(cells.items()):
        if row not in rows or cell.merged_anchor:
            continue
        original = cell.display_text or cell.formula_text
        if not original:
            continue
        evidence_id = f"{scope_id}:{cell.cell_id}"
        ids.append(evidence_id)
        if evidence_id not in existing:
            catalog.append(
                EvidenceItem(
                    evidence_id=evidence_id,
                    scope_id=scope_id,
                    sheet_id=sheet.sheet_id,
                    sheet_name=sheet.sheet_name,
                    cell_range=cell.reference,
                    original_text=original,
                    normalized_text=normalize_evidence_text(original),
                )
            )
            existing.add(evidence_id)
    return ids


def _range_text(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return normalize_evidence_text(str(value))
