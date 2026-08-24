"""Compute privacy-safe ranking metrics from a validated review workbook."""

from __future__ import annotations

import json
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bedding_order_parser.exceptions import BeddingOrderParserError
from bedding_order_parser.materials.loader import compute_sha256
from bedding_order_parser.materials.review_validator import validate_review_workbook
from bedding_order_parser.materials.review_workbook import (
    CANDIDATE_HEADERS,
    CANDIDATE_SHEET,
    REVIEW_HEADERS,
    REVIEW_SHEET,
)


REVIEW_METRICS_SCHEMA_VERSION = "1.0"
POSITIVE_TRUTH_CONCLUSIONS = {
    "推荐编码正确",
    "Top候选中其他编码正确",
    "Top候选外编码正确",
}
NO_MATERIAL_CONCLUSION = "物料库不存在对应物料"


class ReviewMetricsError(BeddingOrderParserError):
    """Raised when a review workbook cannot safely produce metrics."""


@dataclass(frozen=True)
class ReviewMetricsResult:
    workbook_sha256: str
    counts: dict[str, int]
    ranking: dict[str, int | float | None]
    no_material: dict[str, int | float | None]

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": REVIEW_METRICS_SCHEMA_VERSION,
            "workbook_sha256": self.workbook_sha256,
            "counts": self.counts,
            "ranking": self.ranking,
            "no_material": self.no_material,
        }


def evaluate_review_workbook(
    workbook_path: str | Path,
    store_path: str | Path,
) -> ReviewMetricsResult:
    """Validate a filled workbook and calculate aggregate retrieval metrics."""
    validation = validate_review_workbook(workbook_path, store_path)
    if not validation.ok:
        raise ReviewMetricsError(
            f"Review workbook validation failed with {len(validation.errors)} error(s)."
        )
    if validation.sha256_before != validation.sha256_after:
        raise ReviewMetricsError("Review workbook changed during validation.")

    workbook_file = Path(workbook_path).expanduser().resolve()
    sha_before = compute_sha256(workbook_file)
    if sha_before != validation.sha256_after:
        raise ReviewMetricsError("Review workbook changed after validation.")
    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    try:
        review_ws = workbook[REVIEW_SHEET]
        candidate_ws = workbook[CANDIDATE_SHEET]
        review_rows = _rows_by_header(review_ws, REVIEW_HEADERS)
        candidate_ranks = _candidate_ranks(candidate_ws)
        result = _calculate_metrics(review_rows, candidate_ranks, sha_before)
    finally:
        workbook.close()

    if compute_sha256(workbook_file) != sha_before:
        raise ReviewMetricsError("Review workbook changed while metrics were calculated.")
    return result


def write_review_metrics(
    result: ReviewMetricsResult,
    output_path: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    """Atomically write the aggregate-only metrics JSON."""
    output_file = Path(output_path).expanduser().resolve()
    if output_file.exists() and not overwrite:
        raise ReviewMetricsError(f"Output already exists: {output_file}")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=output_file.parent,
            prefix=f".{output_file.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
            json.dump(result.to_dict(), handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, output_file)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()
    return output_file


def _rows_by_header(ws, headers: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            {
                header: _text(value)
                for header, value in zip(headers, values, strict=True)
            }
        )
    return rows


def _candidate_ranks(ws) -> dict[str, dict[str, int]]:
    audit_index = CANDIDATE_HEADERS.index("审核序号")
    rank_index = CANDIDATE_HEADERS.index("候选排名")
    code_index = CANDIDATE_HEADERS.index("物料编码")
    result: dict[str, dict[str, int]] = {}
    for values in ws.iter_rows(
        min_row=2,
        max_col=len(CANDIDATE_HEADERS),
        values_only=True,
    ):
        audit_id = _text(values[audit_index])
        code = _text(values[code_index])
        try:
            rank = int(_text(values[rank_index]))
        except ValueError:
            continue
        if not audit_id or not code or rank < 1:
            continue
        code_ranks = result.setdefault(audit_id, {})
        code_ranks[code] = min(rank, code_ranks.get(code, rank))
    return result


def _calculate_metrics(
    review_rows: list[dict[str, str]],
    candidate_ranks: dict[str, dict[str, int]],
    workbook_sha256: str,
) -> ReviewMetricsResult:
    reviewed_rows = 0
    positive_truth_rows = 0
    no_material_truth_rows = 0
    no_material_no_recommendation_hits = 0
    excluded_rows = 0
    ranks: list[int | None] = []

    for row in review_rows:
        conclusion = row["审核结论"]
        if not conclusion:
            continue
        reviewed_rows += 1
        if conclusion in POSITIVE_TRUTH_CONCLUSIONS:
            positive_truth_rows += 1
            ranks.append(
                candidate_ranks.get(row["审核序号"], {}).get(row["正确物料编码"])
            )
        elif conclusion == NO_MATERIAL_CONCLUSION:
            no_material_truth_rows += 1
            if not row["推荐物料编码"]:
                no_material_no_recommendation_hits += 1
        else:
            excluded_rows += 1

    denominator = positive_truth_rows
    top1_hits = sum(rank is not None and rank <= 1 for rank in ranks)
    top3_hits = sum(rank is not None and rank <= 3 for rank in ranks)
    top10_hits = sum(rank is not None and rank <= 10 for rank in ranks)
    counts = {
        "total_rows": len(review_rows),
        "reviewed_rows": reviewed_rows,
        "positive_truth_rows": positive_truth_rows,
        "no_material_truth_rows": no_material_truth_rows,
        "excluded_rows": excluded_rows,
        "unreviewed_rows": len(review_rows) - reviewed_rows,
    }
    ranking: dict[str, int | float | None] = {
        "denominator": denominator,
        "top1_hits": top1_hits,
        "top3_hits": top3_hits,
        "top10_hits": top10_hits,
        "outside_top10": denominator - top10_hits,
        "top1_rate": _rate(top1_hits, denominator),
        "top3_rate": _rate(top3_hits, denominator),
        "top10_rate": _rate(top10_hits, denominator),
    }
    no_material: dict[str, int | float | None] = {
        "denominator": no_material_truth_rows,
        "no_recommendation_hits": no_material_no_recommendation_hits,
        "false_recommendations": (
            no_material_truth_rows - no_material_no_recommendation_hits
        ),
        "no_recommendation_rate": _rate(
            no_material_no_recommendation_hits, no_material_truth_rows
        ),
    }
    return ReviewMetricsResult(
        workbook_sha256=workbook_sha256,
        counts=counts,
        ranking=ranking,
        no_material=no_material,
    )


def _rate(numerator: int, denominator: int) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
