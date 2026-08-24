"""Build the Gate 3B-D human review workbook for material matching."""

from __future__ import annotations

import json
import shutil
import sqlite3
import uuid
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from bedding_order_parser.exceptions import BeddingOrderParserError


REVIEW_SHEET = "审核清单"
CANDIDATE_SHEET = "候选明细"
ISSUE_SHEET = "重点问题"
INSTRUCTION_SHEET = "填写说明"
MATERIAL_INDEX_SHEET = "物料编码索引"

REVIEW_HEADERS = [
    "审核序号",
    "来源文件",
    "Sheet",
    "订单行号",
    "客户",
    "物料名称",
    "规格",
    "颜色",
    "面料",
    "面料-涤棉成分",
    "款式",
    "加标方式",
    "尺寸类型",
    "当前决策状态",
    "过滤后候选数",
    "推荐物料编码",
    "原型综合分数",
    "Top1与Top2差值",
    "主要问题说明",
    "正确物料编码",
    "审核结论",
    "需要修正的订单字段",
    "正确字段值",
    "审核备注",
    "审核人",
    "审核日期",
]

CANDIDATE_HEADERS = [
    "审核序号",
    "来源文件",
    "订单行号",
    "候选排名",
    "物料编码",
    "原型综合分数",
    "结构化分数",
    "原始向量分数",
    "重复文本组数量",
    "候选物料名称",
    "候选规格",
    "候选颜色",
    "候选面料",
    "候选成分",
    "候选款式",
    "候选加标方式",
    "候选尺寸类型",
    "规格状态",
    "颜色状态",
    "面料状态",
    "成分状态",
    "款式状态",
    "加标方式状态",
    "尺寸类型状态",
    "候选说明",
]

MATERIAL_INDEX_HEADERS = [
    "物料编码",
    "物料名称",
    "规格",
    "颜色",
    "面料",
    "成分",
    "款式",
]

REVIEW_CONCLUSIONS = [
    "推荐编码正确",
    "Top候选中其他编码正确",
    "Top候选外编码正确",
    "物料库不存在对应物料",
    "订单字段解析错误",
    "物料主数据有误",
    "信息不足无法确认",
    "需要补充新的区分字段",
]

STATUS_LABELS = {
    "exact_match": "完全一致",
    "equivalent_match": "等价一致",
    "partial_match": "部分匹配",
    "no_match": "不一致",
    "missing_query": "订单信息缺失",
    "missing_candidate": "物料信息缺失",
    "not_comparable": "无法比较",
    "hard_conflict": "明确冲突",
}

FIELD_TO_STATUS_HEADER = {
    "spec": "规格状态",
    "color": "颜色状态",
    "fabric": "面料状态",
    "composition": "成分状态",
    "style": "款式状态",
    "label_method": "加标方式状态",
    "size_type": "尺寸类型状态",
}

QUERY_TO_REVIEW = {
    "spec": "规格",
    "color": "颜色",
    "fabric": "面料",
    "composition": "面料-涤棉成分",
    "style": "款式",
    "label_method": "加标方式",
    "size_type": "尺寸类型",
}


class ReviewWorkbookError(BeddingOrderParserError):
    """Raised when the Gate 3B-D review workbook cannot be built safely."""


@dataclass(frozen=True)
class ReviewWorkbookResult:
    output_path: Path
    review_records: int
    recommended_codes: int
    no_candidate: int
    ambiguous_tie: int
    insufficient_evidence: int
    unique_best_candidate: int
    candidate_detail_rows: int


@dataclass(frozen=True)
class MaterialIndexRow:
    material_code: str
    material_name: str
    spec: str
    color: str
    fabric: str
    composition: str
    style: str


def build_review_workbook(
    candidates_path: str | Path,
    summary_path: str | Path,
    store_path: str | Path,
    output_path: str | Path,
    *,
    overwrite: bool = False,
) -> ReviewWorkbookResult:
    """Build a workbook for business-owner review from existing prototype outputs."""
    candidates_file = Path(candidates_path).expanduser().resolve()
    summary_file = Path(summary_path).expanduser().resolve()
    store_file = Path(store_path).expanduser().resolve()
    output_file = Path(output_path).expanduser().resolve()
    if output_file.exists() and not overwrite:
        raise ReviewWorkbookError(f"Review workbook already exists: {output_file}")

    payload = _read_json_object(candidates_file)
    summary = _read_json_object(summary_file)
    records = payload.get("records")
    if not isinstance(records, list):
        raise ReviewWorkbookError("Candidate payload does not contain records.")
    _validate_summary_contract(records, summary)

    material_rows = _load_material_index(store_file)
    materials_by_code = {row.material_code: row for row in material_rows}
    formal_lookup = _load_formal_lookup(candidates_file, records)

    workbook = Workbook()
    review_ws = workbook.active
    review_ws.title = REVIEW_SHEET
    candidate_ws = workbook.create_sheet(CANDIDATE_SHEET)
    issue_ws = workbook.create_sheet(ISSUE_SHEET)
    instruction_ws = workbook.create_sheet(INSTRUCTION_SHEET)
    index_ws = workbook.create_sheet(MATERIAL_INDEX_SHEET)
    index_ws.sheet_state = "hidden"

    _write_review_sheet(review_ws, records, formal_lookup)
    candidate_rows = _write_candidate_sheet(candidate_ws, records, materials_by_code)
    _write_issue_sheet(issue_ws, records)
    _write_instruction_sheet(instruction_ws)
    _write_material_index(index_ws, material_rows)
    _apply_workbook_style(workbook)

    _write_workbook_atomically(workbook, output_file, overwrite=overwrite)
    decisions = Counter(str(record["decision"]["status"]) for record in records)
    return ReviewWorkbookResult(
        output_path=output_file,
        review_records=len(records),
        recommended_codes=sum(bool(record.get("candidates")) for record in records),
        no_candidate=decisions["no_candidate"],
        ambiguous_tie=decisions["ambiguous_tie"],
        insufficient_evidence=decisions["insufficient_evidence"],
        unique_best_candidate=decisions["unique_best_candidate"],
        candidate_detail_rows=candidate_rows,
    )


def _read_json_object(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ReviewWorkbookError(f"Unable to read JSON object {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise ReviewWorkbookError(f"Expected JSON object: {path}")
    return payload


def _validate_summary_contract(records: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    summary_values = summary.get("summary")
    if not isinstance(summary_values, dict):
        raise ReviewWorkbookError("Summary payload does not contain summary object.")
    decisions = Counter(str(record.get("decision", {}).get("status", "")) for record in records)
    if int(summary_values.get("order_records", -1)) != len(records):
        raise ReviewWorkbookError("Candidate record count does not match summary.")
    if int(summary_values.get("records_with_candidates", -1)) != sum(
        bool(record.get("candidates")) for record in records
    ):
        raise ReviewWorkbookError("Records-with-candidates count does not match summary.")
    for status in ("unique_best_candidate", "ambiguous_tie", "insufficient_evidence", "no_candidate"):
        expected = int(summary_values.get("decision_statuses", {}).get(status, 0))
        if decisions[status] != expected:
            raise ReviewWorkbookError(f"Decision count mismatch for {status}.")
    actions = {str(record.get("decision", {}).get("action", "")) for record in records}
    if actions != {"manual_review"}:
        raise ReviewWorkbookError("All prototype actions must be manual_review.")


def _load_material_index(store_path: Path) -> list[MaterialIndexRow]:
    if not store_path.exists():
        raise ReviewWorkbookError(f"Material store not found: {store_path}")
    with sqlite3.connect(store_path) as connection:
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            """
            SELECT material_code, material_name_raw, spec_raw, color_raw, fabric_raw,
                   composition_raw, style_raw
            FROM materials
            ORDER BY material_code
            """
        ).fetchall()
    return [
        MaterialIndexRow(
            material_code=str(row["material_code"]),
            material_name=str(row["material_name_raw"]),
            spec=str(row["spec_raw"]),
            color=str(row["color_raw"]),
            fabric=str(row["fabric_raw"]),
            composition=str(row["composition_raw"]),
            style=str(row["style_raw"]),
        )
        for row in rows
    ]


def _load_formal_lookup(candidates_path: Path, records: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    output_root = candidates_path.parent.parent
    result_root = output_root / "gate2d_validation" / "all_results"
    lookup: dict[tuple[str, str], dict[str, Any]] = {}
    for result_name in sorted({str(record.get("result_json", "")) for record in records}):
        if not result_name:
            continue
        result_path = result_root / result_name
        if not result_path.exists():
            continue
        payload = json.loads(result_path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            continue
        for row in payload:
            if isinstance(row, dict):
                lookup[(result_name, str(row.get("行号", "")))] = row
    return lookup


def _write_review_sheet(ws, records: list[dict[str, Any]], formal_lookup: dict[tuple[str, str], dict[str, Any]]) -> None:
    ws.append(REVIEW_HEADERS)
    for index, record in enumerate(records, start=1):
        query = _dict(record.get("query"))
        decision = _dict(record.get("decision"))
        retrieval = _dict(record.get("retrieval"))
        formal = formal_lookup.get((str(record.get("result_json", "")), str(record.get("行号", ""))), {})
        top = _top_candidate(record)
        ws.append(
            [
                index,
                record.get("source_file", ""),
                record.get("sheet", ""),
                record.get("行号", ""),
                formal.get("客户", ""),
                formal.get("物料名称", query.get("product_category", "")),
                query.get("spec", ""),
                query.get("color", ""),
                query.get("fabric", ""),
                query.get("composition", ""),
                query.get("style", ""),
                query.get("label_method", ""),
                query.get("size_type", ""),
                decision.get("status", ""),
                retrieval.get("post_filter_candidates", 0),
                top.get("material_code", ""),
                top.get("prototype_match_score", ""),
                decision.get("top1_margin", ""),
                _problem_summary(record),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
    if ws.max_row >= 2:
        options = ','.join(REVIEW_CONCLUSIONS)
        conclusion_validation = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
        ws.add_data_validation(conclusion_validation)
        conclusion_validation.add(f"U2:U{ws.max_row}")
        code_validation = DataValidation(
            type="custom",
            formula1=f'=OR($T2="",COUNTIF(\'{MATERIAL_INDEX_SHEET}\'!$A:$A,$T2)>0)',
            allow_blank=True,
        )
        ws.add_data_validation(code_validation)
        code_validation.add(f"T2:T{ws.max_row}")
    ws["Q1"].comment = Comment("原型分数，仅供候选排序，不代表正式匹配率。", "Codex")


def _write_candidate_sheet(ws, records: list[dict[str, Any]], materials_by_code: dict[str, MaterialIndexRow]) -> int:
    ws.append(CANDIDATE_HEADERS)
    row_count = 0
    for review_index, record in enumerate(records, start=1):
        for candidate in list(record.get("candidates", []))[:10]:
            code = str(candidate.get("material_code", ""))
            material = materials_by_code.get(code)
            fields = _dict(candidate.get("fields"))
            ws.append(
                [
                    review_index,
                    record.get("source_file", ""),
                    record.get("行号", ""),
                    candidate.get("rank", ""),
                    code,
                    candidate.get("prototype_match_score", ""),
                    candidate.get("structured_score", ""),
                    candidate.get("vector_score", ""),
                    candidate.get("duplicate_group_size", 1),
                    material.material_name if material else "",
                    material.spec if material else _field_value(fields, "spec", "candidate_value"),
                    material.color if material else _field_value(fields, "color", "candidate_value"),
                    material.fabric if material else _field_value(fields, "fabric", "candidate_value"),
                    material.composition if material else _field_value(fields, "composition", "candidate_value"),
                    material.style if material else _field_value(fields, "style", "candidate_value"),
                    _field_value(fields, "label_method", "candidate_value"),
                    _field_value(fields, "size_type", "candidate_value"),
                    _status_label(fields, "spec"),
                    _status_label(fields, "color"),
                    _status_label(fields, "fabric"),
                    _status_label(fields, "composition"),
                    _status_label(fields, "style"),
                    _status_label(fields, "label_method"),
                    _status_label(fields, "size_type"),
                    _candidate_note(candidate),
                ]
            )
            row_count += 1
    ws["F1"].comment = Comment("原型分数，仅供候选排序，不代表正式匹配率。", "Codex")
    return row_count


def _write_issue_sheet(ws, records: list[dict[str, Any]]) -> None:
    _append_section(
        ws,
        "A. 无候选记录",
        ["审核序号", "来源文件", "行号", "订单关键字段", "召回候选数", "主要硬冲突字段", "负责人需要回答的问题"],
    )
    for review_index, record in enumerate(records, start=1):
        if record.get("decision", {}).get("status") != "no_candidate":
            continue
        ws.append(
            [
                review_index,
                record.get("source_file", ""),
                record.get("行号", ""),
                _query_summary(record),
                record.get("retrieval", {}).get("union_candidates", 0),
                _hard_conflict_summary(record),
                "ERP是否存在该规格/颜色/面料组合的物料？若不存在请选择无对应物料。",
            ]
        )
    ws.append([])

    _append_section(
        ws,
        "B. 并列歧义记录",
        ["审核序号", "来源文件", "行号", "并列物料编码", "完全相同字段", "存在差异字段", "需要增加的业务区分信息"],
    )
    for review_index, record in enumerate(records, start=1):
        if record.get("decision", {}).get("status") != "ambiguous_tie":
            continue
        group = _top_duplicate_group(record)
        ws.append(
            [
                review_index,
                record.get("source_file", ""),
                record.get("行号", ""),
                ", ".join(group.get("duplicate_material_codes", [])),
                ", ".join(group.get("identical_fields", [])),
                ", ".join(group.get("differing_fields", [])),
                ", ".join(group.get("required_business_evidence", [])),
            ]
        )
    ws.append([])

    _append_section(
        ws,
        "C. 证据不足记录",
        ["审核序号", "来源文件", "行号", "缺少订单字段", "缺少候选主数据", "当前可比较字段数量", "负责人需要补充或确认的信息"],
    )
    for review_index, record in enumerate(records, start=1):
        if record.get("decision", {}).get("status") != "insufficient_evidence":
            continue
        ws.append(
            [
                review_index,
                record.get("source_file", ""),
                record.get("行号", ""),
                ", ".join(_missing_query_fields(record)),
                ", ".join(_missing_candidate_fields(record)),
                record.get("decision", {}).get("comparable_field_count", 0),
                "请确认正确物料编码，或补充可区分的规格、面料、成分、密度、款式、加标方式、尺寸类型。",
            ]
        )

def _write_instruction_sheet(ws) -> None:
    instructions = [
        "本表共有49条订单记录。",
        "推荐物料编码只是算法候选，不是正式结果。",
        "请优先审核红色和黄色记录。",
        "正确物料编码必须来自现有ERP物料主数据。",
        "如果ERP中不存在，请选择“物料库不存在对应物料”。",
        "如果订单字段提取错误，请填写错误字段及正确值。",
        "如果需要额外区分字段，请在备注中说明。",
        "不要修改系统生成列名和审核序号。",
        "审核完成后保存原Excel并返回项目组。",
        "建议审核顺序：1. 6条无候选；2. 3条并列歧义；3. 26条证据不足；4. 14条唯一最高分候选。",
        "原型分数仅供候选排序，不代表正式匹配率。",
    ]
    ws.append(["填写说明"])
    for line in instructions:
        ws.append([line])


def _write_material_index(ws, material_rows: list[MaterialIndexRow]) -> None:
    ws.append(MATERIAL_INDEX_HEADERS)
    for row in material_rows:
        ws.append([
            row.material_code,
            row.material_name,
            row.spec,
            row.color,
            row.fabric,
            row.composition,
            row.style,
        ])


def _apply_workbook_style(workbook: Workbook) -> None:
    dark = PatternFill("solid", fgColor="1F2937")
    gray = PatternFill("solid", fgColor="E5E7EB")
    yellow = PatternFill("solid", fgColor="FEF3C7")
    green = PatternFill("solid", fgColor="DCFCE7")
    red = PatternFill("solid", fgColor="FEE2E2")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in workbook.worksheets:
        if ws.max_row < 1:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = dark
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _fit_widths(ws)

    review = workbook[REVIEW_SHEET]
    for row in review.iter_rows(min_row=2):
        for cell in row[:19]:
            cell.fill = gray
        for cell in row[19:26]:
            cell.fill = yellow
        status = str(row[13].value or "")
        if status == "unique_best_candidate":
            row[13].fill = green
        elif status in {"ambiguous_tie", "insufficient_evidence"}:
            row[13].fill = yellow
        elif status == "no_candidate":
            row[13].fill = red

    candidate = workbook[CANDIDATE_SHEET]
    status_columns = range(18, 25)
    for row in candidate.iter_rows(min_row=2):
        for index in status_columns:
            cell = row[index - 1]
            value = str(cell.value or "")
            if value in {"完全一致", "等价一致"}:
                cell.fill = green
            elif value in {"部分匹配", "无法比较", "订单信息缺失", "物料信息缺失"}:
                cell.fill = yellow
            elif value in {"不一致", "明确冲突"}:
                cell.fill = red

    for row in workbook[ISSUE_SHEET].iter_rows():
        if row and isinstance(row[0].value, str) and row[0].value.startswith(("A.", "B.", "C.")):
            for cell in row:
                cell.fill = dark
                cell.font = white_font
        elif row and row[0].value in {"审核序号", "来源文件", "行号"}:
            for cell in row:
                cell.fill = gray
                cell.font = header_font

    workbook[INSTRUCTION_SHEET].column_dimensions["A"].width = 110

    max_index_row = workbook[MATERIAL_INDEX_SHEET].max_row
    if max_index_row >= 2:
        workbook[REVIEW_SHEET].auto_filter.ref = f"A1:Z{workbook[REVIEW_SHEET].max_row}"
        workbook[MATERIAL_INDEX_SHEET].auto_filter.ref = f"A1:G{max_index_row}"

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False


def _fit_widths(ws) -> None:
    max_width = 46 if ws.title != INSTRUCTION_SHEET else 110
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = min(max_width, max(10, max(len(str(cell.value or "")) for cell in column_cells[:200]) + 2))
        ws.column_dimensions[letter].width = width


def _write_workbook_atomically(workbook: Workbook, output_path: Path, *, overwrite: bool) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.parent / f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    backup_path: Path | None = None
    try:
        workbook.save(temp_path)
        if output_path.exists():
            if not overwrite:
                raise ReviewWorkbookError(f"Review workbook already exists: {output_path}")
            backup_path = output_path.with_name(f".{output_path.name}.{uuid.uuid4().hex}.backup")
            output_path.rename(backup_path)
        temp_path.rename(output_path)
        if backup_path is not None:
            backup_path.unlink()
    except Exception:
        temp_path.unlink(missing_ok=True)
        if backup_path is not None and backup_path.exists() and not output_path.exists():
            backup_path.rename(output_path)
        raise
    finally:
        workbook.close()


def _append_section(ws, title: str, headers: list[str]) -> None:
    ws.append([title])
    ws.append(headers)


def _top_candidate(record: dict[str, Any]) -> dict[str, Any]:
    candidates = record.get("candidates")
    if isinstance(candidates, list) and candidates:
        candidate = candidates[0]
        if isinstance(candidate, dict):
            return candidate
    return {}


def _top_duplicate_group(record: dict[str, Any]) -> dict[str, Any]:
    top = _top_candidate(record)
    group = top.get("duplicate_group")
    if isinstance(group, dict):
        return group
    candidates = record.get("candidates", [])
    if isinstance(candidates, list):
        rank = top.get("rank")
        return {
            "duplicate_material_codes": [str(item.get("material_code", "")) for item in candidates if item.get("rank") == rank],
            "identical_fields": [],
            "differing_fields": [],
            "required_business_evidence": ["需要业务负责人确认并列物料编码的区分依据"],
        }
    return {}


def _problem_summary(record: dict[str, Any]) -> str:
    decision = _dict(record.get("decision"))
    status = str(decision.get("status", ""))
    reason = str(decision.get("reason", ""))
    if status == "no_candidate":
        return f"无候选；硬冲突字段：{_hard_conflict_summary(record)}；{reason}"
    if status == "ambiguous_tie":
        group = _top_duplicate_group(record)
        codes = ", ".join(group.get("duplicate_material_codes", []))
        return f"并列歧义：{codes}；{reason}"
    if status == "insufficient_evidence":
        return f"证据不足；缺少订单字段：{', '.join(_missing_query_fields(record)) or '无'}；{reason}"
    return reason


def _hard_conflict_summary(record: dict[str, Any]) -> str:
    conflicts = _dict(record.get("retrieval", {}).get("hard_conflicts_by_field"))
    if not conflicts:
        return "无"
    return ", ".join(f"{key}:{value}" for key, value in conflicts.items())


def _query_summary(record: dict[str, Any]) -> str:
    query = _dict(record.get("query"))
    values = [f"{label}:{query.get(key, '')}" for key, label in QUERY_TO_REVIEW.items()]
    return "；".join(values)


def _missing_query_fields(record: dict[str, Any]) -> list[str]:
    query = _dict(record.get("query"))
    return [label for key, label in QUERY_TO_REVIEW.items() if not str(query.get(key, ""))]


def _missing_candidate_fields(record: dict[str, Any]) -> list[str]:
    top = _top_candidate(record)
    fields = _dict(top.get("fields"))
    labels: list[str] = []
    for key, label in QUERY_TO_REVIEW.items():
        status = str(_dict(fields.get(key)).get("status", ""))
        if status == "missing_candidate":
            labels.append(label)
    return labels


def _candidate_note(candidate: dict[str, Any]) -> str:
    notes = []
    if candidate.get("ambiguous_duplicate_group"):
        notes.append("重复文本组")
    fields = _dict(candidate.get("fields"))
    no_match = [label for key, label in QUERY_TO_REVIEW.items() if _dict(fields.get(key)).get("status") == "no_match"]
    if no_match:
        notes.append("不一致字段：" + ", ".join(no_match))
    return "；".join(notes)


def _status_label(fields: dict[str, Any], field_name: str) -> str:
    return STATUS_LABELS.get(str(_dict(fields.get(field_name)).get("status", "")), "")


def _field_value(fields: dict[str, Any], field_name: str, key: str) -> str:
    return str(_dict(fields.get(field_name)).get(key, ""))


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}

