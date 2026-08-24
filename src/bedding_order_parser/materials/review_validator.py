"""Validate the Gate 3B-D material review workbook without modifying it."""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bedding_order_parser.exceptions import BeddingOrderParserError
from bedding_order_parser.materials.loader import compute_sha256
from bedding_order_parser.materials.review_workbook import (
    CANDIDATE_HEADERS,
    CANDIDATE_SHEET,
    ISSUE_SHEET,
    REVIEW_CONCLUSIONS,
    REVIEW_HEADERS,
    REVIEW_SHEET,
)


class ReviewValidationError(BeddingOrderParserError):
    """Raised when a workbook cannot be inspected for review validation."""


@dataclass(frozen=True)
class ValidationIssue:
    level: str
    audit_id: str
    message: str

    def to_dict(self) -> dict[str, str]:
        return {"level": self.level, "audit_id": self.audit_id, "message": self.message}


@dataclass(frozen=True)
class ReviewValidationResult:
    workbook_path: Path
    checked_rows: int
    errors: list[ValidationIssue] = field(default_factory=list)
    warnings: list[ValidationIssue] = field(default_factory=list)
    sha256_before: str = ""
    sha256_after: str = ""

    @property
    def ok(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "checked_rows": self.checked_rows,
            "sha256_before": self.sha256_before,
            "sha256_after": self.sha256_after,
            "errors": [issue.to_dict() for issue in self.errors],
            "warnings": [issue.to_dict() for issue in self.warnings],
        }


def validate_review_workbook(workbook_path: str | Path, store_path: str | Path) -> ReviewValidationResult:
    """Validate reviewer-filled fields against SQLite and the workbook candidate list."""
    workbook_file = Path(workbook_path).expanduser().resolve()
    store_file = Path(store_path).expanduser().resolve()
    if not workbook_file.exists():
        raise ReviewValidationError(f"Workbook not found: {workbook_file}")
    if not store_file.exists():
        raise ReviewValidationError(f"Material store not found: {store_file}")

    sha_before = compute_sha256(workbook_file)
    material_codes = _load_material_codes(store_file)
    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    try:
        if any(sheet not in workbook.sheetnames for sheet in (REVIEW_SHEET, CANDIDATE_SHEET, ISSUE_SHEET)):
            raise ReviewValidationError("Workbook is missing required review sheets.")
        review_ws = workbook[REVIEW_SHEET]
        candidate_ws = workbook[CANDIDATE_SHEET]
        issue_ws = workbook[ISSUE_SHEET]
        _validate_headers(review_ws, REVIEW_HEADERS, REVIEW_SHEET)
        _validate_headers(candidate_ws, CANDIDATE_HEADERS, CANDIDATE_SHEET)
        candidate_codes = _candidate_codes_by_audit_id(candidate_ws)
        expected_ids = set(candidate_codes) | _issue_audit_ids(issue_ws)
        review_rows = list(_review_rows(review_ws))
        errors: list[ValidationIssue] = []
        warnings: list[ValidationIssue] = []
        _validate_audit_ids(review_rows, errors, expected_ids)
        for row in review_rows:
            _validate_review_row(row, material_codes, candidate_codes, errors, warnings)
    finally:
        workbook.close()
    sha_after = compute_sha256(workbook_file)
    return ReviewValidationResult(
        workbook_path=workbook_file,
        checked_rows=len(review_rows),
        errors=errors,
        warnings=warnings,
        sha256_before=sha_before,
        sha256_after=sha_after,
    )


def _load_material_codes(store_path: Path) -> set[str]:
    with sqlite3.connect(store_path) as connection:
        return {str(row[0]) for row in connection.execute("SELECT material_code FROM materials")}


def _validate_headers(ws, expected: list[str], sheet_name: str) -> None:
    actual = [str(cell.value or "") for cell in ws[1][: len(expected)]]
    if actual != expected:
        raise ReviewValidationError(f"Invalid header contract in {sheet_name}.")


def _candidate_codes_by_audit_id(ws) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        audit_id = _text(row[0])
        code = _text(row[4])
        if not audit_id or not code:
            continue
        result.setdefault(audit_id, set()).add(code)
    return result

def _issue_audit_ids(ws) -> set[str]:
    ids: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = _text(row[0])
        if first.isdigit():
            ids.add(first)
    return ids

def _review_rows(ws) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, max_col=len(REVIEW_HEADERS), values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append({header: _text(value) for header, value in zip(REVIEW_HEADERS, row, strict=True)})
    return rows


def _validate_audit_ids(
    rows: list[dict[str, str]],
    errors: list[ValidationIssue],
    expected_ids: set[str],
) -> None:
    ids = [row["审核序号"] for row in rows]
    seen: set[str] = set()
    for audit_id in ids:
        if audit_id in seen:
            errors.append(ValidationIssue("error", audit_id, "审核序号重复。"))
        seen.add(audit_id)
    numeric_ids: list[int] = []
    for audit_id in ids:
        try:
            numeric_ids.append(int(audit_id))
        except ValueError:
            errors.append(ValidationIssue("error", audit_id, "审核序号不是整数。"))
    if numeric_ids:
        expected = set(range(1, max(numeric_ids) + 1))
        missing = sorted(expected - set(numeric_ids))
        for audit_id in missing:
            errors.append(ValidationIssue("error", str(audit_id), "审核序号缺失，疑似订单行被删除。"))
    review_ids = set(ids)
    for audit_id in sorted(expected_ids - review_ids, key=_audit_id_sort_key):
        errors.append(ValidationIssue("error", audit_id, "候选明细中存在但审核清单缺失，疑似订单行被删除。"))


def _audit_id_sort_key(value: str) -> tuple[int, int | str]:
    if value.isdigit():
        return (0, int(value))
    return (1, value)

def _validate_review_row(
    row: dict[str, str],
    material_codes: set[str],
    candidate_codes: dict[str, set[str]],
    errors: list[ValidationIssue],
    warnings: list[ValidationIssue],
) -> None:
    audit_id = row["审核序号"]
    correct_code = row["正确物料编码"]
    conclusion = row["审核结论"]
    recommended = row["推荐物料编码"]
    fix_fields = row["需要修正的订单字段"]

    if conclusion and conclusion not in REVIEW_CONCLUSIONS:
        errors.append(ValidationIssue("error", audit_id, f"审核结论不在允许选项内：{conclusion}"))
    if correct_code and correct_code not in material_codes:
        errors.append(ValidationIssue("error", audit_id, f"正确物料编码不存在于SQLite：{correct_code}"))
    if conclusion == "推荐编码正确":
        if not correct_code:
            errors.append(ValidationIssue("error", audit_id, "选择推荐编码正确时必须填写正确物料编码。"))
        elif correct_code != recommended:
            errors.append(ValidationIssue("error", audit_id, "推荐编码正确时，正确物料编码必须等于推荐物料编码。"))
    if conclusion == "Top候选中其他编码正确":
        if not correct_code:
            errors.append(ValidationIssue("error", audit_id, "选择Top候选中其他编码正确时必须填写正确物料编码。"))
        elif correct_code not in candidate_codes.get(audit_id, set()):
            errors.append(ValidationIssue("error", audit_id, "正确物料编码不在该订单Top 10候选中。"))
    if conclusion == "Top候选外编码正确":
        if not correct_code:
            errors.append(ValidationIssue("error", audit_id, "选择Top候选外编码正确时必须填写正确物料编码。"))
        elif correct_code in candidate_codes.get(audit_id, set()):
            errors.append(ValidationIssue("error", audit_id, "Top候选外正确编码不能出现在该订单Top 10候选中。"))
    if conclusion == "物料库不存在对应物料" and correct_code:
        errors.append(ValidationIssue("error", audit_id, "物料库不存在对应物料时不得填写正确物料编码。"))
    if conclusion == "订单字段解析错误" and not fix_fields:
        errors.append(ValidationIssue("error", audit_id, "订单字段解析错误时必须填写需要修正的订单字段。"))
    if not conclusion and correct_code:
        warnings.append(ValidationIssue("warning", audit_id, "已填写正确物料编码但未选择审核结论。"))


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


