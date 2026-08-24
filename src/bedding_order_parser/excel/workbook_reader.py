"""Validate and load PI workbooks without mutating input files."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from bedding_order_parser.exceptions import InputFileError


ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


@dataclass(frozen=True)
class LoadedWorkbook:
    path: Path
    workbook: Workbook
    sha256: str


def compute_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_pi_workbook(path: Path) -> LoadedWorkbook:
    resolved = path.expanduser().resolve()
    if not resolved.exists():
        raise InputFileError(f"Input file does not exist: {path}")
    if not resolved.is_file():
        raise InputFileError(f"Input path is not a file: {path}")
    if resolved.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise InputFileError(f"Unsupported Excel extension: {resolved.suffix}")

    before_hash = compute_sha256(resolved)
    try:
        workbook = load_workbook(
            resolved,
            data_only=True,
            read_only=False,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl raises several concrete reader errors.
        raise InputFileError(f"Failed to open Excel workbook: {resolved.name}") from exc

    return LoadedWorkbook(path=resolved, workbook=workbook, sha256=before_hash)
