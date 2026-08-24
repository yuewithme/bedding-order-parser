"""Merged-cell value mapping for openpyxl worksheets."""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


CellPosition = tuple[int, int]


def build_merged_cell_value_map(worksheet: Worksheet) -> dict[CellPosition, Any]:
    merged_values: dict[CellPosition, Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left_value = worksheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, column)] = top_left_value
    return merged_values


def value_with_merged_fallback(
    worksheet: Worksheet,
    row: int,
    column: int,
    merged_values: dict[CellPosition, Any],
) -> Any:
    value = worksheet.cell(row=row, column=column).value
    if value is None:
        return merged_values.get((row, column))
    return value
