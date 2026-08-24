"""Convert a PI worksheet into cleaned table rows."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from bedding_order_parser.excel.merged_cells import (
    build_merged_cell_value_map,
    value_with_merged_fallback,
)
from bedding_order_parser.exceptions import WorkbookStructureError


HEADER_KEYWORDS = (
    "no",
    "item",
    "items",
    "description",
    "specification",
    "size",
    "dimension",
    "qty",
    "quantity",
    "unit price",
    "amount",
    "remark",
    "remarks",
    "序号",
    "品名",
    "描述",
    "规格",
    "工艺",
    "尺寸",
    "数量",
)

FOOTER_KEYWORDS = (
    "total",
    "subtotal",
    "grand total",
    "total value",
    "amount due",
    "bank",
    "payment",
    "terms",
    "signature",
    "seller",
    "合计",
    "总计",
    "付款",
)


@dataclass(frozen=True)
class ParsedRow:
    excel_row_number: int
    values: dict[str, str]
    raw_values: list[str]
    section_note: str = ""


@dataclass(frozen=True)
class ParsedTable:
    sheet_title: str
    rows: list[list[str]]
    number_formats: list[list[str]]
    header_index: int
    headers: list[str]
    data_rows: list[ParsedRow]
    pre_header_rows: list[list[str]]
    post_table_rows: list[list[str]]


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.strftime("%Y-%m-%d 00:00:00")
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\t", " ")
    return re.sub(r"\s{2,}", " ", text).strip()


def worksheet_to_rows(worksheet: Worksheet, max_initial_cols: int = 20) -> list[list[str]]:
    rows, _number_formats = worksheet_to_rows_and_number_formats(worksheet, max_initial_cols)
    return rows


def worksheet_to_rows_and_number_formats(
    worksheet: Worksheet,
    max_initial_cols: int = 20,
) -> tuple[list[list[str]], list[list[str]]]:
    max_column = min(worksheet.max_column, max_initial_cols)
    merged_values = build_merged_cell_value_map(worksheet)
    rows: list[list[str]] = []
    number_formats: list[list[str]] = []
    last_non_empty_row = -1
    last_non_empty_col = -1

    for row_number in range(1, worksheet.max_row + 1):
        row_values: list[str] = []
        row_formats: list[str] = []
        for column in range(1, max_column + 1):
            value = value_with_merged_fallback(worksheet, row_number, column, merged_values)
            cleaned = clean_cell_value(value)
            row_values.append(cleaned)
            number_format = clean_cell_value(worksheet.cell(row_number, column).number_format)
            row_formats.append("" if number_format == "General" else number_format)
            if cleaned:
                last_non_empty_row = row_number - 1
                last_non_empty_col = max(last_non_empty_col, column - 1)
        rows.append(row_values)
        number_formats.append(row_formats)

    if last_non_empty_row < 0:
        return [], []

    trimmed_width = last_non_empty_col + 1
    trimmed_rows = [row[:trimmed_width] for row in rows[: last_non_empty_row + 1]]
    trimmed_formats = [row[:trimmed_width] for row in number_formats[: last_non_empty_row + 1]]
    return trimmed_rows, trimmed_formats


def _row_header_score(row: list[str]) -> int:
    normalized_cells = [cell.casefold() for cell in row if cell]
    score = 0
    for keyword in HEADER_KEYWORDS:
        if any(keyword in cell for cell in normalized_cells):
            score += 1
    return score


def locate_header_row(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows):
        if _row_header_score(row) >= 3:
            return index

    for index, row in enumerate(rows):
        if _is_numbered_item_row(row):
            return max(index - 1, 0)

    raise WorkbookStructureError("Could not locate a PI item table header.")


def _locate_header_region(rows: list[list[str]]) -> tuple[int, int]:
    header_index = locate_header_row(rows)
    header_end = header_index
    for next_index in range(header_index + 1, min(header_index + 3, len(rows))):
        row = rows[next_index]
        if _is_blank(row) or _is_numbered_item_row(row):
            break
        if _looks_like_header_continuation(row):
            header_end = next_index
            continue
        break
    return header_index, header_end


def _looks_like_header_continuation(row: list[str]) -> bool:
    return _row_header_score(row) >= 2


def _combine_header_rows(header_rows: list[list[str]]) -> list[str]:
    width = max((len(row) for row in header_rows), default=0)
    headers: list[str] = []
    for column_index in range(width):
        parts: list[str] = []
        seen_parts: set[str] = set()
        for row in header_rows:
            cell = row[column_index] if column_index < len(row) else ""
            if not cell:
                continue
            normalized = cell.casefold()
            if normalized in seen_parts:
                continue
            parts.append(cell)
            seen_parts.add(normalized)
        headers.append(" ".join(parts) if parts else f"Column {column_index + 1}")
    return headers


def _is_numbered_item_row(row: list[str]) -> bool:
    if not row:
        return False
    first = row[0].strip()
    return re.fullmatch(r"\d+(?:\.0)?", first) is not None


def _is_blank(row: list[str]) -> bool:
    return not any(cell.strip() for cell in row)


def _is_footer_row(row: list[str]) -> bool:
    text = _row_text(row).casefold()
    if not text:
        return False
    first_non_empty = next((cell.strip().casefold() for cell in row if cell.strip()), "")
    return first_non_empty.startswith(FOOTER_KEYWORDS) or text.startswith(FOOTER_KEYWORDS)


def _row_text(row: list[str]) -> str:
    return " ".join(cell for cell in row if cell)


def parse_table(worksheet: Worksheet) -> ParsedTable:
    rows, number_formats = worksheet_to_rows_and_number_formats(worksheet)
    if not rows:
        raise WorkbookStructureError("The selected worksheet is empty.")

    header_index, header_end = _locate_header_region(rows)
    headers = _combine_header_rows(rows[header_index : header_end + 1])

    data_rows: list[ParsedRow] = []
    post_table_start = len(rows)
    pending_note = ""
    seen_item = False
    blank_after_items = 0
    first_blank_after_items: int | None = None

    for row_index in range(header_end + 1, len(rows)):
        row = rows[row_index]
        if _is_blank(row):
            if seen_item:
                if first_blank_after_items is None:
                    first_blank_after_items = row_index
                blank_after_items += 1
                if blank_after_items >= 2:
                    post_table_start = first_blank_after_items
                    break
            continue

        blank_after_items = 0
        first_blank_after_items = None

        if _is_footer_row(row):
            if seen_item:
                post_table_start = row_index
                break
            continue

        if _is_numbered_item_row(row):
            values = {
                headers[column_index]: row[column_index] if column_index < len(row) else ""
                for column_index in range(len(headers))
            }
            data_rows.append(
                ParsedRow(
                    excel_row_number=row_index + 1,
                    values=values,
                    raw_values=row,
                    section_note=pending_note,
                )
            )
            pending_note = ""
            seen_item = True
            continue

        pending_note = _row_text(row)

    if not data_rows:
        raise WorkbookStructureError("No numbered PI item rows were found.")

    return ParsedTable(
        sheet_title=worksheet.title,
        rows=rows,
        number_formats=number_formats,
        header_index=header_index,
        headers=headers,
        data_rows=data_rows,
        pre_header_rows=rows[:header_index],
        post_table_rows=rows[post_table_start:],
    )
