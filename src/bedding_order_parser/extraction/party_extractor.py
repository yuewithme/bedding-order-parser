"""Extract buyer organization and seller contact from worksheet regions."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Iterable

from openpyxl.utils import get_column_letter

from bedding_order_parser.diagnostics.models import (
    AMBIGUOUS,
    EXTRACTED,
    NORMALIZED,
    SOURCE_NOT_PROVIDED,
    FieldDiagnostic,
    SourceEvidence,
)
from bedding_order_parser.excel.table_parser import ParsedTable
from bedding_order_parser.normalization.field_normalizer import clean_text


BUYER = "buyer"
SELLER = "seller"

BUYER_LABELS: tuple[tuple[str, int], ...] = (
    ("to buyer", 1),
    ("invoice to", 1),
    ("bill to", 1),
    ("sold to", 1),
    ("buyer", 1),
    ("messers", 1),
    ("messrs", 1),
    ("m/s", 1),
    ("consignee", 2),
    ("customer", 2),
    ("to", 3),
)

SELLER_LABELS: tuple[str, ...] = (
    "sellers",
    "seller",
    "supplier",
    "exporter",
    "from",
)

SALES_LABELS: tuple[tuple[str, int], ...] = (
    ("contact person", 1),
    ("sales person", 2),
    ("salesperson", 2),
    ("merchandiser", 2),
    ("account manager", 2),
    ("sales", 3),
)

CONTACT_BOUNDARY = re.compile(
    r"\b(?:address|add|phone(?:\s+no)?|tel(?:ephone)?|mobile|mob|"
    r"e-?mail|registration\s+no|company\s+reg(?:istration)?(?:\s+no)?|"
    r"tax\s+id|fax|contact\s+person)\s*[:：.]?",
    flags=re.IGNORECASE,
)

LEGAL_SUFFIX = re.compile(
    r"^(.+?(?:"
    r"\bco\.,?\s*ltd\.?|\bcompany\s+limited|\bpvt\.?\s+ltd\.?|\binc\.?|"
    r"\bl\.?l\.?c\.?|\blimited|\bcorp(?:oration)?\.?|\best(?:ablishment)?\.?"
    r")(?:\s*\(branch[^)]*\)|\s+branch\s+[A-Za-z0-9-]+)?)"
    r"(?=\s|$)",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class PartyExtraction:
    customer: FieldDiagnostic
    salesperson: FieldDiagnostic
    buyer_contact: FieldDiagnostic


@dataclass(frozen=True)
class _Cell:
    row: int
    column: int
    value: str

    @property
    def coordinate(self) -> str:
        return f"{get_column_letter(self.column + 1)}{self.row + 1}"


@dataclass(frozen=True)
class _Candidate:
    value: str
    priority: int
    cells: tuple[str, ...]
    label: str
    normalized: bool = False


def extract_parties(table: ParsedTable) -> PartyExtraction:
    cells_by_row = _logical_cells(table.pre_header_rows)
    regions = _classify_regions(cells_by_row)
    customer = _extract_customer(table.sheet_title, cells_by_row, regions)
    salesperson, buyer_contact = _extract_contacts(table, cells_by_row, regions)
    return PartyExtraction(
        customer=customer,
        salesperson=salesperson,
        buyer_contact=buyer_contact,
    )


def _logical_cells(rows: list[list[str]]) -> list[list[_Cell]]:
    logical_rows: list[list[_Cell]] = []
    for row_index, row in enumerate(rows):
        seen: set[str] = set()
        cells: list[_Cell] = []
        for column_index, value in enumerate(row):
            text = clean_text(value)
            if not text:
                continue
            key = text.casefold()
            if key in seen:
                continue
            seen.add(key)
            cells.append(_Cell(row=row_index, column=column_index, value=text))
        logical_rows.append(cells)
    return logical_rows


def _classify_regions(cells_by_row: list[list[_Cell]]) -> dict[tuple[int, int], str]:
    regions: dict[tuple[int, int], str] = {}
    active_split: int | None = None
    vertical_region = ""

    for row_cells in cells_by_row:
        buyer_columns = [
            cell.column for cell in row_cells if _match_labeled_value(cell.value, BUYER_LABELS)
        ]
        seller_columns = [cell.column for cell in row_cells if _is_seller_anchor(cell.value)]
        sales_columns = [
            cell.column for cell in row_cells if _match_labeled_value(cell.value, SALES_LABELS)
        ]

        if buyer_columns and seller_columns:
            active_split = min(seller_columns)
            vertical_region = ""
        elif buyer_columns and sales_columns and min(sales_columns) > min(buyer_columns):
            active_split = min(sales_columns)
            vertical_region = ""
        elif active_split is None and vertical_region == BUYER and sales_columns:
            if min(sales_columns) >= 3:
                active_split = min(sales_columns)
        elif active_split is None:
            if seller_columns:
                vertical_region = SELLER
            elif buyer_columns:
                vertical_region = BUYER

        for cell in row_cells:
            if active_split is not None:
                region = SELLER if cell.column >= active_split else BUYER
            else:
                region = vertical_region
            regions[(cell.row, cell.column)] = region

    return regions


def _extract_customer(
    sheet_name: str,
    cells_by_row: list[list[_Cell]],
    regions: dict[tuple[int, int], str],
) -> FieldDiagnostic:
    candidates: list[_Candidate] = []
    for row_cells in cells_by_row:
        for cell in row_cells:
            matched = _match_labeled_value(cell.value, BUYER_LABELS)
            if not matched:
                continue
            priority, label, inline_value = matched
            for raw_value, value_cell in _buyer_values_near_label(
                cell,
                inline_value,
                cells_by_row,
                regions,
            ):
                value, normalized = _clean_company_candidate(raw_value)
                if not value:
                    continue
                source_cells = (cell.coordinate,) if value_cell.coordinate == cell.coordinate else (
                    cell.coordinate,
                    value_cell.coordinate,
                )
                candidates.append(
                    _Candidate(
                        value=value,
                        priority=priority,
                        cells=source_cells,
                        label=label,
                        normalized=normalized,
                    )
                )
                break
    return _resolve_candidates(
        candidates,
        sheet_name=sheet_name,
        region=BUYER,
        rule="buyer.organization",
        missing_message="未发现明确买方主体",
        conflict_message="发现多个冲突买方主体，需要人工复核",
    )


def _extract_contacts(
    table: ParsedTable,
    cells_by_row: list[list[_Cell]],
    regions: dict[tuple[int, int], str],
) -> tuple[FieldDiagnostic, FieldDiagnostic]:
    seller_candidates: list[_Candidate] = []
    buyer_candidates: list[_Candidate] = []

    for row_cells in cells_by_row:
        for cell in row_cells:
            region = regions.get((cell.row, cell.column), "")
            matched = _match_labeled_value(cell.value, SALES_LABELS)
            if matched:
                priority, label, inline_value = matched
                raw_value, value_cell = _value_near_label(
                    cell,
                    inline_value,
                    cells_by_row,
                    regions,
                    region,
                )
                value = _clean_person_candidate(raw_value)
                if value:
                    candidate = _Candidate(
                        value=value,
                        priority=priority,
                        cells=(cell.coordinate,)
                        if value_cell.coordinate == cell.coordinate
                        else (cell.coordinate, value_cell.coordinate),
                        label=label,
                    )
                    if region == SELLER:
                        seller_candidates.append(candidate)
                    elif region == BUYER:
                        buyer_candidates.append(candidate)

            if region == SELLER:
                embedded = _embedded_contact(cell.value)
                if embedded:
                    value, label = embedded
                    seller_candidates.append(
                        _Candidate(
                            value=value,
                            priority=1,
                            cells=(cell.coordinate,),
                            label=label,
                        )
                    )

    if not seller_candidates:
        seller_candidates.extend(_implicit_seller_contacts(cells_by_row, regions))
    if not seller_candidates:
        seller_candidates.extend(_signature_candidates(table))

    seller = _resolve_candidates(
        seller_candidates,
        sheet_name=table.sheet_title,
        region=SELLER,
        rule="seller.contact",
        missing_message="未发现明确卖方联系人",
        conflict_message="发现多个同优先级卖方联系人，需要人工复核",
    )
    buyer = _resolve_candidates(
        buyer_candidates,
        sheet_name=table.sheet_title,
        region=BUYER,
        rule="buyer.contact",
        missing_message="未发现明确买方联系人",
        conflict_message="发现多个冲突买方联系人",
    )
    return seller, buyer


def _value_near_label(
    label_cell: _Cell,
    inline_value: str,
    cells_by_row: list[list[_Cell]],
    regions: dict[tuple[int, int], str],
    expected_region: str,
) -> tuple[str, _Cell]:
    if clean_text(inline_value):
        return inline_value, label_cell

    same_row = cells_by_row[label_cell.row]
    for cell in same_row:
        if cell.column <= label_cell.column:
            continue
        if expected_region and regions.get((cell.row, cell.column), "") != expected_region:
            continue
        if _is_non_value_label(cell.value):
            continue
        return cell.value, cell

    for row_index in range(label_cell.row + 1, min(len(cells_by_row), label_cell.row + 5)):
        row_cells = cells_by_row[row_index]
        same_column = next((cell for cell in row_cells if cell.column == label_cell.column), None)
        candidates = ([same_column] if same_column else []) + [
            cell for cell in row_cells if cell is not same_column
        ]
        for cell in candidates:
            if expected_region and regions.get((cell.row, cell.column), "") != expected_region:
                continue
            if _is_non_value_label(cell.value):
                continue
            return cell.value, cell
    return "", label_cell


def _buyer_values_near_label(
    label_cell: _Cell,
    inline_value: str,
    cells_by_row: list[list[_Cell]],
    regions: dict[tuple[int, int], str],
) -> list[tuple[str, _Cell]]:
    values: list[tuple[str, _Cell]] = []
    if clean_text(inline_value):
        values.append((inline_value, label_cell))
        return values

    for cell in cells_by_row[label_cell.row]:
        if not (label_cell.column < cell.column <= label_cell.column + 3):
            continue
        if regions.get((cell.row, cell.column), "") != BUYER:
            continue
        if not _is_non_value_label(cell.value):
            values.append((cell.value, cell))

    for row_index in range(label_cell.row + 1, min(len(cells_by_row), label_cell.row + 5)):
        row_cells = cells_by_row[row_index]
        nearby = sorted(
            (
                cell
                for cell in row_cells
                if label_cell.column <= cell.column <= label_cell.column + 2
                and regions.get((cell.row, cell.column), "") == BUYER
                and not _is_non_value_label(cell.value)
            ),
            key=lambda cell: (abs(cell.column - label_cell.column), cell.column),
        )
        values.extend((cell.value, cell) for cell in nearby)
    return values


def _match_labeled_value(
    text: str,
    labels: tuple[tuple[str, int], ...],
) -> tuple[int, str, str] | None:
    stripped = clean_text(text)
    for label, priority in labels:
        label_pattern = r"\s+".join(re.escape(part) for part in label.split())
        match = re.match(
            rf"^{label_pattern}\s*(?:[:：]\s*(.*)|$)",
            stripped,
            flags=re.IGNORECASE,
        )
        if match:
            return priority, label, clean_text(match.group(1) or "")
    return None


def _is_seller_anchor(text: str) -> bool:
    stripped = clean_text(text)
    lowered = stripped.casefold()
    if any(marker in lowered for marker in ("canasin", "康乃馨")):
        return True
    for label in SELLER_LABELS:
        label_pattern = r"\s+".join(re.escape(part) for part in label.split())
        if re.match(rf"^{label_pattern}\s*(?:[:：]|$)", stripped, flags=re.IGNORECASE):
            return True
    return False


def _clean_company_candidate(value: str) -> tuple[str, bool]:
    original = clean_text(value)
    if not original:
        return "", False
    text = re.sub(r"^(?:company|organization)\s*[:：]\s*", "", original, flags=re.IGNORECASE)
    boundary = CONTACT_BOUNDARY.search(text)
    if boundary:
        text = text[: boundary.start()]
    text = text.strip(" ,;/:-")

    legal_match = LEGAL_SUFFIX.match(text)
    if legal_match and legal_match.end() < len(text):
        text = legal_match.group(1).strip(" ,;/:-")

    if not text or _is_non_company_value(text):
        return "", False
    normalized = text != original
    return text, normalized


def _is_non_company_value(value: str) -> bool:
    lowered = value.casefold().strip()
    if any(marker in lowered for marker in ("canasin", "康乃馨")):
        return True
    if lowered.startswith(("use hotel", "project", "contact person", "address", "phone", "tel")):
        return True
    if re.match(r"^(?:mr|mrs|ms|miss)\.?\s+\S+", value, flags=re.IGNORECASE):
        return True
    if _is_non_value_label(value):
        return True
    return False


def _clean_person_candidate(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.split(
        r"\b(?:tel|telephone|mobile|mob|e-?mail|email|fax)\s*[:：]?",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0].strip(" ,;/:-")
    if not text or any(marker in text.casefold() for marker in ("canasin", "company", " co.,", " ltd")):
        return ""
    if "@" in text or re.search(r"\d{4,}", text):
        return ""
    if len(text) > 60:
        return ""
    return text


def _embedded_contact(text: str) -> tuple[str, str] | None:
    match = re.search(
        r"\b(contact(?:\s+person)?|sales(?:\s+person)?|salesperson|"
        r"merchandiser|account\s+manager)\s*[:：]\s*"
        r"([A-Za-z][A-Za-z .'\-/]{0,55}?)"
        r"(?=\s+(?:tel|telephone|mobile|mob|e-?mail|email|fax)\b|$)",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    value = _clean_person_candidate(match.group(2))
    if not value:
        return None
    return value, match.group(1)


def _implicit_seller_contacts(
    cells_by_row: list[list[_Cell]],
    regions: dict[tuple[int, int], str],
) -> list[_Candidate]:
    candidates: list[_Candidate] = []
    pattern = re.compile(
        r"^([A-Za-z][A-Za-z .'\-]{1,45}?)\s+"
        r"(?:tel|telephone|mobile|mob|e-?mail|email)\s*[:：]",
        flags=re.IGNORECASE,
    )
    for row_cells in cells_by_row:
        for cell in row_cells:
            if regions.get((cell.row, cell.column), "") != SELLER:
                continue
            match = pattern.search(cell.value)
            if not match:
                continue
            value = _clean_person_candidate(match.group(1))
            if value:
                candidates.append(
                    _Candidate(
                        value=value,
                        priority=3,
                        cells=(cell.coordinate,),
                        label="seller contact line",
                    )
                )
    if candidates:
        return candidates

    for row_cells in cells_by_row:
        for cell in row_cells:
            if regions.get((cell.row, cell.column), "") != SELLER:
                continue
            value = _clean_person_candidate(cell.value)
            if value and _looks_like_person_name(value):
                candidates.append(
                    _Candidate(
                        value=value,
                        priority=3,
                        cells=(cell.coordinate,),
                        label="seller contact name",
                    )
                )
    return candidates


def _signature_candidates(table: ParsedTable) -> list[_Candidate]:
    candidates: list[_Candidate] = []
    for row_index, row in enumerate(table.rows[table.header_index + 1 :], table.header_index + 1):
        for column_index, value in enumerate(row):
            match = re.match(
                r"^(?:seller|authorized\s+signature|signed\s+by)\s*[:：]\s*(.+)$",
                value,
                flags=re.IGNORECASE,
            )
            if not match:
                continue
            person = _clean_person_candidate(match.group(1))
            if person:
                candidates.append(
                    _Candidate(
                        value=person,
                        priority=4,
                        cells=(f"{get_column_letter(column_index + 1)}{row_index + 1}",),
                        label="signature",
                    )
                )
    return candidates


def _resolve_candidates(
    candidates: Iterable[_Candidate],
    *,
    sheet_name: str,
    region: str,
    rule: str,
    missing_message: str,
    conflict_message: str,
) -> FieldDiagnostic:
    candidate_list = list(candidates)
    if not candidate_list:
        return FieldDiagnostic(
            value="",
            status=SOURCE_NOT_PROVIDED,
            source=SourceEvidence(sheet=sheet_name, region=region),
            rule=rule,
            message=missing_message,
        )

    top_priority = min(candidate.priority for candidate in candidate_list)
    top = [candidate for candidate in candidate_list if candidate.priority == top_priority]
    distinct: dict[str, _Candidate] = {}
    for candidate in top:
        distinct.setdefault(candidate.value.casefold(), candidate)
    if len(distinct) > 1:
        cells = tuple(dict.fromkeys(cell for candidate in top for cell in candidate.cells))
        return FieldDiagnostic(
            value="",
            status=AMBIGUOUS,
            source=SourceEvidence(
                sheet=sheet_name,
                cells=cells,
                region=region,
                label=top[0].label,
            ),
            rule=rule,
            message=conflict_message,
        )

    selected = next(iter(distinct.values()))
    return FieldDiagnostic(
        value=selected.value,
        status=NORMALIZED if selected.normalized else EXTRACTED,
        source=SourceEvidence(
            sheet=sheet_name,
            cells=selected.cells,
            region=region,
            label=selected.label,
        ),
        rule=rule,
        message="已从对应交易方区域提取",
    )


def _is_non_value_label(value: str) -> bool:
    text = clean_text(value)
    if not text:
        return True
    if _match_labeled_value(text, BUYER_LABELS) or _match_labeled_value(text, SALES_LABELS):
        return True
    if _is_seller_anchor(text):
        return True
    lowered = text.casefold()
    return lowered.startswith(
        (
            "address:",
            "address：",
            "phone:",
            "tel:",
            "mobile:",
            "email:",
            "e-mail:",
            "use hotel",
            "proforma invoice",
        )
    )


def _looks_like_person_name(value: str) -> bool:
    lowered = value.casefold().strip(" .:")
    if lowered in {
        "proforma invoice",
        "purchase order",
        "invoice",
        "buyer",
        "seller",
        "from",
    }:
        return False
    if _is_seller_anchor(value) or _match_labeled_value(value, BUYER_LABELS):
        return False
    if re.search(r"\b(?:address|phone|tel|email|invoice|date)\b", lowered):
        return False
    words = re.findall(r"[A-Za-z]+", value)
    return 1 <= len(words) <= 3 and all(len(word) >= 2 for word in words)
