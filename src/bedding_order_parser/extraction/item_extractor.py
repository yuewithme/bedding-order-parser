"""Extract duvet-cover rows, final results, and field diagnostics."""

from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bedding_order_parser.diagnostics.models import (
    DEFAULTED,
    DERIVED,
    EXTRACTED,
    NORMALIZED,
    NOT_IMPLEMENTED,
    SOURCE_NOT_PROVIDED,
    UNRECOGNIZED,
    FieldDiagnostic,
    SourceEvidence,
)
from bedding_order_parser.excel.table_parser import ParsedRow, ParsedTable
from bedding_order_parser.extraction.metadata_extractor import OrderMetadata
from bedding_order_parser.models.final_result import FINAL_FIELD_NAMES, FinalResult
from bedding_order_parser.normalization.field_normalizer import (
    normalize_component,
    normalize_color,
    normalize_fabric,
    normalize_labeling,
    normalize_quantity,
    normalize_row_note,
    normalize_size,
    normalize_size_type,
    normalize_style,
    normalize_yes_no_embroidery,
)


INCLUDE_KEYWORDS = (
    "duvet cover",
    "duvetcover",
    "dubet cover",
    "quilt cover",
    "comforter cover",
)

STRONG_EXCLUDE_KEYWORDS = (
    "bed sheet",
    "bedsheet",
    "flat sheet",
    "flatsheet",
    "pillow",
    "towel",
    "blanket",
    "bath mat",
    "bathrobe",
    "bath robe",
    "mattress protector",
    "duvet insert",
    "duvet inner",
    "down duvet",
)


@dataclass(frozen=True)
class RawItem:
    line_number: str
    item_name: str
    size: str
    specification: str
    quantity: str
    section_note: str
    remarks: str = ""
    excel_row_number: int = 0
    source_cells: dict[str, tuple[str, ...]] = field(default_factory=dict)
    inherited_description: bool = False

    @property
    def combined_text(self) -> str:
        return " ".join(
            part
            for part in (
                self.item_name,
                self.size,
                self.specification,
                self.remarks,
                self.section_note,
            )
            if part
        )


@dataclass(frozen=True)
class BuiltResult:
    result: FinalResult
    item: RawItem
    field_diagnostics: dict[str, FieldDiagnostic]


def is_duvet_cover_row(text: str) -> bool:
    normalized, compact = _normalize_filter_text(text)
    if _contains_keywords(normalized, compact, STRONG_EXCLUDE_KEYWORDS):
        return False
    return _contains_keywords(normalized, compact, INCLUDE_KEYWORDS)


def extract_raw_items(
    table: ParsedTable,
    worksheet: Worksheet | None = None,
) -> list[RawItem]:
    all_items: list[RawItem] = []
    previous_item: RawItem | None = None
    for row in table.data_rows:
        raw_item = _raw_item_from_row(row, table, worksheet)
        raw_item = _inherit_adjacent_description(raw_item, previous_item)
        all_items.append(raw_item)
        previous_item = raw_item
    return [item for item in all_items if _is_duvet_cover_item(item)]


def _is_duvet_cover_item(item: RawItem) -> bool:
    item_normalized, item_compact = _normalize_filter_text(item.item_name)
    if _contains_keywords(item_normalized, item_compact, STRONG_EXCLUDE_KEYWORDS):
        return False
    if _contains_keywords(item_normalized, item_compact, INCLUDE_KEYWORDS):
        return True
    return is_duvet_cover_row(item.combined_text)


def _normalize_filter_text(text: str) -> tuple[str, str]:
    normalized = " ".join(text.casefold().replace("-", " ").split())
    return normalized, normalized.replace(" ", "")


def _contains_keywords(normalized: str, compact: str, keywords: tuple[str, ...]) -> bool:
    for keyword in keywords:
        normalized_keyword = " ".join(keyword.casefold().replace("-", " ").split())
        if normalized_keyword in normalized or normalized_keyword.replace(" ", "") in compact:
            return True
    return False


def build_final_results(
    table: ParsedTable,
    metadata: OrderMetadata,
    worksheet: Worksheet | None = None,
) -> list[FinalResult]:
    return [
        built.result
        for built in build_final_results_with_diagnostics(table, metadata, worksheet)
    ]


def build_final_results_with_diagnostics(
    table: ParsedTable,
    metadata: OrderMetadata,
    worksheet: Worksheet | None = None,
) -> list[BuiltResult]:
    built_results: list[BuiltResult] = []
    for item in extract_raw_items(table, worksheet):
        text = item.combined_text
        record = {
            "客户": metadata.customer,
            "币种": metadata.currency,
            "业务员": metadata.salesperson,
            "表头备注": metadata.header_note,
            "行号": item.line_number,
            "物料编码": "",
            "物料名称": f"{metadata.customer} 被套" if metadata.customer else "被套",
            "规格": normalize_size(item.size, item.specification),
            "颜色": normalize_color(text),
            "面料": normalize_fabric(text),
            "面料-涤棉成分": normalize_component(text),
            "款式": normalize_style(text),
            "加标方式": normalize_labeling(text),
            "尺寸类型": normalize_size_type(text),
            "数量": normalize_quantity(item.quantity),
            "行备注": normalize_row_note(text),
            "计划发货日期": metadata.planned_ship_date,
            "包装方式": metadata.packaging,
            "是否绣花": normalize_yes_no_embroidery(text),
            "相似分数": 0.0,
        }
        result = FinalResult.from_mapping(record)
        diagnostics = _build_field_diagnostics(result, item, metadata, table.sheet_title)
        built_results.append(
            BuiltResult(
                result=result,
                item=item,
                field_diagnostics=diagnostics,
            )
        )
    return built_results


def _raw_item_from_row(
    row: ParsedRow,
    table: ParsedTable,
    worksheet: Worksheet | None,
) -> RawItem:
    item_name, item_index = _field_value(
        row,
        table.headers,
        ("item", "items", "product", "品名"),
        fallback_index=1,
    )
    if not item_name:
        item_name, item_index = _field_value(
            row,
            table.headers,
            ("description", "描述"),
            fallback_index=1,
        )
    size, size_index = _field_value(
        row,
        table.headers,
        (
            "after wash size",
            "delivery size",
            "dimension",
            "size",
            "w*l",
            "w x l",
            "wxl",
            "尺寸",
            "洗涤尺寸",
            "交货尺寸",
        ),
        fallback_index=2,
    )
    specification, specification_index = _field_value(
        row,
        table.headers,
        ("specification", "工艺", "fabric", "description", "描述"),
        fallback_index=3,
    )
    quantity, quantity_index = _field_value(
        row,
        table.headers,
        ("total qty", "qty", "quantity", "数量"),
        fallback_index=4,
    )
    remarks, remarks_index = _field_value(
        row,
        table.headers,
        ("remarks", "remark", "notes", "note", "备注"),
        fallback_index=None,
    )
    line_number, line_index = _field_value(
        row,
        table.headers,
        ("no.", "no", "number", "line", "序号"),
        fallback_index=0,
    )

    source_cells = {
        "line_number": _source_for(row, line_index),
        "item_name": _source_for(row, item_index),
        "size": _source_for(row, size_index),
        "specification": _source_for(row, specification_index),
        "quantity": _source_for(row, quantity_index),
        "remarks": _source_for(row, remarks_index),
    }
    inherited = False
    if worksheet is not None and specification_index is not None and specification:
        source, inherited = _merged_source(
            worksheet,
            row.excel_row_number,
            specification_index + 1,
        )
        if source:
            source_cells["specification"] = (source,)

    return RawItem(
        line_number=line_number,
        item_name=item_name,
        size=size,
        specification=specification,
        quantity=quantity,
        section_note=row.section_note,
        remarks=remarks,
        excel_row_number=row.excel_row_number,
        source_cells=source_cells,
        inherited_description=inherited,
    )


def _field_value(
    row: ParsedRow,
    headers: list[str],
    needles: tuple[str, ...],
    *,
    fallback_index: int | None,
) -> tuple[str, int | None]:
    normalized_headers = [_normalize_header(header) for header in headers]
    for needle in needles:
        for index, header in enumerate(normalized_headers):
            if needle in header and index < len(row.raw_values) and row.raw_values[index]:
                return row.raw_values[index], index
    if fallback_index is not None and fallback_index < len(row.raw_values):
        return row.raw_values[fallback_index], fallback_index
    return "", None


def _normalize_header(header: str) -> str:
    return " ".join(header.casefold().replace("\n", " ").split())


def _source_for(row: ParsedRow, column_index: int | None) -> tuple[str, ...]:
    if column_index is None:
        return ()
    return (f"{get_column_letter(column_index + 1)}{row.excel_row_number}",)


def _merged_source(
    worksheet: Worksheet,
    row_number: int,
    column_number: int,
) -> tuple[str, bool]:
    cell = worksheet.cell(row_number, column_number)
    if not isinstance(cell, MergedCell):
        return cell.coordinate, False
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).coordinate, True
    return cell.coordinate, True


def _inherit_adjacent_description(
    current: RawItem,
    previous: RawItem | None,
) -> RawItem:
    if current.specification or previous is None:
        return current
    if current.section_note:
        return current
    if current.excel_row_number != previous.excel_row_number + 1:
        return current
    if not is_duvet_cover_row(previous.specification):
        return current
    return RawItem(
        line_number=current.line_number,
        item_name=current.item_name,
        size=current.size,
        specification=previous.specification,
        quantity=current.quantity,
        section_note=current.section_note,
        remarks=current.remarks,
        excel_row_number=current.excel_row_number,
        source_cells={
            **current.source_cells,
            "specification": previous.source_cells.get("specification", ()),
        },
        inherited_description=True,
    )


def _build_field_diagnostics(
    result: FinalResult,
    item: RawItem,
    metadata: OrderMetadata,
    sheet_name: str,
) -> dict[str, FieldDiagnostic]:
    text = item.combined_text
    item_source = SourceEvidence(
        sheet=sheet_name,
        cells=_all_item_cells(item),
        region="item",
    )
    diagnostics: dict[str, FieldDiagnostic] = {}
    for field_name in ("客户", "币种", "业务员", "表头备注"):
        diagnostics[field_name] = metadata.field_diagnostics[field_name]

    diagnostics["行号"] = _field(
        result,
        "行号",
        EXTRACTED,
        _item_evidence(item, sheet_name, "line_number"),
        "item.line_number",
        "从商品明细行提取",
    )
    diagnostics["物料编码"] = _field(
        result,
        "物料编码",
        NOT_IMPLEMENTED,
        SourceEvidence(sheet=sheet_name, region="matching"),
        "material_code.not_implemented",
        "物料编码匹配属于后续阶段",
    )
    diagnostics["物料名称"] = _field(
        result,
        "物料名称",
        DERIVED,
        SourceEvidence(
            sheet=sheet_name,
            cells=tuple(
                dict.fromkeys(
                    metadata.field_diagnostics["客户"].source.cells
                    + item.source_cells.get("item_name", ())
                )
            ),
            region="derived",
        ),
        "material_name.customer_cover",
        "由已确定客户与被套品类生成",
    )
    size_status = NORMALIZED if result.values["规格"] else (
        UNRECOGNIZED if item.size else SOURCE_NOT_PROVIDED
    )
    diagnostics["规格"] = _field(
        result,
        "规格",
        size_status,
        _item_evidence(item, sheet_name, "size"),
        "size.length_width_cm",
        "规格已统一为长*宽厘米" if result.values["规格"] else "未能稳定识别规格",
    )
    color_explicit = _has_explicit_color(text)
    diagnostics["颜色"] = _field(
        result,
        "颜色",
        NORMALIZED if color_explicit else DEFAULTED,
        item_source if color_explicit else SourceEvidence(sheet=sheet_name, region="default"),
        "color.standard_or_default",
        "颜色已标准化" if color_explicit else "源文件未写颜色，使用漂白色默认值",
    )

    fabric_value = str(result.values["面料"])
    fabric_status = NORMALIZED if fabric_value else (
        UNRECOGNIZED if _has_fabric_signal(text) else SOURCE_NOT_PROVIDED
    )
    diagnostics["面料"] = _field(
        result,
        "面料",
        fabric_status,
        _item_evidence(item, sheet_name, "specification"),
        "fabric.standard",
        _fabric_message(fabric_status, item.inherited_description),
    )

    component_value = str(result.values["面料-涤棉成分"])
    component_status = NORMALIZED if component_value else (
        UNRECOGNIZED if _has_component_signal(text) else SOURCE_NOT_PROVIDED
    )
    diagnostics["面料-涤棉成分"] = _field(
        result,
        "面料-涤棉成分",
        component_status,
        _item_evidence(item, sheet_name, "specification"),
        "component.standard",
        "涤棉成分已标准化"
        if component_value
        else "存在成分描述但当前规则无法稳定识别"
        if component_status == UNRECOGNIZED
        else "源文件未提供明确涤棉成分",
    )

    style_value = str(result.values["款式"])
    style_status = NORMALIZED if style_value else (
        UNRECOGNIZED if _has_style_signal(text) else SOURCE_NOT_PROVIDED
    )
    diagnostics["款式"] = _field(
        result,
        "款式",
        style_status,
        item_source,
        "style.standard",
        "款式已按结构描述标准化"
        if style_value
        else "存在局部款式信息但无法映射完整标准款式"
        if style_status == UNRECOGNIZED
        else "源文件未提供款式结构信息",
    )

    labeling_explicit = bool(re.search(r"\b(?:label|tag)\b|唛标|商标", text, re.IGNORECASE))
    diagnostics["加标方式"] = _field(
        result,
        "加标方式",
        NORMALIZED if labeling_explicit else DEFAULTED,
        item_source if labeling_explicit else SourceEvidence(sheet=sheet_name, region="default"),
        "labeling.customer",
        "加标方式已标准化" if labeling_explicit else "使用客标默认值",
    )
    size_type_explicit = bool(
        re.search(r"before\s+wash|delivery\s+size|after\s*wash(?:ing)?|洗涤尺寸|交货尺寸", text, re.IGNORECASE)
    )
    diagnostics["尺寸类型"] = _field(
        result,
        "尺寸类型",
        NORMALIZED if size_type_explicit else DEFAULTED,
        item_source if size_type_explicit else SourceEvidence(sheet=sheet_name, region="default"),
        "size_type.standard_or_default",
        "尺寸类型已标准化" if size_type_explicit else "使用洗涤尺寸默认值",
    )
    diagnostics["数量"] = _field(
        result,
        "数量",
        EXTRACTED,
        _item_evidence(item, sheet_name, "quantity"),
        "item.quantity",
        "从商品明细行提取数量",
    )
    row_note_value = str(result.values["行备注"])
    diagnostics["行备注"] = _field(
        result,
        "行备注",
        DERIVED if row_note_value else SOURCE_NOT_PROVIDED,
        item_source,
        "row_note.from_item_description",
        "由已确定工艺描述生成" if row_note_value else "源文件未提供需写入行备注的信息",
    )
    diagnostics["计划发货日期"] = metadata.field_diagnostics["计划发货日期"]
    diagnostics["包装方式"] = metadata.field_diagnostics["包装方式"]

    embroidery_explicit = bool(
        re.search(r"\b(?:no|without)?\s*embroider(?:y|ed|ing)\b|绣花|刺绣", text, re.IGNORECASE)
    )
    diagnostics["是否绣花"] = _field(
        result,
        "是否绣花",
        NORMALIZED if embroidery_explicit else DEFAULTED,
        item_source if embroidery_explicit else SourceEvidence(sheet=sheet_name, region="default"),
        "embroidery.standard_or_default",
        "绣花信息已标准化" if embroidery_explicit else "无明确绣花信息，按批准规则使用N",
    )
    diagnostics["相似分数"] = _field(
        result,
        "相似分数",
        NOT_IMPLEMENTED,
        SourceEvidence(sheet=sheet_name, region="matching"),
        "similarity.not_implemented",
        "相似度匹配属于后续阶段",
    )
    return {field_name: diagnostics[field_name] for field_name in FINAL_FIELD_NAMES}


def _field(
    result: FinalResult,
    field_name: str,
    status: str,
    source: SourceEvidence,
    rule: str,
    message: str,
) -> FieldDiagnostic:
    return FieldDiagnostic(
        value=result.values[field_name],
        status=status,
        source=source,
        rule=rule,
        message=message,
    )


def _item_evidence(item: RawItem, sheet_name: str, key: str) -> SourceEvidence:
    return SourceEvidence(
        sheet=sheet_name,
        cells=item.source_cells.get(key, ()),
        region="item",
        label=key,
    )


def _all_item_cells(item: RawItem) -> tuple[str, ...]:
    return tuple(
        dict.fromkeys(
            cell
            for cells in item.source_cells.values()
            for cell in cells
        )
    )


def _has_explicit_color(text: str) -> bool:
    return bool(re.search(r"\bwhite\b|\bgrey\b|\bgray\b|\bblue\b|\bred\b|\byellow\b|白色|灰色|蓝色|红色|黄色", text, re.IGNORECASE))


def _has_fabric_signal(text: str) -> bool:
    return bool(
        re.search(
            r"cotton|coton|poly(?:ester)?|sateen|satin|twill|stripe|plain\s+weave|"
            r"thread\s+count|\b\d{3}\s*TC\b|贡缎|缎条|斜纹|平布",
            text,
            re.IGNORECASE,
        )
    )


def _has_component_signal(text: str) -> bool:
    return bool(re.search(r"\d{1,3}\s*%|cotton|coton|poly(?:ester)?|涤棉|全棉", text, re.IGNORECASE))


def _has_style_signal(text: str) -> bool:
    return bool(
        re.search(
            r"\bbag\b|envelope|flap|flange|hem|opening|hand\s+holes?|zip(?:per)?|ties?|"
            r"口袋|信封|飞边|开口|手洞|拉链|系带",
            text,
            re.IGNORECASE,
        )
    )


def _fabric_message(status: str, inherited: bool) -> str:
    if status == NORMALIZED:
        return "使用同组共享描述并标准化" if inherited else "面料信息已标准化"
    if status == UNRECOGNIZED:
        return "存在面料描述但当前规则无法稳定识别"
    return "源文件未提供明确面料信息"
